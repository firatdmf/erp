"""
Import opening balances for the Ergene Fabric book from the per-customer
Excel ledgers.

The money comes from the workbooks; the CRM linkage comes from a reviewed
map CSV. Keeping the two apart means every figure traces back to a source
sheet, and every judgement call about who a sheet belongs to is visible in
version control instead of buried in a fuzzy match.

Balances only — one `opening` movement per account, dated the cutover. The
workbooks do carry line-by-line history, but the date column doubles as a
document reference and subtotal rows sit between the real ones, so
replaying them is a separate job. Two things to know if that job ever
happens: the debit/credit pair is positional, not labelled (the
second-to-last column always increases the balance whatever the header
says, and DUCILIE CIPRIAN ROMANYA is the one sheet where the columns are
genuinely swapped), and `DEVİR` rows are themselves opening balances
carried in from an older book.

Every movement is tagged with a batch reference so the run can be undone:

    CariMovement.objects.filter(reference="ERGENE-OB-20260831").delete()

Dry run by default; pass --apply to commit.
"""
import csv
import datetime
import re
import unicodedata
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from accounting.models import Book, CurrencyCategory
from accounting.models_accounts import CariAccount, CariMovement, CariSettings
from crm.models import Company, Contact

# The sheets are named for the customer, and for the foreign ones that name
# carries the country ("GÜRHAN ROMANYA", "WİNN WAŞOVA"). Worth keeping — it is
# the only location data these workbooks hold. Ambiguous markers are left out
# on purpose: ARAP names a language, and IRAKLI is a Georgian given name, not
# a reference to Iraq.
COUNTRY_BY_TOKEN = {
    "ROMANYA": "Romania",
    "RUSYA": "Russia",
    "MOSKOVA": "Russia",
    "OZBEKISTAN": "Uzbekistan",
    "AFGAN": "Afghanistan",
    "IRAN": "Iran",
    "IRANLI": "Iran",
    "VARSOVA": "Poland",
    "WASOVA": "Poland",
    "KIRGIZISTAN": "Kyrgyzstan",
}

DEFAULT_MAP = Path(__file__).resolve().parents[2] / "data" / "ergene_opening_map.csv"
DEFAULT_SOURCE = Path.home() / "Desktop" / "Müşteri"


class _DryRun(Exception):
    """Raised at the end of a dry run to roll the transaction back."""


def nfc(value):
    """Compose a string the way the rest of the world writes it.

    macOS hands back DECOMPOSED filenames — "HACİM" arrives as H A C I
    plus a combining dot — and a name written to the database that way is
    unsearchable, because every browser sends the composed form and
    Postgres compares the bytes. Everything read from a filename or from
    the map CSV goes through this before it is stored.
    """
    return unicodedata.normalize("NFC", value or "")


def fold(value):
    """Case- and diacritic-insensitive key.

    macOS hands back NFD-decomposed filenames, so `İ` arrives as `I` plus a
    combining dot and never compares equal to the NFC literal in a CSV.
    Decompose, drop the combining marks, then fold the dotless ı by hand
    since it has no decomposition of its own.
    """
    text = unicodedata.normalize("NFD", value or "")
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.replace("ı", "i").replace("İ", "I").upper()
    return " ".join(re.sub(r"[^A-Z0-9 ]", " ", text).split())


def derive_country(name):
    """Country implied by the customer name, or '' when it says nothing."""
    for token in fold(name).split():
        if token in COUNTRY_BY_TOKEN:
            return COUNTRY_BY_TOKEN[token]
    return ""


def read_sheet_balance(path):
    """Return the closing balance of a customer workbook.

    This is the last populated cell of the running-balance column, not the
    `Güncel Borcu` cell in row 1. That cell is a formula pinned to a fixed
    row (`=I998`), so it goes stale the moment a sheet grows past it —
    HACİM reads 862k there against a real 1,004k, and HMZ ROMA reads 144k
    against a zeroed-out account.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        rows = list(workbook.worksheets[0].iter_rows(values_only=True))
    finally:
        workbook.close()

    header_at = None
    for index, row in enumerate(rows[:10]):
        if row and isinstance(row[0], str) and row[0].strip() == "Tarih":
            header_at = index
            break
    if header_at is None:
        raise CommandError(f"{path.name}: no 'Tarih' header row in the first 10 rows")

    # Most sheets label the column 'Bakiye'; the older ones leave it as
    # Excel's own 'Sütun1'.
    header = [fold(str(c)) if c is not None else "" for c in rows[header_at]]
    for label in ("BAKIYE", "SUTUN1"):
        if label in header:
            column = header.index(label)
            break
    else:
        raise CommandError(f"{path.name}: no balance column in {header!r}")

    for row in reversed(rows[header_at + 1:]):
        if column < len(row) and isinstance(row[column], (int, float)):
            return Decimal(str(round(float(row[column]), 2)))
    return Decimal("0.00")


class Command(BaseCommand):
    help = "Import Ergene Fabric opening balances from the per-customer Excel ledgers."

    def add_arguments(self, parser):
        parser.add_argument("--source", default=str(DEFAULT_SOURCE),
                            help="Folder holding the per-customer .xlsx ledgers")
        parser.add_argument("--map", dest="map_path", default=str(DEFAULT_MAP),
                            help="Reviewed CRM linkage map")
        parser.add_argument("--book", default="Ergene Fabric", help="Target book name")
        parser.add_argument("--date", default=None,
                            help="Cutover date, YYYY-MM-DD (default: today)")
        parser.add_argument("--prefix", default="ACC",
                            help="Account code prefix for newly minted codes")
        parser.add_argument("--only", nargs="+", default=None, metavar="KEY",
                            help="Restrict the run to these map keys. Balances are "
                                 "re-read from the workbooks every time, so without "
                                 "this a later run also trues up any sheet that has "
                                 "moved since — useful, but rarely what you meant "
                                 "when posting one account.")
        parser.add_argument("--apply", action="store_true",
                            help="Commit. Without it the run is rolled back.")

    def handle(self, *args, **options):
        source = Path(options["source"])
        if not source.is_dir():
            raise CommandError(f"Source folder not found: {source}")
        map_path = Path(options["map_path"])
        if not map_path.is_file():
            raise CommandError(f"Map file not found: {map_path}")

        cutover = (
            datetime.date.fromisoformat(options["date"]) if options["date"]
            else timezone.localdate()
        )
        batch_ref = f"ERGENE-OB-{cutover:%Y%m%d}"

        try:
            book = Book.objects.get(name=options["book"])
        except Book.DoesNotExist:
            raise CommandError(f"No book named {options['book']!r}")
        try:
            usd = CurrencyCategory.objects.get(code="USD")
        except CurrencyCategory.DoesNotExist:
            raise CommandError("No USD currency category")

        # Index the workbooks by folded name so NFD filenames still match.
        workbooks = {fold(p.stem): p for p in source.glob("*.xlsx")
                     if not p.name.startswith("~$")}
        rows = [{k: nfc(v) for k, v in r.items()}
                for r in csv.DictReader(map_path.open(encoding="utf-8"))]
        if options["only"]:
            wanted = {fold(k) for k in options["only"]}
            rows = [r for r in rows if fold(r["key"]) in wanted]
            missing = wanted - {fold(r["key"]) for r in rows}
            if missing:
                raise CommandError("No map row for: " + ", ".join(sorted(missing)))

        self.stdout.write(
            f"book={book.name} (id {book.pk})  cutover={cutover}  ref={batch_ref}  "
            f"{'APPLY' if options['apply'] else 'DRY RUN'}"
        )

        # Parse the sheets before opening the transaction. Reading 60-odd
        # workbooks takes long enough that doing it inside would hold a write
        # transaction open on the live database for no reason.
        balances, missing = {}, []
        for row in rows:
            if row["action"] == "skip":
                continue
            path = workbooks.get(row["key"])
            if path is None:
                missing.append(row["key"])
                continue
            balances[row["key"]] = read_sheet_balance(path)
        if missing:
            raise CommandError(
                "No workbook found for: " + ", ".join(sorted(missing))
            )
        self.stdout.write(f"read {len(balances)} workbooks\n")

        try:
            with transaction.atomic():
                self._run(rows, balances, book, usd, cutover, batch_ref,
                          options["prefix"], options["apply"])
                if not options["apply"]:
                    raise _DryRun
        except _DryRun:
            self.stdout.write(self.style.WARNING(
                "\nDry run — rolled back. Re-run with --apply to commit."))

    def _run(self, rows, balances, book, usd, cutover, batch_ref, prefix, apply):
        settings_obj = CariSettings.for_book(book)
        if settings_obj.cari_code_prefix != prefix:
            self.stdout.write(f"code prefix {settings_obj.cari_code_prefix!r} -> {prefix!r}")
            settings_obj.cari_code_prefix = prefix
            settings_obj.save(update_fields=["cari_code_prefix"])

        created = posted = skipped = 0
        total = Decimal("0.00")
        problems = []

        for row in rows:
            key, action = row["key"], row["action"]
            if action == "skip":
                skipped += 1
                continue

            balance = balances[key]
            link = self._resolve_link(row, problems)
            if link is None:
                continue
            field, target = link

            lookup = {field: target} if field else {"name": row["account_name"]}
            account = CariAccount.objects.filter(book=book, **lookup).first()
            if account is None and action == "movement_only":
                problems.append(f"{key}: map says movement_only but no account is linked")
                continue

            if account is None:
                account = CariAccount(
                    book=book, name=row["account_name"],
                    type=row.get("account_type") or "customer",
                    default_currency=usd, opening_balance=balance,
                    opening_balance_date=cutover,
                    notes=row.get("note") or "",
                    **({field: target} if field else {}),
                )
                account.save()          # blank code -> minted by pre_save signal
                created += 1
                verb = "new"
            else:
                verb = f"exists {account.code}"

            # Post whatever it takes to land on the sheet's closing balance.
            # For an account already carrying live movements (Eurofirany has
            # an order on it) that is a delta, not the full figure.
            delta = balance - account.cached_balance
            if CariMovement.objects.filter(cari=account, reference=batch_ref).exists():
                self.stdout.write(f"  = {row['account_name'][:28]:29} already posted, skipping")
                continue
            if delta == 0:
                self.stdout.write(
                    f"  · {row['account_name'][:28]:29} {balance:>12,.2f}  "
                    f"{account.code}, no movement needed"
                )
                continue

            CariMovement(
                cari=account, book=book, date=cutover, amount=delta, currency=usd,
                movement_type="opening",
                description=row.get("description") or "Opening balance (Ergene ledger migration)",
                reference=batch_ref,
            ).save()                    # never bulk_create: save() fills amount_base
            posted += 1
            total += delta
            where = f"[{field} #{target.pk}]" if field else "[no CRM link]"
            self.stdout.write(
                f"  + {row['account_name'][:28]:29} {delta:>12,.2f}  {verb} -> "
                f"{account.code} {where}"
            )

        self.stdout.write("")
        self.stdout.write(
            f"accounts created {created} | movements posted {posted} | "
            f"skipped by map {skipped} | total {total:,.2f} USD"
        )
        if problems:
            self.stdout.write(self.style.ERROR(f"\n{len(problems)} problem(s):"))
            for line in problems:
                self.stdout.write(self.style.ERROR(f"  {line}"))
            raise CommandError("Refusing to continue with unresolved rows.")

    def _resolve_link(self, row, problems):
        """Return (field_name, crm_object) for the map row, or None.

        link_type "none" returns (None, None): an inter-company position is
        not a client relationship, and inventing a CRM record for the other
        half of your own business would put a company in the customer list
        that nobody is allowed to sell to.
        """
        if (row.get("link_type") or "").strip() == "none":
            return None, None
        model, field = (
            (Company, "company") if row["link_type"] == "company" else (Contact, "contact")
        )
        name = row["account_name"]
        country = derive_country(name)

        crm_id = (row.get("crm_id") or "").strip()
        if crm_id:
            obj = model.objects.filter(pk=crm_id).first()
            if obj is None:
                problems.append(f"{row['key']}: {field} #{crm_id} does not exist")
                return None
        else:
            matches = list(model.objects.filter(name=name))
            if len(matches) > 1:
                problems.append(
                    f"{row['key']}: {len(matches)} {field}s already named {name!r} "
                    f"(ids {[m.pk for m in matches]}) — set crm_id in the map"
                )
                return None
            obj = matches[0] if matches else model.objects.create(name=name, country=country)

        # Existing records keep whatever they already say; only a blank is
        # worth filling in from a sheet name.
        if country and not obj.country:
            obj.country = country
            obj.save(update_fields=["country"])
            self.stdout.write(f"    country {field} #{obj.pk} {name[:24]} -> {country}")
        return field, obj
