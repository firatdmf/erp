"""Excel (.xlsx) export of an Invoice — bordered, document-style sheet that
mirrors the printable invoice and carries the full record: issuer, consignee,
line items and totals. VAT / discount columns and total lines appear ONLY
when the invoice actually uses them.
"""
from io import BytesIO

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from erp.xlsx_utils import (
    cell, merge, merge_border, section, kv_full, kv_pair,
    GRID, RULE, FILL_HEAD, RIGHT, LEFT, TOP,
    F_TITLE, F_SUB, F_DOCNO, F_HEAD, F_VAL, F_VALB, F_TOTAL,
)
from .models import Invoice


def _dec(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _date(v):
    try:
        return v.strftime("%d %b %Y") if v else "—"
    except Exception:
        return "—"


def build_invoice_workbook(invoice):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    cari = invoice.cari
    ccode = invoice.currency.code if (invoice.currency_id and invoice.currency) else ""
    money = (f'#,##0.00" {ccode}"' if ccode else "#,##0.00")

    def s(name):
        return getattr(settings, name, "") or ""

    # Issuer = per-invoice snapshot, else brand defaults.
    iname = invoice.issuer_name or s("BRAND_NAME") or "Nejum"
    iaddr = invoice.issuer_address or s("BRAND_ADDRESS")
    iphone = invoice.issuer_phone or s("BRAND_PHONE")
    ifax = invoice.issuer_fax or s("BRAND_FAX")
    iemail = invoice.issuer_email or s("BRAND_EMAIL")
    itax_office = invoice.issuer_tax_office or s("BRAND_TAX_OFFICE")
    itax_no = invoice.issuer_tax_number or s("BRAND_TAX_NUMBER")

    # Consignee = per-invoice snapshot, else cari master.
    def cv(attr):
        return getattr(cari, attr, "") if cari else ""
    bname = invoice.bill_to_name or (cari.name if cari else "—")
    baddr = invoice.bill_to_address or cv("billing_address")
    bcity = invoice.bill_to_city or cv("billing_city")
    bcountry = invoice.bill_to_country or cv("billing_country")
    bphone = invoice.bill_to_phone or cv("phone")
    bemail = invoice.bill_to_email or cv("email")
    btax_office = invoice.bill_to_tax_office or cv("tax_office")
    btax_no = invoice.bill_to_tax_number or cv("tax_number")

    # Item columns adapt: VAT / Disc only shown when at least one line uses them.
    items = list(invoice.items.all().select_related("product", "variant"))
    has_disc = any((it.discount_rate or 0) for it in items)
    has_vat = any((it.tax_rate or 0) for it in items)
    icols = [("Description", 34, "desc"), ("SKU", 18, "sku"),
             ("Qty", 10, "qty"), ("Unit Price", 14, "unit")]
    if has_disc:
        icols.append(("Disc %", 9, "disc"))
    if has_vat:
        icols.append(("VAT %", 9, "vat"))
    icols.append(("Line Total", 16, "total"))
    NCOLS = len(icols)
    mid = (NCOLS + 1) // 2

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.sheet_view.showGridLines = False
    for i, (h, w, k) in enumerate(icols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    title = "PROFORMA INVOICE" if invoice.type == "proforma" else "INVOICE"

    # ── Header ──
    r = 1
    cell(ws, r, 1, iname.upper(), font=F_TITLE); merge(ws, r, 1, mid)
    cell(ws, r, mid + 1, title, font=F_DOCNO, align=RIGHT); merge(ws, r, mid + 1, NCOLS)
    r += 1
    cell(ws, r, 1, "Tax Invoice", font=F_SUB); merge(ws, r, 1, mid)
    cell(ws, r, mid + 1, f"No: {invoice.number or '—'}", font=F_VALB, align=RIGHT); merge(ws, r, mid + 1, NCOLS)
    for c in range(1, NCOLS + 1):
        ws.cell(r, c).border = RULE
    r += 2

    # ── Invoice details ──
    r = section(ws, r, "INVOICE DETAILS", NCOLS)
    r = kv_pair(ws, r, "Invoice No", invoice.number or "—", "Series", invoice.series or "—", NCOLS)
    r = kv_pair(ws, r, "Type", invoice.get_type_display(), "Status", invoice.get_status_display(), NCOLS)
    r = kv_pair(ws, r, "Date", _date(invoice.date), "Due Date", _date(invoice.due_date), NCOLS)
    r = kv_pair(ws, r, "Delivery Date", _date(invoice.delivery_date), "Currency", ccode or "—", NCOLS)
    linked = ""
    if invoice.order_id and invoice.order:
        linked = invoice.order.order_number or f"#{invoice.order_id}"
    r = kv_pair(ws, r, "Exchange Rate", str(invoice.exchange_rate), "Linked Order", linked or "—", NCOLS)
    if invoice.earsiv_uuid or invoice.earsiv_status:
        r = kv_pair(ws, r, "e-Arşiv UUID", invoice.earsiv_uuid or "—", "e-Arşiv Status", invoice.earsiv_status or "—", NCOLS)
    r += 1

    # ── Issuer ──
    r = section(ws, r, "ISSUER", NCOLS)
    r = kv_full(ws, r, "Name", iname, NCOLS, bold_value=True)
    if iaddr:
        r = kv_full(ws, r, "Address", iaddr, NCOLS)
    if iphone:
        r = kv_full(ws, r, "Phone", iphone, NCOLS)
    if ifax:
        r = kv_full(ws, r, "Fax", ifax, NCOLS)
    if iemail:
        r = kv_full(ws, r, "Email", iemail, NCOLS)
    if itax_office or itax_no:
        r = kv_full(ws, r, "Tax Office / No", f"{itax_office}  ·  {itax_no}".strip(" ·"), NCOLS)
    r += 1

    # ── Bill to ──
    r = section(ws, r, "BILL TO", NCOLS)
    r = kv_full(ws, r, "Name", bname, NCOLS, bold_value=True)
    if baddr:
        r = kv_full(ws, r, "Address", baddr, NCOLS)
    loc = ", ".join([x for x in [bcity, bcountry] if x])
    if loc:
        r = kv_full(ws, r, "City / Country", loc, NCOLS)
    if bphone:
        r = kv_full(ws, r, "Phone", bphone, NCOLS)
    if bemail:
        r = kv_full(ws, r, "Email", bemail, NCOLS)
    if btax_office or btax_no:
        r = kv_full(ws, r, "Tax Office / No", f"{btax_office}  ·  {btax_no}".strip(" ·"), NCOLS)
    r += 1

    # ── Items ──
    r = section(ws, r, f"ITEMS ({len(items)})", NCOLS)
    for i, (h, w, k) in enumerate(icols, 1):
        cell(ws, r, i, h, font=F_HEAD, fill=FILL_HEAD, border=GRID,
             align=(LEFT if k in ("desc", "sku") else RIGHT))
    r += 1

    def _val(it, kind, sku):
        if kind == "desc":
            return (it.description or (str(it.product) if it.product else "—")), None, TOP
        if kind == "sku":
            return (sku or "—"), None, LEFT
        if kind == "qty":
            return _dec(it.quantity), "#,##0.00", RIGHT
        if kind == "unit":
            return _dec(it.unit_price), money, RIGHT
        if kind == "disc":
            return _dec(it.discount_rate), '0.##"%"', RIGHT
        if kind == "vat":
            return _dec(it.tax_rate), '0.##"%"', RIGHT
        return _dec(it.total), money, RIGHT  # total

    for it in items:
        if it.variant_id and it.variant:
            sku = getattr(it.variant, "variant_sku", "") or ""
        elif it.product_id and it.product:
            sku = getattr(it.product, "sku", "") or ""
        else:
            sku = ""
        for i, (h, w, k) in enumerate(icols, 1):
            v, fmt, al = _val(it, k, sku)
            cell(ws, r, i, v, font=F_VAL, border=GRID, align=al, fmt=fmt)
        r += 1

    # ── Totals (only the lines that actually apply) ──
    r += 1
    lc1, lc2, vc = NCOLS - 2, NCOLS - 1, NCOLS
    rows = []
    if has_disc or has_vat:
        rows.append(("Subtotal", invoice.subtotal, False))
    if invoice.discount_amount:
        rows.append(("Discount", invoice.discount_amount, False))
    if invoice.tax_amount:
        rows.append(("VAT", invoice.tax_amount, False))
    if invoice.other_charges:
        rows.append(("Other charges", invoice.other_charges, False))
    rows.append(("TOTAL", invoice.total, True))
    if invoice.paid_amount:
        rows.append(("Paid", invoice.paid_amount, False))
        rows.append(("Balance", invoice.balance, True))
    for lbl, val, strong in rows:
        cell(ws, r, lc1, lbl, font=(F_TOTAL if strong else F_VALB), border=GRID, align=RIGHT)
        merge(ws, r, lc1, lc2)
        merge_border(ws, r, lc1, lc2, GRID)
        cell(ws, r, vc, _dec(val), font=(F_TOTAL if strong else F_VALB), border=GRID, align=RIGHT, fmt=money)
        r += 1

    # ── Notes ──
    if invoice.notes:
        r += 1
        r = section(ws, r, "NOTES", NCOLS)
        cell(ws, r, 1, invoice.notes, font=F_VAL, border=GRID, align=TOP)
        merge(ws, r, 1, NCOLS)
        merge_border(ws, r, 1, NCOLS, GRID)
        ws.row_dimensions[r].height = 46

    return wb


@login_required
def invoice_excel(request, pk):
    """Download the invoice as an .xlsx file."""
    invoice = get_object_or_404(Invoice, pk=pk)
    wb = build_invoice_workbook(invoice)
    buf = BytesIO()
    wb.save(buf)
    label = f"{invoice.series}-{invoice.number}" if invoice.number else f"invoice-{invoice.pk}"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{label}.xlsx"'
    return resp
