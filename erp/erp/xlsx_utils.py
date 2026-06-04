"""Shared helpers for bordered, document-style .xlsx exports (order &
invoice) that resemble the printable PDFs — full cell borders ("çizgili"),
shaded label cells, a coloured items header and a totals block."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK = "FF111111"
MUT = "FF6B7280"
HEADBG = "FF0F766E"
LBLBG = "FFEFF3F6"

_thin = Side(style="thin", color="FFB9C2CC")
_med = Side(style="medium", color="FF111111")
GRID = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
RULE = Border(bottom=_med)

F_TITLE = Font(size=20, bold=True, color="FF944F05")   # brand wordmark
F_SUB = Font(size=9, color=MUT)
F_DOCNO = Font(size=14, bold=True, color=INK)
F_SECT = Font(size=10, bold=True, color=INK)
F_LBL = Font(size=9, bold=True, color="FF374151")
F_VAL = Font(size=10, color=INK)
F_VALB = Font(size=10, bold=True, color=INK)
F_HEAD = Font(size=9, bold=True, color="FFFFFFFF")
F_TOTAL = Font(size=11, bold=True, color=INK)

FILL_HEAD = PatternFill("solid", fgColor=HEADBG)
FILL_LBL = PatternFill("solid", fgColor=LBLBG)

RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def cell(ws, r, c, value="", font=None, align=None, border=None, fill=None, fmt=None):
    cl = ws.cell(r, c, value if value is not None else "")
    if font:
        cl.font = font
    cl.alignment = align or LEFT
    if border:
        cl.border = border
    if fill:
        cl.fill = fill
    if fmt:
        cl.number_format = fmt
    return cl


def merge(ws, r, c1, c2):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)


def merge_border(ws, r, c1, c2, border):
    """Apply a border across every cell of a horizontal merged range
    (openpyxl only borders the top-left cell otherwise)."""
    for c in range(c1, c2 + 1):
        ws.cell(r, c).border = border


def section(ws, r, text, ncols):
    """Section header: bold label with a medium rule underneath, full width."""
    c = ws.cell(r, 1, text)
    c.font = F_SECT
    c.alignment = LEFT
    for col in range(1, ncols + 1):
        ws.cell(r, col).border = RULE
    return r + 1


def kv_full(ws, r, label, value, ncols, bold_value=False):
    """One full-width label/value row: shaded label in col A, value spanning
    the rest. Bordered."""
    lc = ws.cell(r, 1, label)
    lc.font = F_LBL
    lc.alignment = LEFT
    lc.fill = FILL_LBL
    lc.border = GRID
    val = value if (value not in (None, "")) else "—"
    vc = ws.cell(r, 2, val)
    vc.font = F_VALB if bold_value else F_VAL
    vc.alignment = LEFT
    merge(ws, r, 2, ncols)
    merge_border(ws, r, 2, ncols, GRID)
    return r + 1


def kv_pair(ws, r, l1, v1, l2, v2, ncols):
    """Two label/value pairs on one row. Layout adapts to column count:
    A=label, B..mid=value, mid+1=label, ...=value."""
    mid = (ncols + 1) // 2
    # left pair: label col 1, value cols 2..mid
    ws.cell(r, 1, l1).font = F_LBL
    ws.cell(r, 1).fill = FILL_LBL
    ws.cell(r, 1).border = GRID
    ws.cell(r, 1).alignment = LEFT
    ws.cell(r, 2, v1 if v1 not in (None, "") else "—").font = F_VAL
    ws.cell(r, 2).alignment = LEFT
    merge(ws, r, 2, mid)
    merge_border(ws, r, 2, mid, GRID)
    # right pair: label col mid+1, value cols mid+2..ncols
    ws.cell(r, mid + 1, l2).font = F_LBL
    ws.cell(r, mid + 1).fill = FILL_LBL
    ws.cell(r, mid + 1).border = GRID
    ws.cell(r, mid + 1).alignment = LEFT
    ws.cell(r, mid + 2, v2 if v2 not in (None, "") else "—").font = F_VAL
    ws.cell(r, mid + 2).alignment = LEFT
    merge(ws, r, mid + 2, ncols)
    merge_border(ws, r, mid + 2, ncols, GRID)
    return r + 1
