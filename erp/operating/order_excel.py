"""Excel (.xlsx) export of an Order — bordered, document-style sheet that
mirrors the printable order PDF and carries the full order record.

Two of them: one order, and several of one customer's orders together
(the sheet OrderPrintCombined prints). The combined workbook is the one
people take away to work on, so it is a table before it is a document:
one row per line item, the rows filtered and the header frozen, and
every figure a number rather than something that only looks like one.

It does NOT carry the order each line came from. The printed sheet says
that in a heading above each group, and this deliberately does not
repeat it — asked for as a list of goods, not a reconciliation.
"""
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

from erp.xlsx_utils import (
    cell, merge, merge_border, section, kv_full, kv_pair,
    GRID, RULE, FILL_HEAD, RIGHT, LEFT, TOP, INK,
    F_TITLE, F_SUB, F_DOCNO, F_HEAD, F_VAL, F_VALB, F_TOTAL,
)
from .models import Order

# The totals row, dressed so the eye lands on it from across the sheet:
# the accountant's rule above and double rule below, a tinted ground, and
# type a size up from the lines it totals. It used to differ from a line
# item by its font weight alone, which is not enough at the foot of
# thirty rows that are themselves half bold.
_rule = Side(style="medium", color=INK)
_double = Side(style="double", color=INK)
BORDER_TOTAL = Border(top=_rule, bottom=_double)
FILL_TOTAL = PatternFill("solid", fgColor="FFF3F6F8")
F_GRAND = Font(size=12, bold=True, color=INK)

NCOLS = 6  # A..F


def _dec(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _dt(v, fmt="%d %b %Y · %H:%M"):
    try:
        return v.strftime(fmt) if v else "—"
    except Exception:
        return "—"


def _currency_symbol(code):
    """The sign a figure in this currency wears — "$", "₺", "€".

    Asked of CurrencyCategory, which is where the ledger keeps it, so a
    workbook and an invoice cannot disagree about what a lira looks
    like. Falls back to the ISO code itself: a column reading "USD 2.45"
    is worse than "$2.45" but better than a blank sign, and it must
    never raise — a spreadsheet has to build on a database that has no
    currency rows at all.
    """
    try:
        from accounting.models import CurrencyCategory
        sym = (CurrencyCategory.objects.filter(code=code)
               .values_list("symbol", flat=True).first() or "").strip()
        if sym:
            return sym
    except Exception:
        pass
    return {"USD": "$", "EUR": "€", "GBP": "£", "TRY": "₺"}.get(code, code)


def _money_format(code):
    """An Excel CURRENCY format: the sign is worn by the number, not
    typed beside it. The cell still holds 2.45 and still sums; only its
    display carries the sign, which is the difference between a figure
    and a label that looks like one."""
    sym = _currency_symbol(code).replace('"', "")
    return f'"{sym}"#,##0.00'


def _break_words(text, every=4):
    """Hard-wrap a name every `every` words.

    Excel wraps on its own, but only where the column happens to run
    out: the same name breaks in a different place on every sheet, and
    one that fits by a character does not break at all — it just runs
    on across the page. A break the reader can predict beats one the
    column width decides.

    Four words, because that is what halves the names this is for.
    "GREK TAŞLI VE İNCİ EKRU İNCİ BEYAZ ZEMİN" is eight of them and
    comes out square.
    """
    words = str(text or "").split()
    if len(words) <= every:
        return str(text or "").strip()
    return "\n".join(" ".join(words[i:i + every])
                     for i in range(0, len(words), every))


def _display_len(cl):
    """How wide a cell READS, which is not how wide its value is.

    A number carries a format: 243.53 in a currency column shows as
    "$243.53", two characters more than the value has. And a value
    broken over lines is only as wide as its longest line, not the sum
    of them.
    """
    v = cl.value
    if v is None:
        return 0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        fmt = cl.number_format or ""
        # Whatever the format sets in quotes is printed literally beside
        # the figure — the currency sign, in practice.
        literal = sum(len(part) for part in fmt.split('"')[1::2])
        return len(f"{v:,.2f}") + literal
    return max((len(line) for line in str(v).split("\n")), default=0)


def fit_columns(ws, first_row, last_row, ncols, header_row=None,
                min_w=6, max_w=34, pad=1, arrow=3):
    """Size each column to the widest thing actually in it.

    Fixed widths were guesswork, and guessed generously: columns holding
    "Fabric" and a pack count were as wide as the room a long name might
    one day need. Measured against the rows themselves there is no such
    slack, and a narrower sheet is one the fit-to-width scaling has less
    to shrink.

    `pad` is one character, which is all a cell needs: Excel already
    insets its text a couple of pixels from the border. `arrow` is the
    three the header row gets on top — the filter's dropdown sits
    INSIDE the cell and covers the end of the label otherwise, which is
    what several of these columns were doing before they were measured.
    """
    from openpyxl.utils import get_column_letter

    for c in range(1, ncols + 1):
        widest = 0
        for r in range(first_row, last_row + 1):
            n = _display_len(ws.cell(r, c))
            if r == header_row and n:
                n += arrow
            widest = max(widest, n)
        ws.column_dimensions[get_column_letter(c)].width = \
            min(max(widest + pad, min_w), max_w)


def fit_rows(ws, first_row, last_row, ncols, cap=4):
    """Give each row the height its longest cell needs.

    Runs AFTER fit_columns, because how many lines a value takes is a
    question about the column it ended up in. openpyxl writes no row
    heights at all, and Excel then leaves every row one line tall
    whatever its cells say about wrapping.
    """
    from openpyxl.utils import get_column_letter

    for r in range(first_row, last_row + 1):
        lines = 1
        for c in range(1, ncols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                w = ws.column_dimensions[get_column_letter(c)].width or 10
                lines = max(lines, _wrapped_lines(v, w))
        if lines > 1:
            ws.row_dimensions[r].height = 13.5 * min(lines, cap)


def _wrapped_lines(text, width_chars):
    """How many lines a value takes in a column that wide.

    Rough — it counts characters, not the words they fall into — but it
    only ever has to answer "more than one?", and openpyxl writes no row
    heights, so without an answer Excel leaves the row one line tall and
    a 40-character product name runs off the side of its column instead
    of wrapping inside it.
    """
    if not text:
        return 1
    lines = 0
    for para in str(text).split("\n"):
        lines += max(1, -(-len(para) // max(1, int(width_chars) - 1)))
    return lines


def _firstv(v):
    if isinstance(v, (list, tuple)):
        return (v[0] if v else "") or ""
    return v or ""


def build_order_workbook(order):
    from openpyxl import Workbook

    # The order's own print header wins, then the ledger book's brand
    # name — same precedence as order_print.html, so the two documents
    # for one order sign with the same name.
    from accounting.services_accounts import brand_name_for
    brand = (order.print_header or "").strip() or brand_name_for()
    ccode = (order.original_currency or order.paid_currency or "USD")
    # The same currency FORMAT the combined sheet uses — "$2.45", the
    # sign worn by the number. This carried the ISO code as a suffix
    # ("2.45 USD"), which is still a number underneath but reads as a
    # label and cannot be told apart from one at a glance.
    money = _money_format(ccode)

    wb = Workbook()
    ws = wb.active
    ws.title = "Order"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", (32, 22, 18, 11, 15, 16)):
        ws.column_dimensions[col].width = w

    num = order.order_number or f"#{order.pk}"

    # ── Header ──
    r = 1
    cell(ws, r, 1, brand, font=F_TITLE)
    merge(ws, r, 1, 3)
    cell(ws, r, 4, f"ORDER {num}", font=F_DOCNO, align=RIGHT)
    merge(ws, r, 4, 6)
    # 20pt type in a row sized for 11pt: the wordmark's descenders were
    # cut off by the row beneath it. Row heights are in points.
    ws.row_dimensions[r].height = 28
    r += 1
    cell(ws, r, 1, "Order Confirmation", font=F_SUB)
    merge(ws, r, 1, 3)
    cell(ws, r, 4, _dt(order.order_date) if order.order_date else _dt(order.created_at), font=F_SUB, align=RIGHT)
    merge(ws, r, 4, 6)
    for c in range(1, NCOLS + 1):
        ws.cell(r, c).border = RULE
    r += 2

    # ── Order details ──
    r = section(ws, r, "ORDER DETAILS", NCOLS)
    payment = ""
    if getattr(order, "payment_status", None):
        try:
            payment = order.get_payment_status_display()
        except Exception:
            payment = order.payment_status
    r = kv_pair(ws, r, "Order No", num, "Status", order.get_order_status_display() or order.order_status, NCOLS)
    r = kv_pair(ws, r, "Order Date", _dt(order.order_date) if order.order_date else _dt(order.created_at), "Payment", payment or "—", NCOLS)
    cari_name = order.cari.name if (order.cari_id and order.cari) else "—"
    r = kv_pair(ws, r, "Last Update", _dt(order.updated_at), "Currency", ccode, NCOLS)
    r = kv_full(ws, r, "Linked Account", cari_name, NCOLS)
    r += 1

    # ── Customer ──
    r = section(ws, r, "CUSTOMER", NCOLS)
    if order.contact_id and order.contact:
        ct = order.contact
        r = kv_full(ws, r, "Type", "Contact", NCOLS)
        r = kv_full(ws, r, "Name", ct.name or "—", NCOLS, bold_value=True)
        r = kv_full(ws, r, "Email", _firstv(ct.email), NCOLS)
        r = kv_full(ws, r, "Phone", _firstv(ct.phone), NCOLS)
        if getattr(ct, "address", ""):
            r = kv_full(ws, r, "Address", ct.address, NCOLS)
        if getattr(ct, "company", ""):
            r = kv_full(ws, r, "Company", ct.company, NCOLS)
    elif order.company_id and order.company:
        co = order.company
        r = kv_full(ws, r, "Type", "Company", NCOLS)
        r = kv_full(ws, r, "Name", co.name or "—", NCOLS, bold_value=True)
        r = kv_full(ws, r, "Email", _firstv(getattr(co, "email", "")), NCOLS)
        r = kv_full(ws, r, "Phone", _firstv(getattr(co, "phone", "")), NCOLS)
        if getattr(co, "address", ""):
            r = kv_full(ws, r, "Address", co.address, NCOLS)
        if getattr(co, "tax_office", ""):
            r = kv_full(ws, r, "Tax Office", co.tax_office, NCOLS)
        if getattr(co, "tax_number", ""):
            r = kv_full(ws, r, "Tax No", co.tax_number, NCOLS)
    elif order.web_client_id and order.web_client:
        wc = order.web_client
        r = kv_full(ws, r, "Type", "Web Customer", NCOLS)
        r = kv_full(ws, r, "Name", getattr(wc, "name", "") or getattr(wc, "username", "") or "—", NCOLS, bold_value=True)
        r = kv_full(ws, r, "Email", getattr(wc, "email", ""), NCOLS)
        r = kv_full(ws, r, "Phone", getattr(wc, "phone", ""), NCOLS)
    elif getattr(order, "is_guest_order", False):
        r = kv_full(ws, r, "Type", "Guest", NCOLS)
        gname = " ".join(p for p in [order.guest_first_name or "", order.guest_last_name or ""] if p)
        r = kv_full(ws, r, "Name", gname or "—", NCOLS, bold_value=True)
        r = kv_full(ws, r, "Email", order.guest_email or "", NCOLS)
        r = kv_full(ws, r, "Phone", order.guest_phone or "", NCOLS)
    else:
        r = kv_full(ws, r, "Name", "—", NCOLS, bold_value=True)
    r += 1

    # ── Delivery address ──
    r = section(ws, r, "DELIVERY ADDRESS", NCOLS)
    if order.delivery_address or order.delivery_city:
        if order.delivery_address_title:
            r = kv_full(ws, r, "Title", order.delivery_address_title, NCOLS)
        if order.delivery_address:
            r = kv_full(ws, r, "Address", order.delivery_address, NCOLS)
        loc = ", ".join([x for x in [order.delivery_city, order.delivery_country] if x])
        if loc:
            r = kv_full(ws, r, "City / Country", loc, NCOLS)
        if order.delivery_phone:
            r = kv_full(ws, r, "Phone", order.delivery_phone, NCOLS)
    else:
        r = kv_full(ws, r, "Address", "Same as customer", NCOLS)
    r += 1

    # ── Billing address (only if set) ──
    if order.billing_address or order.billing_city:
        r = section(ws, r, "BILLING ADDRESS", NCOLS)
        if order.billing_address_title:
            r = kv_full(ws, r, "Title", order.billing_address_title, NCOLS)
        if order.billing_address:
            r = kv_full(ws, r, "Address", order.billing_address, NCOLS)
        loc = ", ".join([x for x in [order.billing_city, order.billing_country] if x])
        if loc:
            r = kv_full(ws, r, "City / Country", loc, NCOLS)
        if order.billing_phone:
            r = kv_full(ws, r, "Phone", order.billing_phone, NCOLS)
        r += 1

    # ── Items ──
    items = list(order.items.all().select_related("product", "product_variant"))
    r = section(ws, r, f"PRODUCTS ({len(items)})", NCOLS)
    heads = ["Product", "SKU", "Variant", "Qty", "Unit", "Amount"]
    for i, h in enumerate(heads, 1):
        cell(ws, r, i, h, font=F_HEAD, fill=FILL_HEAD, border=GRID,
             align=(RIGHT if i >= 4 else LEFT))
    r += 1

    total = Decimal("0.00")
    for it in items:
        qty = it.quantity or Decimal("0")
        price = it.price or Decimal("0")
        line = (qty * price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        total += line
        title = _break_words(getattr(it.product, "title", None)
                             or str(it.product or "—"))
        if getattr(it, "description", ""):
            title = f"{title}\n{it.description}"
        vsku = it.product_variant.variant_sku if (it.product_variant_id and it.product_variant) else "—"
        cell(ws, r, 1, title, font=F_VAL, border=GRID, align=TOP)
        cell(ws, r, 2, getattr(it.product, "sku", "") or "—", font=F_VAL, border=GRID)
        cell(ws, r, 3, vsku or "—", font=F_VAL, border=GRID)
        cell(ws, r, 4, _dec(qty), font=F_VAL, border=GRID, align=RIGHT, fmt="#,##0.00")
        cell(ws, r, 5, _dec(price), font=F_VAL, border=GRID, align=RIGHT, fmt=money)
        cell(ws, r, 6, _dec(line), font=F_VAL, border=GRID, align=RIGHT, fmt=money)
        r += 1

    # ── Totals ──
    paid = _dec(getattr(order, "paid_amount", 0))
    grand = _dec(total)
    rows = [("Total", grand, True)]
    if paid:
        rows += [("Paid", paid, False), ("Balance", grand - paid, True)]
    for lbl, val, strong in rows:
        cell(ws, r, 4, lbl, font=(F_TOTAL if strong else F_VALB), border=GRID, align=RIGHT)
        merge(ws, r, 4, 5)
        merge_border(ws, r, 4, 5, GRID)
        cell(ws, r, 6, val, font=(F_TOTAL if strong else F_VALB), border=GRID, align=RIGHT, fmt=money)
        r += 1

    # ── Notes ──
    if order.notes:
        r += 1
        r = section(ws, r, "NOTES", NCOLS)
        cell(ws, r, 1, order.notes, font=F_VAL, border=GRID, align=TOP)
        merge(ws, r, 1, NCOLS)
        merge_border(ws, r, 1, NCOLS, GRID)
        ws.row_dimensions[r].height = 46

    return wb


# ── Combined: several of one customer's orders ───────────────────────
# The printed sheet's columns, with each SKU given one of its own rather
# than folded under the name it belongs to: someone who exports this is
# about to sort or pivot, and neither works on a value tucked under
# another. The money columns name the currency in the HEADING and hold
# bare numbers — see the note on `money` below.
COMBINED_HEADS = ["Product", "SKU", "Variant", "Variant SKU", "Type",
                  "Quantity", "Packs", "Price", "Amount"]
CNCOLS = len(COMBINED_HEADS)
C_QTY, C_PACKS, C_PRICE, C_AMOUNT = 6, 7, 8, 9


def build_combined_workbook(orders):
    """One customer's orders as a single sheet of line items.

    The lines come from build_order_print_rows — the same builder the
    printed sheet uses — so the two documents cannot disagree about what
    a line says or what it comes to. Quantities and money are written as
    NUMBERS with a display format, not as text: the point of taking this
    into a spreadsheet is to do arithmetic on it, and a figure carrying
    its own currency is a figure you have to strip before you can.
    """
    from openpyxl import Workbook
    from accounting.services_accounts import brand_name_for
    from .views import build_order_print_rows

    primary = orders[-1]
    brand = (primary.print_header or "").strip() or brand_name_for()
    ccode = (primary.original_currency or primary.paid_currency or "USD")
    # A currency FORMAT, not a currency written into the cell: the cell
    # holds 2.45 and shows $2.45, so it sums, charts and multiplies like
    # the number it is. What it must never be is the string "2.45 USD",
    # which has to be stripped before any of that works.
    money = _money_format(ccode)

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.sheet_view.showGridLines = False

    dates = [o.order_date for o in orders if o.order_date]
    span = "—"
    if dates:
        span = _dt(dates[0], "%d %b %Y")
        if dates[-1] != dates[0]:
            span = f"{span} – {_dt(dates[-1], '%d %b %Y')}"

    # ── Header ──
    r = 1
    cell(ws, r, 1, brand, font=F_TITLE)
    merge(ws, r, 1, 5)
    cell(ws, r, 6, f"{len(orders)} ORDERS" if len(orders) != 1 else "ORDER",
         font=F_DOCNO, align=RIGHT)
    merge(ws, r, 6, CNCOLS)
    # The wordmark is 20pt in a row sized for 11pt text, so its descenders
    # were being cut off by the row below it. Row heights are in points.
    ws.row_dimensions[r].height = 28
    r += 1
    cell(ws, r, 1, "Order Confirmation", font=F_SUB)
    merge(ws, r, 1, 5)
    cell(ws, r, 6, span, font=F_SUB, align=RIGHT)
    merge(ws, r, 6, CNCOLS)
    for c in range(1, CNCOLS + 1):
        ws.cell(r, c).border = RULE
    r += 2

    # ── Customer. One customer for the whole sheet — select_combined_orders
    # refuses a selection that spans two — so this is asked once. ──
    r = section(ws, r, "CUSTOMER", CNCOLS)
    who = primary.contact or primary.company or primary.web_client
    name = (getattr(who, "name", None) or getattr(who, "username", None)
            or "—") if who else "—"
    r = kv_full(ws, r, "Name", name, CNCOLS, bold_value=True)
    r = kv_full(ws, r, "Orders", ", ".join(
        o.order_number or f"#{o.pk}" for o in orders), CNCOLS)
    # The books these orders post to. Named because they are the reason
    # this sheet exists: one customer, two ledgers, one document.
    books = []
    for o in orders:
        b = o.cari.book.name if (o.cari_id and o.cari and o.cari.book_id) else None
        if b and b not in books:
            books.append(b)
    if books:
        r = kv_full(ws, r, "Books", ", ".join(books), CNCOLS)
    r += 1

    # ── Items ──
    rows, packs = [], set()
    for o in orders:
        o_rows, _t, _q, o_packs = build_order_print_rows(o)
        rows.extend((o, it) for it in o_rows)
        # A pack scanned onto two lines is one pack, so the total is the
        # size of the SET — the rule the printed sheet's foot goes by.
        packs |= o_packs
    r = section(ws, r, f"PRODUCTS ({len(rows)})", CNCOLS)
    for i, h in enumerate(COMBINED_HEADS, 1):
        if i in (C_PRICE, C_AMOUNT):
            # The sign is in the cells now, so the heading names the
            # currency only where a sign is ambiguous — "$" is worn by
            # more than one dollar, "₺" by exactly one lira.
            if _currency_symbol(ccode) in ("$", "kr", ccode):
                h = f"{h} ({ccode})"
        cell(ws, r, i, h, font=F_HEAD, fill=FILL_HEAD, border=GRID,
             align=(RIGHT if i >= C_QTY else LEFT))
    r += 1
    head_row = r

    total = Decimal("0.00")
    total_qty = Decimal("0")
    for o, it in rows:
        qty = it.quantity or Decimal("0")
        line = it.line_total_calc
        total += line
        total_qty += qty
        title = _break_words(getattr(it.product, "title", None)
                             or str(it.product or "—"))
        if getattr(it, "description", ""):
            title = f"{title}\n{it.description}"
        vsku = (it.product_variant.variant_sku
                if (it.product_variant_id and it.product_variant) else None)
        cell(ws, r, 1, title, font=F_VAL, border=GRID, align=TOP)
        cell(ws, r, 2, getattr(it.product, "sku", "") or "—", font=F_VAL, border=GRID)
        cell(ws, r, 3, it.variant_label or "—", font=F_VAL, border=GRID)
        cell(ws, r, 4, vsku or "—", font=F_VAL, border=GRID)
        cell(ws, r, 5, it.product_type_label or "—", font=F_VAL, border=GRID)
        cell(ws, r, C_QTY, _dec(qty), font=F_VAL, border=GRID, align=RIGHT, fmt="#,##0.00")
        cell(ws, r, C_PACKS, it.pack_count or 0, font=F_VAL, border=GRID, align=RIGHT, fmt="#,##0")
        cell(ws, r, C_PRICE, _dec(it.price), font=F_VAL, border=GRID, align=RIGHT, fmt=money)
        cell(ws, r, C_AMOUNT, _dec(line), font=F_VAL, border=GRID, align=RIGHT, fmt=money)
        r += 1

    # ── Total ──
    # The label's span is NOT merged. A fill does not survive being
    # assigned to a merged-over cell, so the tint stopped at column A
    # and the row read as half-shaded — and a merge across a table
    # people are going to sort and filter is worth avoiding anyway. The
    # word sits in A and the cells beside it are simply dressed to
    # match, which looks the same and behaves better.
    for c in range(1, C_QTY):
        cell(ws, r, c, "TOTAL" if c == 1 else "", font=F_GRAND,
             fill=FILL_TOTAL, border=BORDER_TOTAL)
    cell(ws, r, C_QTY, _dec(total_qty), font=F_GRAND, fill=FILL_TOTAL,
         border=BORDER_TOTAL, align=RIGHT, fmt="#,##0.00")
    cell(ws, r, C_PACKS, len(packs), font=F_GRAND, fill=FILL_TOTAL,
         border=BORDER_TOTAL, align=RIGHT, fmt="#,##0")
    cell(ws, r, C_PRICE, "", font=F_GRAND, fill=FILL_TOTAL, border=BORDER_TOTAL)
    cell(ws, r, C_AMOUNT, _dec(total), font=F_GRAND, fill=FILL_TOTAL,
         border=BORDER_TOTAL, align=RIGHT, fmt=money)
    ws.row_dimensions[r].height = 22

    # The item rows get a filter, which is the first thing anyone who
    # exported this will want — by order, by product, by colour.
    # ── Fit the table to what is in it ───────────────────────────────
    # Both after the fact, and in this order: a column is sized to its
    # widest cell, and only then can a row be asked how many lines its
    # cells take. Measured over the item table alone — the header block
    # above it is merged across the sheet, and a merged value would size
    # every column to the whole of itself.
    fit_columns(ws, head_row - 1, r, CNCOLS, header_row=head_row - 1)
    fit_rows(ws, head_row, r, CNCOLS)

    last_col = chr(ord("A") + CNCOLS - 1)
    if rows:
        ws.auto_filter.ref = f"A{head_row - 1}:{last_col}{head_row + len(rows) - 1}"
    ws.freeze_panes = ws.cell(head_row, 1)

    # ── How it prints ────────────────────────────────────────────────
    # Nine columns of a customer's goods do not fit the width of a
    # portrait page, and a spreadsheet that spills its last two columns
    # onto sheets of their own is not a document anyone can hand over.
    # Landscape, scaled to exactly one page WIDE and as many pages long
    # as it takes (fitToHeight=0 means "don't limit the height"), with
    # the column headings repeated at the top of every page so page two
    # can be read at all.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    if rows:
        ws.print_title_rows = f"{head_row - 1}:{head_row - 1}"
    ws.print_area = f"A1:{last_col}{r}"

    return wb


@login_required
def combined_order_excel(request):
    """Download several of one customer's orders as one .xlsx."""
    from .views import select_combined_orders
    orders, refusal = select_combined_orders(request)
    if refusal is not None:
        return refusal
    wb = build_combined_workbook(orders)
    buf = BytesIO()
    wb.save(buf)
    who = orders[-1].contact or orders[-1].company or orders[-1].web_client
    label = (getattr(who, "name", None) or getattr(who, "username", None) or "orders")
    label = "".join(ch if ch.isalnum() or ch in " -_" else "-" for ch in label).strip()
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{label or "orders"}.xlsx"'
    return resp


@login_required
def order_excel(request, pk):
    """Download the order as an .xlsx file."""
    order = get_object_or_404(Order, pk=pk)
    wb = build_order_workbook(order)
    buf = BytesIO()
    wb.save(buf)
    label = order.order_number or f"order-{order.pk}"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{label}.xlsx"'
    return resp
