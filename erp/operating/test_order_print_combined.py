"""Several of one customer's orders on one printable sheet.

A customer who buys from two divisions holds an account in each book —
that is how the ledger is built, and it is right: their money posts in
two places. But they are one person, and asking them to read three
separate sheets for three orders is the ledger's shape leaking onto the
customer's desk.

So: tick the orders on the contact page, print them together. The sheet
records NOTHING — no invoice, no ledger row — which is exactly what lets
it cross books: the receivable stays on each order, in the book it
belongs to, and this is only the page the customer reads.

Run:
    python manage.py test operating.test_order_print_combined
"""
from decimal import Decimal
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from accounting.models import Book, CurrencyCategory
from accounting.models_accounts import CariAccount
from crm.models import Contact
from marketing.models import Product

from .models import Order, OrderItem

User = get_user_model()


class OlegOrders:
    """Three orders for one customer across two books — the split that
    sent this customer three documents. Shared by the printed sheet and
    the Excel of it, which are one document in two formats."""

    @patch("marketing.utils.bunny_storage.upload_to_bunny")
    def setUp(self, mock_upload):
        mock_upload.return_value = "https://mock-cdn.net/qr.png"
        self.usd = CurrencyCategory.objects.create(
            code="USD", name="US Dollar", symbol="$")
        self.laleli = Book.objects.create(name="Laleli Fabric")
        self.ergene = Book.objects.create(name="Ergene Fabric")

        self.oleg = Contact.objects.create(name="OLEG MOTUZENKO")
        self.someone_else = Contact.objects.create(name="BURSA TEKSTIL")

        self.crepe = Product.objects.create(title="Crepe", sku="KZL000315", price=10)

        # Two Laleli orders and one from the factory — the split that
        # sent this customer three documents.
        self.laleli_a = self._order(self.laleli, self.oleg, "DK-284",
                                    Decimal("156.00"), Decimal("2.50"))
        self.laleli_b = self._order(self.laleli, self.oleg, "DK-275",
                                    Decimal("40.00"), Decimal("4.41"))
        self.ergene_a = self._order(self.ergene, self.oleg, "DK-291",
                                    Decimal("19.50"), Decimal("4.50"))

        self.boss = User.objects.create_superuser("boss", "b@t.com", "pw")
        self.client.force_login(self.boss)
        self.url = reverse(self.url_name)

    def _order(self, book, contact, number, qty, price):
        cari, _ = CariAccount.objects.get_or_create(
            book=book, contact=contact,
            defaults=dict(code=f"C-{book.pk}-{contact.pk}", name=contact.name,
                          type="customer", default_currency=self.usd))
        order = Order.objects.create(order_number=number, cari=cari,
                                     contact=contact)
        OrderItem.objects.create(order=order, product=self.crepe,
                                 quantity=qty, price=price)
        return order

    def _get(self, *orders):
        ids = ",".join(str(o.pk) for o in orders)
        return self.client.get(self.url, {"ids": ids})


class CombinedOrderSheet(OlegOrders, TestCase):
    url_name = "operating:order_print_combined"

    # ── what it prints ────────────────────────────────────────────────
    def test_it_totals_every_order_on_the_sheet(self):
        """156.00×2.50 + 40.00×4.41 + 19.50×4.50 = 390.00 + 176.40 + 87.75."""
        resp = self._get(self.laleli_a, self.laleli_b, self.ergene_a)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["order_total"], Decimal("654.15"))
        self.assertEqual(resp.context["order_total_quantity"], Decimal("215.50"))
        self.assertContains(resp, "$654.15")

    def test_each_order_is_named_above_its_own_lines(self):
        """Three orders on one page, and the customer can still tell
        which line came from which."""
        resp = self._get(self.laleli_a, self.laleli_b, self.ergene_a)
        for number in ("DK-284", "DK-275", "DK-291"):
            self.assertContains(resp, number)
        self.assertEqual([g["order"].pk for g in resp.context["order_groups"]],
                         [self.laleli_a.pk, self.laleli_b.pk, self.ergene_a.pk])

    def test_it_spans_books(self):
        """The whole point. An Invoice cannot do this — its book is the
        book its money posts to — but this sheet posts nothing."""
        resp = self._get(self.laleli_a, self.ergene_a)
        self.assertEqual(resp.status_code, 200)
        books = {g["order"].cari.book.name for g in resp.context["order_groups"]}
        self.assertEqual(books, {"Laleli Fabric", "Ergene Fabric"})

    def test_it_writes_nothing(self):
        """No invoice is raised and no ledger row written. The
        receivable already sits on each order; a document that posted
        again would claim the same money twice."""
        from accounting.models_accounts import CariMovement, Invoice
        before = (Invoice.objects.count(), CariMovement.objects.count())
        self._get(self.laleli_a, self.laleli_b, self.ergene_a)
        self.assertEqual(
            (Invoice.objects.count(), CariMovement.objects.count()), before)

    # ── what it refuses ───────────────────────────────────────────────
    def test_two_customers_cannot_share_a_sheet(self):
        """The orders are named by id in a URL anyone can retype. One
        sheet is addressed to one customer."""
        theirs = self._order(self.laleli, self.someone_else, "DK-999",
                             Decimal("10"), Decimal("1"))
        resp = self._get(self.laleli_a, theirs)
        self.assertEqual(resp.status_code, 400)

    def test_an_order_with_no_customer_is_not_printable(self):
        """(None, None, None) matches (None, None, None), so without
        this an unattached order would combine with any other."""
        nobody = Order.objects.create(order_number="DK-NOCARI")
        resp = self._get(nobody)
        self.assertEqual(resp.status_code, 400)

    def test_no_orders_picked(self):
        self.assertEqual(self.client.get(self.url).status_code, 400)
        self.assertEqual(self.client.get(self.url, {"ids": "x"}).status_code, 400)

    def test_a_missing_order_is_not_silently_dropped(self):
        """Printing two of the three asked for would be a wrong total on
        a document the customer is handed."""
        resp = self.client.get(self.url, {"ids": f"{self.laleli_a.pk},9999999"})
        self.assertEqual(resp.status_code, 404)

    def test_a_signed_out_visitor_gets_nothing(self):
        """The sheet carries a customer's prices. Every order with a
        cari is already refused — member_can_use_book says no to a
        viewer with no member — and this covers the one without."""
        self.client.logout()
        resp = self._get(self.laleli_a)
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/authentication/signin", resp["Location"])

    def test_a_book_the_viewer_is_not_assigned_is_refused(self):
        """The rule book_guarded applies to an order page, asked here
        per order — a sheet may legitimately span books, but only the
        viewer's own."""
        laleli_only = User.objects.create_user("laleli_only", password="pw")
        laleli_only.member.books.add(self.laleli)
        laleli_only.member.default_book = self.laleli
        laleli_only.member.save()
        self.client.force_login(laleli_only)

        self.assertEqual(self._get(self.laleli_a, self.laleli_b).status_code, 200)
        self.assertEqual(self._get(self.laleli_a, self.ergene_a).status_code, 404)


class CombinedOrderExcel(OlegOrders, TestCase):
    """The same sheet as a spreadsheet."""

    url_name = "operating:order_excel_combined"

    def _book(self, *orders):
        import openpyxl
        from io import BytesIO
        resp = self._get(*orders)
        self.assertEqual(resp.status_code, 200)
        return openpyxl.load_workbook(BytesIO(resp.content)).active

    AMOUNT = 9  # the last column — see COMBINED_HEADS

    def test_it_comes_to_what_the_printed_sheet_comes_to(self):
        ws = self._book(self.laleli_a, self.laleli_b, self.ergene_a)
        self.assertEqual(ws.cell(ws.max_row, self.AMOUNT).value, 654.15)

    def test_money_is_a_number_a_spreadsheet_can_add(self):
        """The point of exporting is to do arithmetic on it, so the
        figures are numbers with a display format — not text that only
        looks like money."""
        ws = self._book(self.laleli_a)
        # int or float — openpyxl reads a whole number back as an int.
        # What matters is that it is not a string that looks like money.
        cell = ws.cell(ws.max_row, self.AMOUNT)
        self.assertIsInstance(cell.value, (int, float))
        # A number FORMAT — the sign and separators are display, and
        # what the cell holds is a figure arithmetic can reach.
        self.assertIn("#,##0.00", cell.number_format)

    def test_money_wears_its_sign_as_a_format_not_as_text(self):
        """"$2.45" the cell SHOWS, 2.45 the cell HOLDS. A figure
        carrying its currency as text has to be stripped before it can
        be summed, which is the one thing a spreadsheet is for."""
        ws = self._book(self.laleli_a)
        cell = ws.cell(ws.max_row, self.AMOUNT)
        self.assertEqual(cell.number_format, '"$"#,##0.00')
        self.assertIsInstance(cell.value, (int, float))
        for row in ws.iter_rows(min_row=1, values_only=True):
            for v in row:
                if isinstance(v, str):
                    self.assertNotIn("0.00 USD", v)

    def test_the_heading_still_names_the_currency(self):
        """More than one currency signs itself "$"."""
        heads = [c.value for c in self._book(self.laleli_a)[10]]
        self.assertIn("Price (USD)", heads)
        self.assertIn("Amount (USD)", heads)

    def test_a_long_product_name_breaks_across_lines_in_its_column(self):
        """The cells wrap already, but openpyxl writes no row heights
        and Excel then leaves the row one line tall, so a 40-character
        name ran off the side of its column instead of breaking inside
        it."""
        long_name = Product.objects.create(
            title="GREK TAŞLI VE İNCİ EKRU İNCİ BEYAZ ZEMİN",
            sku="HKN00011", price=10)
        OrderItem.objects.create(order=self.laleli_a, product=long_name,
                                 quantity=Decimal("150"), price=Decimal("4.50"))
        ws = self._book(self.laleli_a)
        row = next(r for r in range(11, ws.max_row + 1)
                   if str(ws.cell(r, 1).value or "").startswith("GREK"))
        # Broken where the rule says, not where the column happens to
        # run out: four words a line, so eight words come out square.
        self.assertEqual(ws.cell(row, 1).value,
                         "GREK TAŞLI VE İNCİ\nEKRU İNCİ BEYAZ ZEMİN")
        self.assertTrue(ws.cell(row, 1).alignment.wrap_text)
        self.assertGreaterEqual(ws.row_dimensions[row].height or 0, 27)
        # A short name is left alone rather than padded to match.
        short = next(r for r in range(11, ws.max_row + 1)
                     if ws.cell(r, 1).value == "Crepe")
        self.assertIsNone(ws.row_dimensions[short].height)

    def test_a_name_of_four_words_or_fewer_is_left_whole(self):
        four = Product.objects.create(title="Ekru Grek Taşlı İnci",
                                      sku="PIL01", price=1)
        OrderItem.objects.create(order=self.laleli_a, product=four,
                                 quantity=Decimal("1"), price=Decimal("1"))
        titles = [self._book(self.laleli_a).cell(r, 1).value
                  for r in range(11, 40)]
        self.assertIn("Ekru Grek Taşlı İnci", titles)

    def test_the_total_row_stands_out_from_the_lines_it_totals(self):
        """At the foot of thirty half-bold rows, bold alone is not a
        difference anyone sees."""
        ws = self._book(self.laleli_a, self.laleli_b, self.ergene_a)
        last, line = ws.max_row, ws.max_row - 1
        for c in range(1, 10):
            total, item = ws.cell(last, c), ws.cell(line, c)
            self.assertEqual(total.fill.fgColor.rgb, "FFF3F6F8")
            self.assertEqual(total.border.top.style, "medium")
            self.assertEqual(total.border.bottom.style, "double")
            self.assertNotEqual(total.fill.fgColor.rgb, item.fill.fgColor.rgb)
        self.assertGreater(ws.cell(last, 9).font.size, ws.cell(line, 9).font.size)

    def test_it_prints_onto_a_page(self):
        """Nine columns do not fit a portrait page, and a sheet that
        spills its last two columns onto pages of their own is not a
        document anyone can hand over."""
        ws = self._book(self.laleli_a, self.laleli_b, self.ergene_a)
        self.assertEqual(ws.page_setup.orientation, "landscape")
        self.assertTrue(ws.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(ws.page_setup.fitToWidth, 1)
        # 0 means "as many pages tall as it takes" — only the WIDTH is
        # being forced onto one page.
        self.assertEqual(ws.page_setup.fitToHeight, 0)
        # The column headings repeat, so page two can be read at all.
        self.assertTrue(ws.print_title_rows)

    def test_the_wordmark_gets_a_row_tall_enough_to_hold_it(self):
        """20pt type in a row sized for 11pt loses its descenders under
        the row beneath."""
        self.assertGreaterEqual(self._book(self.laleli_a).row_dimensions[1].height, 26)

    def test_it_is_delivered_as_a_spreadsheet(self):
        resp = self._get(self.laleli_a)
        self.assertIn("spreadsheetml.sheet", resp["Content-Type"])
        self.assertIn("attachment;", resp["Content-Disposition"])
        self.assertIn("OLEG MOTUZENKO", resp["Content-Disposition"])

    def test_it_refuses_what_the_printed_sheet_refuses(self):
        """Both go through select_combined_orders — a rule enforced in
        only one of two formats is not a rule."""
        theirs = self._order(self.laleli, self.someone_else, "DK-999",
                             Decimal("10"), Decimal("1"))
        self.assertEqual(self._get(self.laleli_a, theirs).status_code, 400)
        self.assertEqual(self.client.get(self.url).status_code, 400)
        self.assertEqual(
            self.client.get(self.url,
                            {"ids": f"{self.laleli_a.pk},9999999"}).status_code, 404)

    def test_a_signed_out_visitor_gets_nothing(self):
        self.client.logout()
        self.assertEqual(self._get(self.laleli_a).status_code, 302)

    def test_it_carries_the_variant_name_and_both_codes(self):
        ws = self._book(self.laleli_a)
        heads = [ws.cell(r, c).value
                 for r in range(1, ws.max_row + 1) for c in range(1, 10)]
        for h in ("Product", "SKU", "Variant", "Variant SKU", "Type",
                  "Quantity", "Packs"):
            self.assertIn(h, heads)


class SingleOrderPrintUnchanged(TestCase):
    """The one-order sheet builds its rows through the same helper now.
    It must still print exactly what it printed before — one order needs
    no per-order heading above its lines."""

    @patch("marketing.utils.bunny_storage.upload_to_bunny")
    def setUp(self, mock_upload):
        mock_upload.return_value = "https://mock-cdn.net/qr.png"
        usd = CurrencyCategory.objects.create(code="USD", name="US Dollar", symbol="$")
        book = Book.objects.create(name="Laleli Fabric")
        cari = CariAccount.objects.create(
            book=book, code="C-1", name="Oleg", type="customer",
            default_currency=usd)
        self.order = Order.objects.create(order_number="DK-284", cari=cari,
                                          contact=Contact.objects.create(name="OLEG"))
        OrderItem.objects.create(
            order=self.order,
            product=Product.objects.create(title="Crepe", sku="KZL000315", price=10),
            quantity=Decimal("156.00"), price=Decimal("2.50"))
        self.client.force_login(User.objects.create_superuser("boss", "b@t.com", "pw"))

    def test_it_still_prints_its_total(self):
        resp = self.client.get(
            reverse("operating:order_print", kwargs={"pk": self.order.pk}))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["order_total"], Decimal("390.00"))
        self.assertContains(resp, "$390.00")

    def test_it_carries_no_group_heading(self):
        resp = self.client.get(
            reverse("operating:order_print", kwargs={"pk": self.order.pk}))
        self.assertEqual(len(resp.context["order_groups"]), 1)
        self.assertNotContains(resp, 'class="grp"')
