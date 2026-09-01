"""Packing list: item numbering and the product-type column.

Lives outside tests.py because that module currently fails to import
(a stale `from .models import Product` that no longer resolves), which
would take these down with it.
"""
from decimal import Decimal
from unittest.mock import patch

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse

from accounting.models import Book
from marketing.models import (Product, ProductCategory, ProductVariant,
                              ProductVariantAttribute,
                              ProductVariantAttributeValue)
from .models import (Order, OrderItem, OrderRollReservation, Pack, Warehouse,
                     WarehouseProduct, WarehouseProductRoll)
from .views import (_long_date, _pack_roll_rows, _product_type_label,
                    _variant_label)


class PackingListColumns(TestCase):
    @patch("marketing.utils.bunny_storage.upload_to_bunny")
    def setUp(self, mock_upload):
        mock_upload.return_value = "https://mock-cdn.net/qr.png"
        fabric = ProductCategory.objects.create(name="fabric")
        self.colour = ProductVariantAttribute.objects.create(name="colour")
        # Group names are stored slugged; the printed column un-slugs them.
        linen = ProductCategory.objects.create(name="bed_linen")
        wh = Warehouse.objects.create(name="Test WH",
            accounting_book=Book.objects.get_or_create(name="Laleli Fabric")[0])
        self.order = Order.objects.create(order_number="DK0000270")

        # Two packages of two rolls each — enough to prove the item
        # number keeps counting across a package boundary.
        for pack_no in (1, 2):
            pack = Pack.objects.create(order=self.order, pack_number=pack_no)
            for i, cat in enumerate((fabric, linen)):
                product = Product.objects.create(
                    title=f"Bergamo {pack_no}{i}", sku=f"SKU-{pack_no}{i}",
                    category=cat, price=10)
                variant = None
                if cat is fabric:
                    # Fabric is sold by colour, so its rolls are variants.
                    variant = ProductVariant.objects.create(
                        product=product, variant_sku=f"V-{pack_no}{i}")
                    # (attribute, value) is a shared lookup row.
                    value, _ = ProductVariantAttributeValue.objects.get_or_create(
                        product_variant_attribute=self.colour,
                        product_variant_attribute_value="light_cream")
                    variant.product_variant_attribute_values.add(value)
                item = OrderItem.objects.create(
                    order=self.order, product=product, quantity=1, price=10,
                    product_variant=variant)
                wp = WarehouseProduct.objects.create(
                    warehouse=wh, name=product.title, sku=product.sku, quantity=50)
                roll = WarehouseProductRoll.objects.create(
                    product=wp, meters=Decimal("50.00"),
                    meters_remaining=Decimal("50.00"), barcode=f"BC{pack_no}{i}")
                OrderRollReservation.objects.create(
                    order=self.order, order_item=item, roll=roll,
                    warehouse_product=wp, meters=Decimal("32.45"), pack=pack)

    def test_rows_carry_the_product_type(self):
        pack = self.order.packs.get(pack_number=1)
        types = [r["product_type"] for r in _pack_roll_rows(pack)]
        self.assertEqual(types, ["Fabric", "Bed Linen"])

    def test_a_typeless_product_falls_back_to_a_dash(self):
        Product.objects.update(category=None, type=None)
        pack = self.order.packs.get(pack_number=1)
        self.assertEqual([r["product_type"] for r in _pack_roll_rows(pack)], ["-", "-"])

    def test_free_text_type_stands_in_for_a_missing_group(self):
        Product.objects.update(category=None, type="curtain")
        pack = self.order.packs.get(pack_number=1)
        self.assertEqual([r["product_type"] for r in _pack_roll_rows(pack)], ["Curtain"] * 2)

    def test_item_numbers_run_across_packages_not_within_them(self):
        self.client.force_login(
            User.objects.create_superuser("t", "t@t.com", "pw"))
        resp = self.client.get(
            reverse("operating:order_packing_list", kwargs={"pk": self.order.pk}))
        self.assertEqual(resp.status_code, 200)
        rows = [r for p in resp.context["packs"] for r in p.rows]
        self.assertEqual([r["item_no"] for r in rows], [1, 2, 3, 4])

    def test_rows_carry_the_variant(self):
        """Attribute values are stored slugged (`light_cream`); the
        column prints them the way the warehouse picker does."""
        pack = self.order.packs.get(pack_number=1)
        self.assertEqual([r["variant"] for r in _pack_roll_rows(pack)],
                         ["Light Cream", "-"])

    def test_a_product_with_no_variants_prints_a_dash(self):
        self.assertEqual(_variant_label(None), "-")

    def test_multiple_attributes_are_joined(self):
        v = ProductVariant.objects.create(
            product=Product.objects.first(), variant_sku="V-MULTI")
        width = ProductVariantAttribute.objects.create(name="width")
        for attr, value in ((self.colour, "dark_beige"), (width, "150cm")):
            row, _ = ProductVariantAttributeValue.objects.get_or_create(
                product_variant_attribute=attr,
                product_variant_attribute_value=value)
            v.product_variant_attribute_values.add(row)
        self.assertEqual(_variant_label(v), "Dark Beige / 150Cm")

    def test_the_pdf_opens_in_a_tab_instead_of_downloading(self):
        resp = self.client.get(
            reverse("operating:order_packing_list_pdf", kwargs={"pk": self.order.pk}))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertTrue(resp["Content-Disposition"].startswith("inline;"))
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_the_excel_is_the_downloadable_editable_copy(self):
        resp = self.client.get(
            reverse("operating:export_packing_list_excel", kwargs={"pk": self.order.pk}))
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp["Content-Disposition"].startswith("attachment;"))

        import io
        import openpyxl
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        self.assertEqual(ws["A1"].value, "DEMFIRAT\u00ae | Karven Home Collection")
        self.assertEqual(ws["A2"].value, f"Packing List {self.order.pk}")
        self.assertEqual(
            [c.value for c in ws[7]],
            ["Pack", "No", "Product", "Variant", "Product Type", "SKU",
             "Barcode", "Metres"])
        # One merged block per package, down its own items.
        self.assertIn("A8:A9", [str(m) for m in ws.merged_cells.ranges])
        tail = [[c for c in row if c is not None]
                for row in ws.iter_rows(min_row=ws.max_row - 2, values_only=True)]
        self.assertEqual(tail, [["Total Packages", 2], ["Total Items", 4],
                                ["Total Metres", 129.8]])

    def test_totals_say_items_not_rolls(self):
        """"Rolls" only makes sense for fabric; the same list also
        carries curtains and bed linen."""
        import io
        import openpyxl
        resp = self.client.get(
            reverse("operating:export_packing_list_excel", kwargs={"pk": self.order.pk}))
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        text = " ".join(str(c.value) for row in ws.iter_rows() for c in row)
        self.assertIn("Total Items", text)
        self.assertNotIn("Rolls", text)


class LongDate(TestCase):
    def test_english_reads_month_day_year(self):
        from datetime import date
        self.assertEqual(_long_date(date(2026, 8, 20), False), "August 20, 2026")

    def test_turkish_reads_day_month_year(self):
        from datetime import date
        self.assertEqual(_long_date(date(2026, 8, 20), True), "20 A\u011fustos 2026")


class BrandHeader(TestCase):
    """Every document an order produces signs with the name held on the
    ledger BOOK, not the short BRAND_NAME used for UI chrome."""

    LOCKUP = "DEMFIRAT® | Karven Home Collection"

    def setUp(self):
        from accounting.models import Book
        self.book = Book.objects.create(name="DEMFIRAT",
                                        brand_name=self.LOCKUP)
        self.order = Order.objects.create(order_number="DK0000270")
        self.client.force_login(
            User.objects.create_superuser("b", "b@b.com", "pw"))

    def test_editing_the_book_changes_every_document(self):
        """The point of moving this onto the book: one edit, and the
        order print, its Excel and the packing list all follow."""
        import io
        import openpyxl
        self.book.brand_name = "KARVEN | Contract Textiles"
        self.book.save(update_fields=["brand_name"])

        printed = self.client.get(
            reverse("operating:order_print", kwargs={"pk": self.order.pk})
        ).content.decode()
        self.assertIn("KARVEN | Contract Textiles", printed)
        self.assertNotIn(self.LOCKUP, printed)

        for name in ("order_excel", "export_packing_list_excel"):
            ws = openpyxl.load_workbook(io.BytesIO(self.client.get(
                reverse(f"operating:{name}", kwargs={"pk": self.order.pk})
            ).content)).active
            self.assertEqual(ws["A1"].value, "KARVEN | Contract Textiles", name)

    def test_a_blank_book_name_falls_back_to_the_brand_default(self):
        self.book.brand_name = ""
        self.book.save(update_fields=["brand_name"])
        self.assertEqual(self.book.effective_brand_name, self.LOCKUP)

    def test_the_printable_order_defaults_to_the_lockup(self):
        resp = self.client.get(
            reverse("operating:order_print", kwargs={"pk": self.order.pk}))
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn(self.LOCKUP, body)
        # Not shouted: the old template ran BRAND_NAME through |upper.
        self.assertNotIn("KARVEN HOME COLLECTION", body)

    def test_a_per_order_print_header_still_wins(self):
        self.order.print_header = "Karven Contract Division"
        self.order.save(update_fields=["print_header"])
        body = self.client.get(
            reverse("operating:order_print", kwargs={"pk": self.order.pk})
        ).content.decode()
        self.assertIn("Karven Contract Division", body)
        self.assertNotIn(self.LOCKUP, body)

    def test_the_order_excel_signs_the_same_way(self):
        import io
        import openpyxl
        resp = self.client.get(
            reverse("operating:order_excel", kwargs={"pk": self.order.pk}))
        self.assertEqual(resp.status_code, 200)
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        self.assertEqual(ws["A1"].value, self.LOCKUP)

    def test_the_packing_list_signs_the_same_way(self):
        import io
        import openpyxl
        resp = self.client.get(
            reverse("operating:export_packing_list_excel",
                    kwargs={"pk": self.order.pk}))
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        self.assertEqual(ws["A1"].value, self.LOCKUP)


class OrderPrintProductType(TestCase):
    """The printed ORDER identifies its lines the way the packing list
    does — kind of goods and variant NAME. It used to show neither: a
    customer reading it could not tell that "MT-3016" was fabric, nor
    that "MRK00061" was the beige-silver they ordered."""

    @patch("marketing.utils.bunny_storage.upload_to_bunny")
    def setUp(self, mock_upload):
        mock_upload.return_value = "https://mock-cdn.net/qr.png"
        self.order = Order.objects.create(order_number="DK0000289")
        self.product = Product.objects.create(
            title="MT-3016", sku="MT-3016", price=10,
            category=ProductCategory.objects.create(name="fabric"))
        self.variant = ProductVariant.objects.create(
            product=self.product, variant_sku="MRK00061")
        value, _ = ProductVariantAttributeValue.objects.get_or_create(
            product_variant_attribute=ProductVariantAttribute.objects.create(
                name="colour"),
            product_variant_attribute_value="bej-gumus")
        self.variant.product_variant_attribute_values.add(value)
        OrderItem.objects.create(order=self.order, product=self.product,
                                 quantity=Decimal("81.10"), price=10,
                                 product_variant=self.variant)
        self.client.force_login(
            User.objects.create_superuser("p", "p@t.com", "pw"))

    def _resp(self):
        resp = self.client.get(
            reverse("operating:order_print", kwargs={"pk": self.order.pk}))
        self.assertEqual(resp.status_code, 200)
        return resp

    def _items(self):
        return self._resp().context["order_items"]

    def _page(self):
        return self._resp().content.decode()

    def test_the_line_says_what_kind_of_goods_it_is(self):
        self.assertEqual(self._items()[0].product_type_label, "Fabric")

    def test_it_reads_the_same_word_the_packing_list_prints(self):
        """Both documents go through _product_type_label, so a customer
        holding the order and the packing list never sees two names for
        one kind of goods."""
        self.assertEqual(self._items()[0].product_type_label,
                         _product_type_label(self.product))

    def test_a_slugged_group_loses_its_underscore(self):
        self.product.category = ProductCategory.objects.create(
            name="ready-made_curtain")
        self.product.save()
        self.assertEqual(self._items()[0].product_type_label,
                         "Ready-Made Curtain")

    def test_an_unclassified_product_drops_the_line_instead_of_printing_a_dash(self):
        """The packing list has a column to fill so it prints "-"; this
        layout puts the kind under the product name, where a stray dash
        would read as missing data."""
        self.product.category = None
        self.product.type = None
        self.product.save()
        self.assertIsNone(self._items()[0].product_type_label)

    def test_the_line_names_the_variant_instead_of_its_sku(self):
        """"MRK00061" is a warehouse code; the customer ordered a colour."""
        item = self._items()[0]
        self.assertEqual(item.variant_label, "Bej-Gumus")
        self.assertNotIn("MRK00061", self._page())

    def test_the_variant_reads_the_same_as_on_the_packing_list(self):
        self.assertEqual(self._items()[0].variant_label,
                         _variant_label(self.variant))

    def test_a_line_with_no_variant_drops_the_chip(self):
        OrderItem.objects.update(product_variant=None)
        self.assertIsNone(self._items()[0].variant_label)

    def test_a_mill_coded_product_does_not_print_its_code_twice(self):
        """The title IS the SKU on 894 of 920 warehouse products, so
        showing both read "MT-3016 MT-3016"."""
        self.assertIsNone(self._items()[0].sku_label)

    def test_a_sku_that_says_something_the_title_does_not_still_shows(self):
        self.product.title = "Bergamo"
        self.product.save()
        self.assertEqual(self._items()[0].sku_label, "MT-3016")

    def _query_count(self):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext
        with CaptureQueriesContext(connection) as cap:
            self._items()
        return len(cap)

    def _add_lines(self, n, start=0):
        for i in range(start, start + n):
            p = Product.objects.create(title=f"P{i}", sku=f"S{i}", price=1,
                                       category=self.product.category)
            v = ProductVariant.objects.create(product=p, variant_sku=f"V{i}")
            v.product_variant_attribute_values.add(
                *self.variant.product_variant_attribute_values.all())
            OrderItem.objects.create(order=self.order, product=p, quantity=1,
                                     price=1, product_variant=v)

    def test_extra_lines_do_not_each_cost_their_own_queries(self):
        """Product group and variant attributes are pulled with the lines
        (select_related + prefetch_related), so a six-line order costs
        the same as an eleven-line one. Without them every line fetched
        its own category and its own attribute M2M.

        Both measurements are taken after the page has been rendered
        once: the first render also pays one-off costs (session, content
        types) that would otherwise swamp the comparison."""
        self._add_lines(5)
        self._items()                      # warm the one-off queries
        before = self._query_count()
        self._add_lines(5, start=5)
        self.assertEqual(self._query_count(), before)
