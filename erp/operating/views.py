import traceback
from django.shortcuts import get_object_or_404, render, redirect
from django.http import JsonResponse
from django.views import View
from django.contrib import messages
from .models import (
    Order,
    OrderItemUnit,
    Pack,
    PackedItem,
    RawMaterialGood,
    RawMaterialGoodReceipt,
    RawMaterialGoodItem,
)
from .forms import (
    OrderForm,
    OrderItemUnitForm,
    RawMaterialGoodReceiptForm,
    RawMaterialGoodItemForm,
)
from django.views.generic.edit import CreateView, UpdateView
from django.views.generic.detail import DetailView
from django.views.generic import ListView
from django.db import transaction
from django.forms import ValidationError, formset_factory
from django.views.decorators.http import require_POST

from .models import STATUS_CHOICES

# from accounting.models import Book, CurrencyCategory, AssetAccountsReceivable, Invoice


# segno is for making qr codes, and it is cleaner and more efficient than qrcode.
import segno

import tempfile
from marketing.utils.bunny_storage import upload_to_bunny

# make the qr codes jso


from django.utils.html import escape
from .models import (
    OrderItem,
)  # if it's not already in __init__.py
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from .utils import get_machine_from_api_key
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

# to filter products by multiple fields
from django.db.models import Q
from django.db import models


from marketing.models import Product, ProductVariant


from crm.models import Contact, Company


# functions are here:
def generate_machine_qr_for_order(order):
    payload = {"order_id": order.pk, "action": "update_status"}

    qr = segno.make(json.dumps(payload))  # Convert dict to JSON string

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
        qr.save(temp_file.name, scale=5)
        path = f"media/orders/{order.pk}/qr_order_{order.pk}.png"
        with open(temp_file.name, 'rb') as f:
            url = upload_to_bunny(f, path)

    order.qr_code_url = url
    order.save(update_fields=["qr_code_url"])


def generate_qr_for_order_item_unit(order_item_unit, status="scheduled"):
    order_id = order_item_unit.order_item.order.pk
    order_item_id = order_item_unit.order_item.pk
    order_item_unit_id = order_item_unit.pk
    payload = {
        "order_id": order_id,
        "order_item_id": order_item_id,
        "order_item_unit_id": order_item_unit_id,
        "status": status,
    }

    # this is how you make the qr with json data
    qr = segno.make(json.dumps(payload))

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
        qr.save(temp_file.name, scale=5)
        path = f"media/orders/{order_id}/items/{order_item_id}/units/{order_item_unit_id}/qr_{order_item_unit.pk}_{status}.png"
        with open(temp_file.name, 'rb') as f:
            url = upload_to_bunny(f, path)

    return url


# This is for reading the QR code on mobile for order_item_unit status updates.
def process_qr_payload(request):
    try:
        data = json.loads(request.body)
        order_id = data.get("order_id")
        order_item_id = data.get("order_item_id")
        order_item_unit_pk = data.get("order_item_unit_id")
        # new_status = data.get("status")
        new_status = data.get(
            "status", "pending"
        )  # Default to 'pending' if not provided

        # item = OrderItem.objects.get(pk=order_item_id)
        # item.status = new_status
        # item.save()
        order_item_unit = OrderItemUnit.objects.get(pk=order_item_unit_pk)
        order_item_unit.status = new_status
        order_item_unit.save(update_fields=["status"])

        return JsonResponse(
            {
                "success": True,
                "message": f"Item {order_item_unit.pk} updated to {new_status}",
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


def process_qr_payload_pack(request):
    try:
        with transaction.atomic():
            data = json.loads(request.body)
            order_id = data.get("order_id")
            order_item_id = data.get("order_item_id")
            order_item_unit_pk = data.get("order_item_unit_id")
            pack_number = data.get("pack_number")

            order = get_object_or_404(Order, pk=order_id)
            order_item_unit = OrderItemUnit.objects.get(pk=order_item_unit_pk)
            pack, created = Pack.objects.get_or_create(
                order=order, pack_number=pack_number
            )

            try:
                packed_item = PackedItem.objects.get(order_item_unit=order_item_unit)
                # Already packed: move to new pack
                packed_item.pack = pack
                packed_item.save()
                message = f"Unit {order_item_unit_pk} moved to pack {pack_number}."

            except PackedItem.DoesNotExist:
                # Not packed: create new PackedItem
                PackedItem.objects.create(pack=pack, order_item_unit=order_item_unit)
                message = f"Unit {order_item_unit_pk} packed in pack {pack_number}."

            return JsonResponse({"success": True, "message": message})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


def process_qr_payload_goods_received(request):
    try:
        with transaction.atomic():
            data = json.loads(request.body)
            receipt_number = data.get(
                "receipt_number"
            )  # this would be the ideal case to have.
            item_name = data.get("name")
            item_sku = data.get("sku")
            item_serial = data.get("serial")
            item_batch = data.get("batch")

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


# Create your views here.


class index(View):
    def get(self, request):
        context = {"message": "Hello from Operating index!"}
        # response = HttpResponse()
        # response.write("<h1>Hello</h1>")
        # return response
        return render(request, "operating/index.html", context)


class OrderDetail(DetailView):
    model = Order
    template_name = "operating/order_detail.html"
    context_object_name = "order"

    def get_queryset(self):
        # Prefetch items + their product/variant so gross_profit() can
        # read costs without extra queries (the template also iterates
        # items for the line breakdown).
        return Order.objects.select_related(
            "contact", "company", "web_client", "cari",
        ).prefetch_related(
            "items__product", "items__product_variant",
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Attach available-stock metadata to each item so the template
        # can show "Stok: N" next to the qty input and compute the max
        # the user can bump it to (current qty + remaining stock).
        for it in self.object.items.all():
            if it.product_variant_id and it.product_variant:
                stock = it.product_variant.variant_quantity
            elif it.product_id and it.product:
                stock = it.product.quantity
            else:
                stock = None
            it.available_stock = stock
            it.allow_oversell = bool(getattr(it.product, "selling_while_out_of_stock", False))
            try:
                cur = float(it.quantity or 0)
                avail = float(stock) if stock is not None else None
                it.max_qty = (cur + avail) if avail is not None else None
            except (TypeError, ValueError):
                it.max_qty = None
        # Alphabetical sort by product title (case-insensitive), with a
        # stable fallback to variant SKU then primary key so identical
        # titles keep a deterministic order.
        ctx["order_items_sorted"] = sorted(
            list(self.object.items.all()),
            key=lambda i: (
                (getattr(i.product, "title", "") or "").casefold(),
                (getattr(i.product_variant, "variant_sku", "") or ""),
                i.pk,
            ),
        )
        cari = self.object.cari
        if cari:
            # Last 5 ledger movements for the sidebar widget — cheap
            # query thanks to the (cari, -date) composite index.
            ctx["cari_recent_movements"] = list(
                cari.movements.select_related("currency").order_by("-date", "-id")[:5]
            )
            ctx["cari_invoices_count"] = cari.invoices.count()

        # Invoices linked to THIS order. Used to (a) show the user
        # what's already invoiced and (b) hide the "Create invoice"
        # button when an active invoice already exists — preventing
        # accidental duplicates that the user would have to clean up.
        order_invoices = list(
            self.object.invoices
            .select_related("cari", "currency")
            .order_by("-date", "-id")
        )
        ctx["order_invoices"] = order_invoices
        ctx["has_active_invoice"] = any(
            inv.status != "cancelled" for inv in order_invoices
        )

        # Status flow widget on the detail page — 3 ordered primary
        # stages (Açık → Paketleniyor → Gönderildi) with an "active up
        # to" highlight. We compare against ORDER_PRIMARY_STAGES so the
        # bar still reads correctly even if order_status is a legacy
        # value (e.g. "delivered" — that comes after "shipped", so all
        # stages are highlighted).
        from .models import ORDER_PRIMARY_STAGES, ORDER_STATUS_CHOICES, CARRIER_CHOICES
        primary = list(ORDER_PRIMARY_STAGES)
        current = self.object.order_status or "pending"
        # Position the bar fills up to. Statuses past "shipped"
        # (in_transit / delivered / …) all count as 'shipped reached'.
        if current in primary:
            active_idx = primary.index(current)
        elif current in {"in_transit", "out_for_delivery", "delivered"}:
            active_idx = len(primary) - 1   # shipped + later → fill all
        elif current in {"cancelled", "returned"}:
            active_idx = -1                  # nothing filled
        else:
            active_idx = 0                   # legacy pre-ship (confirmed/preparing) → Açık

        ctx["status_stages"] = [
            {"key": k, "label": dict(ORDER_STATUS_CHOICES).get(k, k), "filled": i <= active_idx}
            for i, k in enumerate(primary)
        ]
        ctx["status_choices"]  = ORDER_STATUS_CHOICES
        ctx["carrier_choices"] = CARRIER_CHOICES
        ctx["is_terminal"] = current in {"cancelled", "returned"}
        # Reserved rolls scanned during packing (for the "Gönderildi"
        # gate + a summary on the detail page). Unconsumed = still held.
        ctx["reservation_count"] = self.object.roll_reservations.filter(consumed=False).count()
        ctx["has_cargo_info"] = bool((self.object.carrier or "").strip() and (self.object.tracking_number or "").strip())
        # Change history preview (full page at order_changes).
        ctx["order_changes_preview"] = _decorate_order_changes(list(
            self.object.change_logs.select_related("created_by")[:8]
        ))
        ctx["order_changes_count"] = self.object.change_logs.count()
        return ctx

    def post(self, request, *args, **kwargs):
        """Inline edit endpoints. Handles three actions:
          - update_status: status / carrier / tracking_number bundle
          - update_item:   single OrderItem field (quantity or price)
          - update_notes:  order.notes
        """
        # NB: use a plain queryset here — get_queryset() prefetches items,
        # which means order.items.all() would return STALE cached rows even
        # after we save a new price/qty in this same request. That made the
        # movement / order_total snapshot wrong by one revision. Fresh fetch
        # forces every order.items.all() call to hit the DB.
        order = get_object_or_404(Order, pk=self.kwargs["pk"])
        from django.utils import timezone
        from decimal import Decimal, InvalidOperation
        from .models import ORDER_STATUS_CHOICES, CARRIER_CHOICES, OrderItem
        from current_account.services import (
            get_or_create_cari_for_order, post_order_movement, reverse_order_movement,
        )

        action = (request.POST.get("action") or "").strip()
        member = getattr(request.user, "member", None)

        def _sync_cari_total():
            """Re-post the cari movement for this order after its total
            changed (qty/price/add/remove). Idempotent — updates in place
            when the movement already exists."""
            try:
                if order.cari_id:
                    post_order_movement(order, member=member)
            except Exception:
                pass

        def _profit_snapshot():
            """Order-level cost + profit numbers for the totals widget.
            Refreshes after qty/price/add/remove so the user sees real
            margin without a page reload."""
            try:
                tc = float(order.total_cost() or 0)
                gp = float(order.gross_profit() or 0)
                ov = float(order.total_value() or 0)
                margin = round(gp / ov * 100, 1) if ov > 0 else None
                return {"total_cost": tc, "gross_profit": gp, "margin": margin}
            except Exception:
                return None

        def _cari_snapshot():
            """Return a small JSON-able dict with the current cari balance
            so the client can refresh the sidebar without a page reload.

            We deliberately fetch a fresh CariAccount + CariMovement from
            the DB here — the signal handler has already updated the
            movement.amount and re-aggregated cached_balance via
            CariAccount.objects.filter(...).update(), which does NOT
            invalidate any in-memory Cari instance. A direct DB read is
            the only way to get the post-update value."""
            if not order.cari_id:
                return None
            try:
                from current_account.models import CariAccount, CariMovement
                from django.contrib.contenttypes.models import ContentType
                cari = CariAccount.objects.get(pk=order.cari_id)
                ct = ContentType.objects.get_for_model(order.__class__)
                mv = CariMovement.objects.filter(
                    source_type=ct, source_id=order.pk, movement_type="order_sale",
                ).first()
                mv_amount = float(mv.amount) if mv else None
                mv_symbol = (mv.currency.symbol if (mv and mv.currency_id) else "") or ""
                bal = float(cari.cached_balance or 0)
                abs_bal = abs(bal)
                if bal > 0: color = "#B45309"
                elif bal < 0: color = "#147F74"
                else: color = "#374151"
                return {
                    "balance": bal,
                    "absolute_balance": abs_bal,
                    "balance_label": cari.balance_label,
                    "currency_symbol": cari.display_currency_symbol or "",
                    "color": color,
                    "mv_amount": mv_amount,
                    "mv_symbol": mv_symbol,
                }
            except Exception:
                return None

        # ── Inline item edit (qty / price) ──────────────────────
        if action == "update_item":
            try:
                item_id = int(request.POST.get("item_id") or 0)
            except (TypeError, ValueError):
                return JsonResponse({"ok": False, "error": "Bad item_id"}, status=400)
            field = (request.POST.get("field") or "").strip()
            value = (request.POST.get("value") or "").strip().replace(",", ".")

            if field not in {"quantity", "price"}:
                return JsonResponse({"ok": False, "error": "Field not editable"}, status=400)
            try:
                dec = Decimal(value)
            except (InvalidOperation, ValueError):
                return JsonResponse({"ok": False, "error": "Invalid number"}, status=400)
            if dec < 0:
                return JsonResponse({"ok": False, "error": "Must be ≥ 0"}, status=400)

            item = get_object_or_404(OrderItem, pk=item_id, order=order)

            # Stock cap intentionally not enforced here — qty changes
            # while the order is in a pre-fulfilment state don't touch
            # the catalog (see operating.signals). If the order is
            # already shipped, the signal will apply the delta and may
            # go negative if the user really wants that; their call.

            setattr(item, field, dec)
            item.save(update_fields=[field, "updated_at"])
            _sync_cari_total()

            # Refresh in-memory variant/product stock so the response
            # mirrors the post-deduction state.
            if item.product_variant_id:
                item.product_variant.refresh_from_db(fields=["variant_quantity"])
                new_stock = item.product_variant.variant_quantity
            elif item.product_id:
                item.product.refresh_from_db(fields=["quantity"])
                new_stock = item.product.quantity
            else:
                new_stock = None
            allow_oversell = bool(getattr(item.product, "selling_while_out_of_stock", False))
            max_qty = None
            if new_stock is not None and not allow_oversell:
                max_qty = float(item.quantity or 0) + float(new_stock)

            return JsonResponse({
                "ok": True,
                "item_id": item.pk,
                "subtotal": float(item.subtotal()),
                "order_total": float(order.total_value()),
                "cari": _cari_snapshot(),
                "profit": _profit_snapshot(),
                "stock": float(new_stock) if new_stock is not None else None,
                "max_qty": max_qty,
                "allow_oversell": allow_oversell,
            })

        # ── Inline notes edit ───────────────────────────────────
        if action == "update_notes":
            notes = request.POST.get("notes") or ""
            order.notes = notes
            order.save(update_fields=["notes", "updated_at"])
            return JsonResponse({"ok": True})

        # ── Toggle customer-notification opt-in ─────────────────
        if action == "update_notify":
            raw = (request.POST.get("notify_customer") or "").strip().lower()
            order.notify_customer = raw in ("1", "true", "on", "yes")
            order.save(update_fields=["notify_customer", "updated_at"])
            return JsonResponse({"ok": True, "notify_customer": order.notify_customer})

        # ── Preview add item (no database save) ──────────────────
        if action == "preview_add_item":
            sku = (request.POST.get("sku") or "").strip()
            is_variant = (request.POST.get("is_variant") or "").lower() == "true"
            qty_raw = (request.POST.get("quantity") or "1").replace(",", ".")
            price_raw = (request.POST.get("price") or "0").replace(",", ".")
            if not sku:
                return JsonResponse({"ok": False, "error": "Missing SKU"}, status=400)
            try:
                qty = Decimal(qty_raw)
                price = Decimal(price_raw)
            except (InvalidOperation, ValueError):
                return JsonResponse({"ok": False, "error": "Invalid number"}, status=400)
            if qty <= 0 or price < 0:
                return JsonResponse({"ok": False, "error": "Bad qty/price"}, status=400)

            variant = None
            if is_variant:
                variant = ProductVariant.objects.filter(variant_sku=sku).select_related("product").first()
                if not variant:
                    return JsonResponse({"ok": False, "error": "Variant not found"}, status=404)
                product = variant.product
            else:
                product = Product.objects.filter(sku=sku).first()
                if not product:
                    return JsonResponse({"ok": False, "error": "Product not found"}, status=404)

            allow_oversell = bool(getattr(product, "selling_while_out_of_stock", False))
            stock = variant.variant_quantity if is_variant else product.quantity
            max_qty = None
            if stock is not None and not allow_oversell:
                max_qty = float(qty) + float(stock)
                
            unit_cost = variant.variant_cost if is_variant else product.cost

            return JsonResponse({
                "ok": True,
                "item": {
                    "product_title": product.title or "",
                    "product_url": reverse("marketing:product_detail", args=[product.pk]),
                    "product_sku": product.sku or "",
                    "variant_sku": variant.variant_sku if variant else "",
                    "image_url": (product.files.first().file_url
                                  if product.files.exists() and product.files.first().file_url else ""),
                    "quantity": float(qty),
                    "price": float(price),
                    "subtotal": float(qty * price),
                    "unit_cost": float(unit_cost or 0),
                    "stock": float(stock) if stock is not None else None,
                    "max_qty": max_qty,
                    "allow_oversell": allow_oversell,
                }
            })

        # ── Add item ────────────────────────────────────────────
        if action == "add_item":
            sku = (request.POST.get("sku") or "").strip()
            is_variant = (request.POST.get("is_variant") or "").lower() == "true"
            qty_raw = (request.POST.get("quantity") or "1").replace(",", ".")
            price_raw = (request.POST.get("price") or "0").replace(",", ".")
            if not sku:
                return JsonResponse({"ok": False, "error": "Missing SKU"}, status=400)
            try:
                qty = Decimal(qty_raw)
                price = Decimal(price_raw)
            except (InvalidOperation, ValueError):
                return JsonResponse({"ok": False, "error": "Invalid number"}, status=400)
            if qty <= 0 or price < 0:
                return JsonResponse({"ok": False, "error": "Bad qty/price"}, status=400)

            variant = None
            if is_variant:
                variant = ProductVariant.objects.filter(variant_sku=sku).select_related("product").first()
                if not variant:
                    return JsonResponse({"ok": False, "error": "Variant not found"}, status=404)
                product = variant.product
            else:
                product = Product.objects.filter(sku=sku).first()
                if not product:
                    return JsonResponse({"ok": False, "error": "Product not found"}, status=404)

            # Back-order is allowed — adding a product whose stock is
            # below the requested qty is fine. Stock isn't touched until
            # the order ships (see operating.signals).
            allow_oversell = bool(getattr(product, "selling_while_out_of_stock", False))

            item = OrderItem.objects.create(
                order=order, product=product, product_variant=variant,
                quantity=qty, price=price,
            )
            _sync_cari_total()

            # Refresh stock for the response so the new row shows the
            # post-deduction figure right away.
            if variant:
                variant.refresh_from_db(fields=["variant_quantity"])
                new_stock = variant.variant_quantity
            else:
                product.refresh_from_db(fields=["quantity"])
                new_stock = product.quantity
            max_qty = None
            if new_stock is not None and not allow_oversell:
                max_qty = float(qty) + float(new_stock)

            return JsonResponse({
                "ok": True,
                "item": {
                    "id": item.pk,
                    "product_title": product.title or "",
                    "product_url": reverse("marketing:product_detail", args=[product.pk]),
                    "product_sku": product.sku or "",
                    "variant_sku": variant.variant_sku if variant else "",
                    "image_url": (product.files.first().file_url
                                  if product.files.exists() and product.files.first().file_url else ""),
                    "quantity": float(qty),
                    "price": float(price),
                    "subtotal": float(item.subtotal()),
                    "unit_cost": float(item.unit_cost() or 0),
                    "stock": float(new_stock) if new_stock is not None else None,
                    "max_qty": max_qty,
                    "allow_oversell": allow_oversell,
                },
                "order_total": float(order.total_value()),
                "cari": _cari_snapshot(),
                "profit": _profit_snapshot(),
            })

        # ── Remove item ─────────────────────────────────────────
        if action == "remove_item":
            try:
                item_id = int(request.POST.get("item_id") or 0)
            except (TypeError, ValueError):
                return JsonResponse({"ok": False, "error": "Bad item_id"}, status=400)
            item = get_object_or_404(OrderItem, pk=item_id, order=order)
            item.delete()
            _sync_cari_total()
            return JsonResponse({
                "ok": True,
                "order_total": float(order.total_value()),
                "cari": _cari_snapshot(),
                "profit": _profit_snapshot(),
            })

        # ── Update customer ─────────────────────────────────────
        if action == "update_customer":
            ctype = (request.POST.get("customer_type") or "").strip()
            try:
                cpk = int(request.POST.get("customer_pk") or 0)
            except (TypeError, ValueError):
                cpk = 0

            if ctype == "contact" and cpk:
                c = get_object_or_404(Contact, pk=cpk)
                order.contact = c
                order.company = None
            elif ctype == "company" and cpk:
                co = get_object_or_404(Company, pk=cpk)
                order.company = co
                order.contact = None
            elif ctype == "none":
                order.contact = None
                order.company = None
            else:
                return JsonResponse({"ok": False, "error": "Bad customer"}, status=400)
            order.save(update_fields=["contact", "company", "updated_at"])

            # Customer changed → re-resolve the cari and move the ledger
            # row to the new account. The old movement (if any) is reversed
            # before posting fresh to avoid orphaning balance on the old cari.
            try:
                new_cari = get_or_create_cari_for_order(order, member=member)
                if order.cari_id and (not new_cari or new_cari.pk != order.cari_id):
                    reverse_order_movement(order)
                if new_cari and order.cari_id != new_cari.pk:
                    order.cari = new_cari
                    order.save(update_fields=["cari", "updated_at"])
                if new_cari:
                    post_order_movement(order, member=member)
                elif order.cari_id:
                    # Customer cleared → drop the cari link entirely.
                    reverse_order_movement(order)
                    order.cari = None
                    order.save(update_fields=["cari", "updated_at"])
            except Exception:
                pass
            return JsonResponse({"ok": True})

        # ── Update guest info ───────────────────────────────────
        if action == "update_guest":
            order.guest_first_name = (request.POST.get("guest_first_name") or "").strip() or None
            order.guest_last_name = (request.POST.get("guest_last_name") or "").strip() or None
            order.guest_email = (request.POST.get("guest_email") or "").strip() or None
            order.guest_phone = (request.POST.get("guest_phone") or "").strip() or None
            order.save(update_fields=["guest_first_name", "guest_last_name",
                                       "guest_email", "guest_phone", "updated_at"])
            return JsonResponse({"ok": True})

        if action != "update_status":
            return self.get(request, *args, **kwargs)

        new_status = (request.POST.get("order_status") or "").strip()
        carrier = request.POST.get("carrier")
        tracking = request.POST.get("tracking_number")

        # All status changes funnel through the single atomic helper so
        # the ship gate (cargo required) + the reservation → stock-out
        # conversion can never be bypassed or partially applied.
        from .views_warehouse import apply_order_status_change
        ok, code = apply_order_status_change(
            order, new_status, carrier=carrier, tracking=tracking, user=request.user,
        )
        if not ok:
            if code == "cargo_required":
                messages.error(request, "Siparişi tamamlamak (Gönderildi) için kargo şirketi ve takip numarası gerekli.")
            elif code == "packing_requires_scan":
                messages.error(request, "Paketleniyor'a geçmek için önce en az bir top okutun.")
            else:
                messages.error(request, f"Sipariş güncellenemedi: {(code or '').replace('error:', '')}")
            return redirect("operating:order_detail", pk=order.pk)
        messages.success(request, "Order updated.")
        return redirect("operating:order_detail", pk=order.pk)


# ---------------------------------------------------------------------------
# Packing-scan flow — reserve warehouse rolls for an order.
#
# The middle stage of the 3-step lifecycle (Açık → Paketleniyor →
# Gönderildi). Staff open the packing page and scan the barcodes of the
# physical rolls (tops) being packed. Each scanned roll must belong to
# one of the order's products (else it's rejected) and reserves its
# metres — editable, so a partial cut reserves less and leaves the rest
# on the roll. Reservations are a SOFT hold: NO stock is physically cut
# until the order ships. The page is re-openable — reservations can be
# added, edited or removed any time before shipping.
# ---------------------------------------------------------------------------
from decimal import Decimal as _PDecimal

_SHIPPED_CLASS = {"shipped", "in_transit", "out_for_delivery", "delivered"}


def _order_item_match_maps(order):
    """Lookup sets from an order's items so we can decide whether a
    scanned roll belongs to this order."""
    variant_ids, product_ids, skus = set(), set(), set()
    items = list(order.items.all().select_related("product", "product_variant"))
    for it in items:
        if it.product_variant_id:
            variant_ids.add(it.product_variant_id)
            vs = getattr(it.product_variant, "variant_sku", None)
            if vs:
                skus.add(vs.strip().lower())
        if it.product_id:
            product_ids.add(it.product_id)
            ps = getattr(it.product, "sku", None)
            if ps:
                skus.add(ps.strip().lower())
    return items, variant_ids, product_ids, skus


def _match_roll_to_order(roll, items, variant_ids, product_ids, skus):
    """Return the best-matching OrderItem for a scanned roll, or None if
    the roll's product is not part of the order. Precedence:
    catalog variant → catalog product → SKU."""
    wp = roll.product
    if wp is None:
        return None
    cv_id = wp.catalog_variant_id
    cv = wp.catalog_variant if cv_id else None
    cv_product_id = cv.product_id if cv else None
    wp_sku = (wp.sku or "").strip().lower()

    def _pick(pred):
        for it in items:
            if pred(it):
                return it
        return None

    if cv_id and cv_id in variant_ids:
        hit = _pick(lambda it: it.product_variant_id == cv_id)
        if hit:
            return hit
    if cv_product_id and cv_product_id in product_ids:
        # Only bind the parent-product to a VARIANT-LESS order line. A
        # line that names a specific variant must match at the variant
        # level (branch above) — otherwise a sibling variant's roll
        # (same base product) would wrongly satisfy it.
        hit = _pick(lambda it: it.product_id == cv_product_id and it.product_variant_id is None)
        if hit:
            return hit
    if wp_sku and wp_sku in skus:
        hit = _pick(lambda it: (
            (it.product_variant and (getattr(it.product_variant, "variant_sku", "") or "").strip().lower() == wp_sku)
            or (it.product and (getattr(it.product, "sku", "") or "").strip().lower() == wp_sku)
        ))
        if hit:
            return hit
        return items[0] if items else None  # sku matched but couldn't bind a specific line
    return None


def _roll_available_meters(roll, exclude_reservation_id=None):
    """Physical metres on the roll minus what OTHER active reservations
    already hold — the ceiling so a roll can't be over-committed."""
    from .models import OrderRollReservation
    from django.db.models import Sum
    phys = roll.meters_remaining if roll.meters_remaining is not None else roll.meters
    phys = phys or _PDecimal("0")
    other = (OrderRollReservation.objects
             .filter(roll=roll, consumed=False)
             .exclude(pk=exclude_reservation_id)
             .aggregate(s=Sum("meters"))["s"] or _PDecimal("0"))
    avail = phys - other
    return avail if avail > 0 else _PDecimal("0")


def _reservation_payload(r):
    wp = r.warehouse_product
    return {
        "id": r.id,
        "roll_id": r.roll_id,
        "barcode": (r.roll.barcode if r.roll else None),
        "product_id": r.warehouse_product_id,
        "product_name": (wp.name if wp else ""),
        "sku": (wp.sku if wp else ""),
        "warehouse": (wp.warehouse.name if (wp and wp.warehouse_id) else ""),
        "meters": float(r.meters or 0),
        "roll_meters": (float(r.roll.meters or 0) if r.roll else None),
        "roll_remaining": (float(r.roll.meters_remaining) if (r.roll and r.roll.meters_remaining is not None) else None),
        "consumed": r.consumed,
    }


@login_required
def order_pack_scan(request, pk):
    """The packing page for an order — item checklist + roll scanner +
    live list of reserved rolls."""
    order = get_object_or_404(Order, pk=pk)
    reservations = list(
        order.roll_reservations
        .select_related("roll", "warehouse_product", "warehouse_product__warehouse")
        .order_by("-created_at")
    )
    items = list(order.items.all().select_related("product", "product_variant"))
    return render(request, "operating/order_pack_scan.html", {
        "order": order,
        "items": items,
        "reservations": reservations,
        "reservation_count": sum(1 for r in reservations if not r.consumed),
        "shipped": order.order_status in _SHIPPED_CLASS,
    })


@login_required
@require_POST
def order_pack_reserve_add(request, pk):
    """Scan/enter a roll barcode → reserve it for this order. Rejects
    rolls whose product is not part of the order."""
    from .models import WarehouseProductRoll, OrderRollReservation
    order = get_object_or_404(Order, pk=pk)
    if order.order_status in _SHIPPED_CLASS:
        return JsonResponse({"ok": False, "error": "Sipariş gönderildi — paketleme kilitli."}, status=400)

    code = (request.POST.get("barcode") or "").strip()
    if not code:
        return JsonResponse({"ok": False, "error": "Barkod boş."}, status=400)

    items, variant_ids, product_ids, skus = _order_item_match_maps(order)
    if not items:
        return JsonResponse({"ok": False, "error": "Siparişte ürün yok."}, status=400)

    rolls = list(WarehouseProductRoll.objects
                 .select_related("product", "product__catalog_variant", "product__warehouse")
                 .filter(barcode__iexact=code))
    if not rolls:
        return JsonResponse({"ok": False, "kind": "not_found",
                             "error": "Bu barkodla bir top (roll) bulunamadı."}, status=404)

    # All rolls sharing this barcode whose product belongs to the order
    # (a barcode is code-unique but not DB-unique, so there can be more
    # than one across warehouses).
    matched = []
    for roll in rolls:
        mi = _match_roll_to_order(roll, items, variant_ids, product_ids, skus)
        if mi is not None:
            matched.append((roll, mi))
    if not matched:
        return JsonResponse({
            "ok": False, "kind": "wrong_product",
            "error": "Bu top bu siparişteki ürünlere ait değil — eklenmedi.",
            "product_name": (rolls[0].product.name if rolls[0].product else ""),
        }, status=409)

    # Already reserved for this order? Re-scan is a no-op (no duplicate).
    for roll, _mi in matched:
        existing = OrderRollReservation.objects.filter(order=order, roll=roll, consumed=False).first()
        if existing:
            return JsonResponse({"ok": True, "duplicate": True,
                                 "reservation": _reservation_payload(existing)})

    # Prefer the first matching roll that still has reservable metres.
    pick = next(((roll, mi) for (roll, mi) in matched if _roll_available_meters(roll) > 0), None)
    if pick is None:
        return JsonResponse({"ok": False, "kind": "no_stock",
                             "error": "Bu topta uygun/rezerv edilebilir metre kalmadı."}, status=409)
    roll_pick, matched_item = pick
    wp = roll_pick.product

    # Parse the requested metres (validation errors before we lock).
    raw = (request.POST.get("meters") or "").strip()
    req_meters = None
    if raw:
        try:
            req_meters = _PDecimal(raw)
        except Exception:
            return JsonResponse({"ok": False, "error": "Geçersiz metre."}, status=400)
        if req_meters <= 0:
            return JsonResponse({"ok": False, "error": "Metre sıfırdan büyük olmalı."}, status=400)

    # Lock the roll row and re-check availability inside the lock so two
    # concurrent scans of the same roll can't over-commit it.
    from django.db import transaction as _tx
    capped = False
    with _tx.atomic():
        locked = WarehouseProductRoll.objects.select_for_update().get(pk=roll_pick.pk)
        avail = _roll_available_meters(locked)
        if avail <= 0:
            return JsonResponse({"ok": False, "kind": "no_stock",
                                 "error": "Bu topta uygun/rezerv edilebilir metre kalmadı."}, status=409)
        if req_meters is not None:
            meters = req_meters
            if meters > avail:
                meters = avail
                capped = True
        else:
            meters = avail
        r = OrderRollReservation.objects.create(
            order=order, order_item=matched_item, roll=locked, warehouse_product=wp,
            meters=meters, created_by=request.user if request.user.is_authenticated else None,
        )
    return JsonResponse({"ok": True, "capped": capped, "reservation": _reservation_payload(r)})


@login_required
@require_POST
def order_pack_reserve_update(request, pk):
    """Edit the reserved metres of one reservation (partial cut)."""
    from .models import OrderRollReservation
    order = get_object_or_404(Order, pk=pk)
    if order.order_status in _SHIPPED_CLASS:
        return JsonResponse({"ok": False, "error": "Sipariş gönderildi — paketleme kilitli."}, status=400)
    rid = (request.POST.get("reservation_id") or "").strip()
    if not rid.isdigit():
        return JsonResponse({"ok": False, "error": "Geçersiz rezervasyon."}, status=400)
    r = OrderRollReservation.objects.filter(pk=rid, order=order, consumed=False).first()
    if r is None:
        return JsonResponse({"ok": False, "error": "Rezervasyon bulunamadı."}, status=404)
    raw = (request.POST.get("meters") or "").strip()
    try:
        meters = _PDecimal(raw)
    except Exception:
        return JsonResponse({"ok": False, "error": "Geçersiz metre."}, status=400)
    if meters <= 0:
        return JsonResponse({"ok": False, "error": "Metre sıfırdan büyük olmalı."}, status=400)
    avail = _roll_available_meters(r.roll, exclude_reservation_id=r.pk)
    capped = meters > avail
    if capped:
        meters = avail
    r.meters = meters
    r.save(update_fields=["meters"])
    return JsonResponse({"ok": True, "capped": capped, "reservation": _reservation_payload(r)})


@login_required
@require_POST
def order_pack_reserve_remove(request, pk):
    """Release a reservation (remove a scanned roll) before shipping."""
    from .models import OrderRollReservation
    order = get_object_or_404(Order, pk=pk)
    if order.order_status in _SHIPPED_CLASS:
        return JsonResponse({"ok": False, "error": "Sipariş gönderildi — paketleme kilitli."}, status=400)
    rid = (request.POST.get("reservation_id") or "").strip()
    if not rid.isdigit():
        return JsonResponse({"ok": False, "error": "Geçersiz rezervasyon."}, status=400)
    r = OrderRollReservation.objects.filter(pk=rid, order=order, consumed=False).first()
    if r is None:
        return JsonResponse({"ok": False, "error": "Rezervasyon bulunamadı."}, status=404)
    removed = r.id
    r.delete()
    return JsonResponse({"ok": True, "removed": removed})


@login_required
@require_POST
def order_pack_complete(request, pk):
    """Confirm packing — needs at least one reserved roll. Moves the
    order to 'packaging' (Paketleniyor). Re-runnable; no stock cut."""
    order = get_object_or_404(Order, pk=pk)
    if order.order_status in _SHIPPED_CLASS:
        return JsonResponse({"ok": False, "error": "Sipariş zaten gönderildi."}, status=400)
    n = order.roll_reservations.filter(consumed=False).count()
    if n < 1:
        return JsonResponse({"ok": False,
                             "error": "Paketlemeyi tamamlamak için en az bir top okutmalısınız."}, status=400)
    if order.order_status != "packaging":
        order.order_status = "packaging"
        order.save(update_fields=["order_status", "updated_at"])
    return JsonResponse({"ok": True, "status": "packaging",
                         "redirect": reverse("operating:order_detail", args=[order.pk])})


# ---------------------------------------------------------------------------
# Order change history — the audit trail written by operating/audit.py.
# Preview card on the detail page + a dedicated filterable page.
# ---------------------------------------------------------------------------
_CHANGE_FIELD_TR = {
    "order_status": "Sipariş durumu", "carrier": "Kargo şirketi",
    "tracking_number": "Takip numarası", "notes": "Notlar",
    "print_header": "Yazdırma başlığı", "ettn": "ETTN",
    "guest_first_name": "Misafir adı", "guest_last_name": "Misafir soyadı",
    "guest_email": "Misafir e-posta", "guest_phone": "Misafir telefon",
    "customer": "Müşteri", "quantity": "Miktar", "price": "Fiyat",
    "product": "Ürün",
}
_CHANGE_ACTION_TR = {
    "created": "Oluşturuldu", "status": "Durum",
    "item_added": "Ürün eklendi", "item_removed": "Ürün çıkarıldı",
    "item_updated": "Ürün güncellendi", "field": "Bilgi güncellendi",
}


def _decorate_order_changes(changes):
    """Attach display attrs (Turkish labels, resolved status/carrier
    names) to OrderChange rows for the templates."""
    from django.utils import translation
    from .models import ORDER_STATUS_CHOICES, CARRIER_CHOICES
    is_tr = (translation.get_language() or "").startswith("tr")
    status_map = {k: str(v) for k, v in ORDER_STATUS_CHOICES}
    carrier_map = dict(CARRIER_CHOICES)

    def val_display(c, v):
        if v is None:
            return None
        if c.field == "order_status":
            return status_map.get(v, v)
        if c.field == "carrier":
            return carrier_map.get(v, v)
        return v

    for c in changes:
        c.action_display = (_CHANGE_ACTION_TR.get(c.action, c.action) if is_tr
                            else c.get_action_display())
        if c.field:
            c.field_display = (_CHANGE_FIELD_TR.get(c.field, c.field) if is_tr
                               else c.field.replace("_", " "))
        else:
            c.field_display = None
        c.old_display = val_display(c, c.old_value)
        c.new_display = val_display(c, c.new_value)
    return changes


@login_required
def order_changes(request, pk):
    """Dedicated filterable page for an order's full change history.
    Query params: action ('' | created | status | item_added |
    item_removed | item_updated | field), from/to (YYYY-MM-DD), q, page."""
    from datetime import datetime as _dt
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Q as _Q
    from .models import OrderChange

    order = get_object_or_404(Order, pk=pk)
    qs = (OrderChange.objects.filter(order=order)
          .select_related("created_by"))

    action = (request.GET.get("action") or "").strip()
    if action in {k for k, _ in OrderChange.ACTION_CHOICES}:
        qs = qs.filter(action=action)

    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()
    try:
        if date_from:
            qs = qs.filter(created_at__date__gte=_dt.strptime(date_from, "%Y-%m-%d").date())
    except ValueError:
        pass
    try:
        if date_to:
            qs = qs.filter(created_at__date__lte=_dt.strptime(date_to, "%Y-%m-%d").date())
    except ValueError:
        pass

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(_Q(item_label__icontains=q) | _Q(field__icontains=q)
                       | _Q(old_value__icontains=q) | _Q(new_value__icontains=q)
                       | _Q(created_by__username__icontains=q))

    qs = qs.order_by("-created_at", "-id")

    paginator = Paginator(qs, 50)
    try:
        page = paginator.page(request.GET.get("page", 1))
    except (EmptyPage, PageNotAnInteger, ValueError):
        page = paginator.page(1)

    qd = request.GET.copy()
    qd.pop("page", None)

    return render(request, "operating/order_changes.html", {
        "order": order,
        "page_obj": page,
        "paginator": paginator,
        "changes": _decorate_order_changes(list(page.object_list)),
        "filter_qs": qd.urlencode(),
        "filters": {"action": action, "from": date_from, "to": date_to, "q": q},
    })


# ---------------------------------------------------------------------------
# Customer card partial — full HTMX flow.
# Single endpoint that handles GET (refresh) + POST (mutations) and always
# returns the rendered card HTML so HTMX can swap it into the page without
# any client-side reload.
# ---------------------------------------------------------------------------
def order_customer_card_view(request, pk):
    from .models import Order
    from crm.models import Contact, Company
    from current_account.services import (
        get_or_create_cari_for_order, post_order_movement, reverse_order_movement,
    )

    order = get_object_or_404(Order, pk=pk)
    member = getattr(request.user, "member", None)

    def _save_cari_link(prev_cari_id):
        """After contact/company on the order changed, re-resolve the cari
        and move the order_sale movement to the new account."""
        try:
            new_cari = get_or_create_cari_for_order(order, member=member)
            # If the cari is changing, reverse the old movement first.
            if prev_cari_id and (not new_cari or new_cari.pk != prev_cari_id):
                reverse_order_movement(order)
            if new_cari and order.cari_id != new_cari.pk:
                order.cari = new_cari
                order.save(update_fields=["cari", "updated_at"])
            if new_cari:
                post_order_movement(order, member=member)
            elif order.cari_id:
                reverse_order_movement(order)
                order.cari = None
                order.save(update_fields=["cari", "updated_at"])
        except Exception:
            pass

    mode = "view"

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        # ── Switch to an existing customer ───────────────────
        if action == "switch":
            ctype = (request.POST.get("customer_type") or "").strip()
            try:
                cpk = int(request.POST.get("customer_pk") or 0)
            except (TypeError, ValueError):
                cpk = 0
            prev_cari_id = order.cari_id
            fields_changed = ["contact", "company", "web_client",
                              "is_guest_order", "updated_at"]
            # Clear all customer linkages first so the swap is atomic.
            order.contact = None
            order.company = None
            order.web_client = None
            order.is_guest_order = False
            if ctype == "contact" and cpk:
                order.contact = get_object_or_404(Contact, pk=cpk)
            elif ctype == "company" and cpk:
                order.company = get_object_or_404(Company, pk=cpk)
            elif ctype == "web_client" and cpk:
                from authentication.models import WebClient
                order.web_client = get_object_or_404(WebClient, pk=cpk)
            elif ctype == "guest":
                order.is_guest_order = True
            order.save(update_fields=fields_changed)
            _save_cari_link(prev_cari_id)
            mode = "edit"

        # ── Clear linked customer ────────────────────────────
        elif action == "clear":
            prev_cari_id = order.cari_id
            order.contact = None
            order.company = None
            order.save(update_fields=["contact", "company", "updated_at"])
            _save_cari_link(prev_cari_id)
            mode = "edit"

        # ── Inline field edit on the linked contact/company ──
        elif action == "update_field":
            from django.contrib.postgres.fields import ArrayField as _ArrayField
            field = (request.POST.get("field") or "").strip()
            value = (request.POST.get("value") or "").strip()
            target = order.contact or order.company
            allowed = {
                "contact": {"name", "job_title", "email", "phone", "address", "country"},
                "company": {"name", "email", "phone", "address", "country", "website"},
            }
            if target:
                kind = "contact" if order.contact_id else "company"
                if field in allowed[kind]:
                    # email / phone are ArrayField on both models — store as
                    # a single-item list so the existing display logic shows it.
                    try:
                        model_field = target.__class__._meta.get_field(field)
                        if isinstance(model_field, _ArrayField):
                            setattr(target, field, [value] if value else [])
                        else:
                            setattr(target, field, value)
                        target.save(update_fields=[field])
                    except Exception:
                        pass
                    # If the contact's name changed AND the cari was named
                    # after it, keep the cari label in sync. Safe / cheap.
                    if field == "name" and order.cari_id:
                        try:
                            order.cari.name = value or order.cari.name
                            order.cari.save(update_fields=["name"])
                        except Exception:
                            pass
            mode = "edit"

        # ── Inline delivery-address edit (order-level) ──────
        elif action == "update_delivery":
            field = (request.POST.get("field") or "").strip()
            value = (request.POST.get("value") or "").strip()
            allowed = {"delivery_address_title", "delivery_address",
                       "delivery_city", "delivery_country", "delivery_phone"}
            if field in allowed:
                setattr(order, field, value)
                order.save(update_fields=[field, "updated_at"])
            mode = "edit"

        # ── Inline guest field edit ──────────────────────────
        elif action == "update_guest":
            field = (request.POST.get("field") or "").strip()
            value = (request.POST.get(field) or "").strip() or None
            if field in {"guest_first_name", "guest_last_name",
                          "guest_email", "guest_phone"}:
                setattr(order, field, value)
                order.save(update_fields=[field, "updated_at"])
            mode = "edit"

        # ── Create brand-new Contact + link + cari ───────────
        elif action == "create_contact":
            name = (request.POST.get("name") or "").strip()
            if not name:
                mode = "create"
            else:
                prev_cari_id = order.cari_id
                email_val = (request.POST.get("email") or "").strip()
                phone_val = (request.POST.get("phone") or "").strip()
                new_contact = Contact.objects.create(
                    name=name,
                    email=[email_val] if email_val else [],
                    phone=[phone_val] if phone_val else [],
                    address=(request.POST.get("address") or "").strip(),
                    country=(request.POST.get("country") or "").strip(),
                )
                order.contact = new_contact
                order.company = None
                order.save(update_fields=["contact", "company", "updated_at"])
                _save_cari_link(prev_cari_id)
                mode = "view"

    elif request.method == "GET":
        mode = request.GET.get("mode") or "view"

    # Re-fetch order to ensure related .cari, .contact etc. are fresh
    # for the template (the snapshot used by HTMX swap is independent of
    # whatever client mode toggle was in effect when the form fired).
    order = get_object_or_404(Order, pk=pk)
    return render(request, "operating/partials/_customer_card.html", {
        "order": order,
        "cust_card_mode": mode,
    })


@require_POST
def update_order_print_header(request, pk):
    """Save the custom top-left header text shown on the order's printable PDF.
    Empty clears it (falls back to the global brand)."""
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "auth"}, status=403)
    order = get_object_or_404(Order, pk=pk)
    header = (request.POST.get("print_header") or "").strip()[:120]
    order.print_header = header or None
    order.save(update_fields=["print_header", "updated_at"])
    return JsonResponse({"success": True, "print_header": order.print_header or ""})


class OrderPrint(DetailView):
    """Printable order view.

    Serves the order as styled HTML and lets the BROWSER render it and
    "Save as PDF". We deliberately do NOT use xhtml2pdf here: it cannot
    render this card/badge/two-column layout cleanly (it scatters the
    blocks down the page and spills onto a second sheet), whereas the
    browser reproduces the on-screen design pixel-perfect. The template
    auto-opens the print dialog when is_pdf is False."""
    model = Order
    template_name = "operating/order_print.html"
    context_object_name = "order"

    def get_context_data(self, **kwargs):
        from decimal import Decimal
        ctx = super().get_context_data(**kwargs)
        order = self.object
        items = []
        total = Decimal("0.00")
        for it in order.items.all().select_related("product", "product_variant"):
            qty = it.quantity or Decimal("0")
            price = it.price or Decimal("0")
            line_total = (qty * price).quantize(Decimal("0.01"))
            it.line_total_calc = line_total
            items.append(it)
            total += line_total
        ctx["order_items"] = items
        ctx["order_total"] = total
        ctx["is_pdf"] = True   # template can strip JS auto-print when rendering for PDF
        return ctx

    def render_to_response(self, context, **response_kwargs):
        from django.template.loader import render_to_string
        from django.http import HttpResponse

        # Render the printable HTML and let the browser handle PDF output
        # (Print → Save as PDF). is_pdf=False triggers the auto-print
        # dialog and keeps the page looking exactly like the on-screen
        # design — no xhtml2pdf re-layout that breaks the cards/badges.
        ctx = dict(context)
        ctx["is_pdf"] = False
        html = render_to_string(self.template_name, ctx, request=self.request)
        return HttpResponse(html)


class OrderCreate(View):
    def get(self, request):
        form = OrderForm()
        if request.headers.get("HX-Request"):
            # Return partial for sidebar
            # We need to manually construct formset if needed, but OrderForm usually handles basics?
            # Wait, create_order.html uses {{ formset.management_form }}.
            # OrderCreate view currently DOES NOT pass a formset in GET context (see original code).
            # Original code: creates OrderForm(), passes {"form": form}.
            # But template references {{ formset.management_form }}. This implies formset might be missing or empty?
            # Let's check post method: commented out # order_item_formset = OrderItemFormSet(request.POST)
            # It seems the formset is handled via JS building JSON, so management_form might be unused or expected to be manually added?
            # In create_order.html, {{ formset.management_form }} is used.
            # If I don't pass 'formset', {{ formset.management_form }} renders nothing (if variable missing).
            # So I will just pass form.
            return render(request, "operating/partials/create_order_form.html", {"form": form})
            
        return render(
            request,
            "operating/create_order.html",
            {"form": form},
        )

    def post(self, request):
        form = OrderForm(request.POST)

        customer_pk = request.POST.get("customer_pk")
        customer_type = request.POST.get("customer_type")
        product_json_input = request.POST.get("product_json_input")
        if product_json_input:
            try:
                product_json_input = json.loads(product_json_input)
            except json.JSONDecodeError:
                messages.error(request, "Invalid product data format.")
                return render(request, "operating/create_order.html", {"form": form})

        # Stock validation removed — orders can be placed for products
        # that exceed current stock (back-order). Catalog stock isn't
        # deducted until the order moves into a fulfilment status
        # (shipped / delivered), so insufficient stock at create time
        # is fine. UI shows an amber back-order chip as a heads-up.

        if form.is_valid():
            try:
                with transaction.atomic():
                    order = form.save(commit=False)

                    # Customer
                    customer_source = request.POST.get("customer_source", "crm")
                    if customer_source == "crm":
                        if customer_type == "contact" and customer_pk:
                            try:
                                order.contact = Contact.objects.get(pk=customer_pk)
                            except Contact.DoesNotExist:
                                pass
                        elif customer_type == "company" and customer_pk:
                            try:
                                order.company = Company.objects.get(pk=customer_pk)
                            except Company.DoesNotExist:
                                pass
                    elif customer_source == "web":
                        webclient_pk = request.POST.get("webclient_pk")
                        if webclient_pk:
                            from authentication.models import WebClient
                            try:
                                order.web_client = WebClient.objects.get(pk=webclient_pk)
                            except WebClient.DoesNotExist:
                                pass
                    elif customer_source == "guest":
                        order.is_guest_order = True
                        order.guest_first_name = request.POST.get("guest_first_name", "")
                        order.guest_last_name = request.POST.get("guest_last_name", "")
                        order.guest_email = request.POST.get("guest_email", "")
                        order.guest_phone = request.POST.get("guest_phone", "")

                    # Retail (Perakende) flag — Step 1's order_type radio.
                    # Persisted so the order list can label it, unlike
                    # customer_source it isn't re-derivable after save.
                    order.is_retail_order = request.POST.get("order_type") == "perakende"

                    # Notification opt-in — staff ticked "Send customer
                    # emails for this order" in the create sidebar.
                    order.notify_customer = bool(request.POST.get("notify_customer"))

                    # Payment method / amount-paid fields removed from
                    # the form — deposit is handled separately below as
                    # a proper cari collection. We leave the fields on
                    # the model alone (they stay NULL for new orders).

                    # Delivery Address
                    order.delivery_address_title = request.POST.get("delivery_address_title", "")
                    order.delivery_address = request.POST.get("delivery_address", "")
                    order.delivery_city = request.POST.get("delivery_city", "")
                    order.delivery_country = request.POST.get("delivery_country", "")
                    order.delivery_phone = request.POST.get("delivery_phone", "")

                    order.save()

                    # Order Items
                    for item_data in product_json_input:
                        sku = item_data["product"]["sku"]
                        is_variant = item_data["product"]["variant"]
                        is_custom = item_data.get("is_custom_curtain", False)

                        if is_variant:
                            variant = get_object_or_404(ProductVariant, variant_sku=sku)
                            product = variant.product
                        else:
                            product = get_object_or_404(Product, sku=sku)
                            variant = None

                        order_item = OrderItem.objects.create(
                            order=order,
                            product=product,
                            product_variant=variant,
                            description=item_data.get("description", ""),
                            quantity=item_data.get("quantity", 1),
                            price=item_data.get("price", 0),
                            is_custom_curtain=is_custom,
                        )

                        # Custom curtain fields
                        if is_custom:
                            order_item.custom_width = item_data.get("custom_width") or None
                            order_item.custom_height = item_data.get("custom_height") or None
                            order_item.custom_pleat_type = item_data.get("custom_pleat_type") or None
                            order_item.custom_pleat_density = item_data.get("custom_pleat_density") or None
                            order_item.custom_mounting_type = item_data.get("custom_mounting_type") or None
                            order_item.custom_wing_type = item_data.get("custom_wing_type") or None
                            order_item.custom_fabric_used_meters = item_data.get("custom_fabric_used_meters") or None
                            order_item.save()

                    # Generate QR code
                    generate_machine_qr_for_order(order)

                # ── NO warehouse stock-out at create ──────────────────
                # New orders start "Açık" (Open) and reserve/deduct
                # nothing. Warehouse rolls are reserved during the
                # packing-scan step and only physically cut when the
                # order is shipped (cargo info entered). See
                # views_warehouse.consume_reservations_for_order.

                # ── Auto-link to a CariAccount + log the sales-order
                # movement so the customer's ledger reflects this
                # order immediately. Web orders are not routed through
                # this view, so we don't gate on order kind here.
                try:
                    from current_account.services import (
                        get_or_create_cari_for_order, post_order_movement,
                    )
                    member = getattr(request.user, "member", None)
                    cari = get_or_create_cari_for_order(order, member=member)
                    if cari and order.cari_id != cari.pk:
                        order.cari = cari
                        order.save(update_fields=["cari"])
                    post_order_movement(order, member=member)
                except Exception as _e:
                    messages.warning(request, f"Order saved but cari sync had an issue: {_e}")

                # ── Deposit (collection) — post against the cari if the
                # user ticked "Deposit received" and entered an amount.
                # Creates + confirms a Payment so it lands in the cari
                # ledger as a deduction.
                deposit_flag = (request.POST.get("deposit_received") or "").strip()
                try:
                    deposit_amount = float(request.POST.get("deposit_amount") or 0)
                except (ValueError, TypeError):
                    deposit_amount = 0
                if deposit_flag and deposit_amount > 0 and order.cari_id:
                    try:
                        from decimal import Decimal
                        from datetime import date
                        from current_account.models import Payment
                        from current_account.views_payment import _next_payment_number
                        from current_account.services import get_default_book, _resolve_currency

                        book = get_default_book()
                        currency = _resolve_currency(order)
                        pay = Payment.objects.create(
                            cari=order.cari, book=book,
                            number=_next_payment_number(book, "collection"),
                            type="collection", method="cash", status="draft",
                            date=date.today(),
                            amount=Decimal(str(deposit_amount)),
                            currency=currency,
                            description=f"Deposit for Order #{order.pk}",
                            notes=f"ORD-{order.pk}",
                            created_by=member,
                        )
                        pay.confirm(user=request.user if request.user.is_authenticated else None)
                    except Exception as _e:
                        messages.warning(request, f"Order saved but deposit posting failed: {_e}")

                # ── Customer notification email ───────────────────────
                # Best-effort — never blocks the order save. Sends only
                # when notify_customer is True (staff opt-in) AND the
                # order has a usable customer email.
                try:
                    if order.notify_customer:
                        from .order_notifications import send_order_event_email
                        send_order_event_email(order, "created", attach_pdf=True)
                except Exception as _e:
                    messages.warning(request, f"Order saved but confirmation email failed: {_e}")

                return redirect("operating:order_detail", pk=order.pk)
            except Exception as e:
                messages.error(request, f"Order creation failed: {e}")
                return render(request, "operating/create_order.html", {"form": form})


class OrderEdit(UpdateView):
    model = Order
    form_class = OrderForm
    template_name = "operating/edit_order.html"

    def _is_ajax(self):
        return self.request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def get_template_names(self):
        # When opened in the side panel, reuse the SAME partial as the
        # create-order sidebar so create and edit feel identical. The
        # partial branches on `order` in context to switch action URL,
        # submit label, and to hydrate existing items / customer.
        if self._is_ajax():
            return ["operating/partials/create_order_form.html"]
        return [self.template_name]

    # prevent editing completed orders.
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status == "completed":
            messages.error(request, "Completed orders cannot be edited.")
            return HttpResponseForbidden("You cannot edit a completed order.")
        # Once an order is shipped its rolls are already cut and its
        # reservations consumed. Editing items then would desync catalog
        # vs warehouse stock (the item signals adjust catalog while the
        # consumed reservations are never revisited), so block it.
        if self.object.order_status in {"shipped", "in_transit", "out_for_delivery", "delivered"}:
            messages.error(request, "Gönderilen siparişler düzenlenemez.")
            return redirect("operating:order_detail", pk=self.object.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = self.object.items.all()
        context["order_items"] = items
        # JSON-safe payload for the JS to hydrate order_data with.
        # Using a dedicated list keeps it parser-tolerant of any title
        # containing apostrophes, quotes, or newlines.
        context["order_items_payload"] = [
            {
                "item_id": it.pk,
                "sku": (
                    it.product_variant.variant_sku
                    if it.product_variant
                    else (it.product.sku if it.product else "")
                ),
                "variant": bool(it.product_variant),
                "description": it.description or "",
                "quantity": float(it.quantity) if it.quantity is not None else 0,
                "price": float(it.price) if it.price is not None else 0,
            }
            for it in items
        ]
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        # Stock validation removed — back-order is allowed; deduction
        # happens only when the order ships (see operating.signals).
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def form_valid(self, form):
        try:
            with transaction.atomic():
                # Save the order
                self.object = form.save()
                # order = self.object

                # Notification preference — read from POST since OrderForm
                # doesn't include the field. Lets staff toggle it from
                # the same Edit sidebar.
                new_notify = bool(self.request.POST.get("notify_customer"))
                if self.object.notify_customer != new_notify:
                    self.object.notify_customer = new_notify
                    self.object.save(update_fields=["notify_customer", "updated_at"])

                # Handle customer update
                customer = self.request.POST.get("customer")
                customer_pk = self.request.POST.get("customer_pk")
                customer_type = self.request.POST.get("customer_type")

                if customer_pk and customer_type:
                    if customer_type == "contact":
                        self.object.contact = get_object_or_404(Contact, pk=customer_pk)
                        self.object.company = None
                        # change later: book, currency
                        # accounts_receivable = AssetAccountsReceivable(
                        #     book=Book.objects.get(pk=1),
                        #     currency=CurrencyCategory.objects.get(pk=1),
                        #     amount=self.object.total_value(),
                        #     contact=self.object.contact,
                        # )
                        # invoice = Invoice(
                        #     book=Book.objects.get(pk=1),
                        #     currency=CurrencyCategory.objects.get(code="usd"),
                        #     order=order,
                        #     contact=order.contact,
                        # )
                        # invoice = Invoice.objects.get_or_create(order=order, book=Book.objects.get(pk=1),currency=CurrencyCategory.objects.get(code="usd"), contact=order.contact,)
                        # asset_accounts_receivable = AssetAccountsReceivable(
                        #     book=Book.objects.get(pk=1),
                        #     currency=CurrencyCategory.objects.get(code="usd"),
                        #     contact=order.contact,
                        # )
                    elif customer_type == "company":
                        self.object.company = get_object_or_404(Company, pk=customer_pk)
                        self.object.contact = None
                        # invoice = Invoice.objects.get_or_create(order=order, book=Book.objects.get(pk=1),currency=CurrencyCategory.objects.get(code="usd"), company=order.company,)

                        # accounts_receivable = AssetAccountsReceivable(
                        #     book=1,
                        #     currency=1,
                        #     amount=self.object.total_value(),
                        #     company=self.object.company,
                        # )
                        # # AssetAccountsReceivable.objects.update_or_create(
                        # #     book=1,
                        # #     currency=1,
                        # #     amount=self.object.total_value(),
                        # #     company=self.object.company,
                        # # )
                        # accounts_receivable.full_clean()  # Will raise ValidationError if something is wrong
                        # accounts_receivable.save()

                # Handle order items
                product_json_input = self.request.POST.get("product_json_input")
                deleted_items_json = self.request.POST.get("deleted_items")

                # Delete removed items
                if deleted_items_json:
                    try:
                        deleted_items = json.loads(deleted_items_json)
                        OrderItem.objects.filter(
                            pk__in=deleted_items, order=self.object
                        ).delete()
                    except json.JSONDecodeError:
                        messages.error(self.request, "Invalid deleted items format.")

                # Update existing items and add new ones
                if product_json_input:
                    try:
                        product_json_input = json.loads(product_json_input)
                        # product_json_input = [
                        #   {
                        #     "item_no": 1,
                        #     "product": {
                        #       "sku": "RK12471GW8",
                        #       "variant": true
                        #     },
                        #     "description": "firat",
                        #     "quantity": 1,
                        #     "price": 2,
                        #     "item_id": 7
                        #   },
                        #   {
                        #     "item_no": 2,
                        #     "product": {
                        #       "sku": "yoursku",
                        #       "variant": false
                        #     },
                        #     "description": "firat2",
                        #     "quantity": 3,
                        #     "price": 3,
                        #     "item_id": 8
                        #   }]

                        for item_data in product_json_input:
                            if "item_id" in item_data:
                                # Update existing item
                                try:
                                    order_item = OrderItem.objects.get(
                                        pk=item_data["item_id"], order=self.object
                                    )
                                    order_item.description = item_data.get(
                                        "description", ""
                                    )
                                    order_item.quantity = item_data.get("quantity", 1)
                                    order_item.price = item_data.get("price", 0)
                                    order_item.save()
                                except OrderItem.DoesNotExist:
                                    continue
                            else:
                                # Add new item
                                if item_data["product"]["variant"] == False:
                                    product = get_object_or_404(
                                        Product, sku=item_data["product"]["sku"]
                                    )
                                    order_item = OrderItem(
                                        order=self.object,
                                        product=product,
                                        product_variant=None,
                                        description=item_data.get("description", ""),
                                        quantity=item_data.get("quantity", 1),
                                        price=item_data.get("price", 0),
                                    )
                                elif item_data["product"]["variant"] == True:
                                    variant = get_object_or_404(
                                        ProductVariant,
                                        variant_sku=item_data["product"]["sku"],
                                    )
                                    order_item = OrderItem(
                                        order=self.object,
                                        product=variant.product,
                                        product_variant=variant,
                                        description=item_data.get("description", ""),
                                        quantity=item_data.get("quantity", 1),
                                        price=item_data.get("price", 0),
                                    )
                                # add the qr code to the order item
                                try:
                                    order_item.save()
                                except Exception as e:
                                    print("youre dead")
                                    messages.error(
                                        self.request, f"Failed to save order item: {e}"
                                    )
                                    return self.form_invalid(form)
                                # qr_code_url = generate_qr_for_order_item(
                                #     order_item, "pending"
                                # )
                                # order_item.qr_code_url = qr_code_url
                                # order_item.save()
                    except json.JSONDecodeError:
                        messages.error(self.request, "Invalid product data format.")
                        return self.form_invalid(form)
                self.object.save()

                # ── NO warehouse stock movement on edit ───────────────
                # Editing an Açık/Paketleniyor order never touches
                # warehouse rolls. Physical stock is only cut at ship
                # time from the reservations scanned during packing, so
                # there is nothing to reverse/reapply here.

                # ── Sync cari + movement after edit. If the customer
                # was swapped, the order moves to the new cari and
                # post_order_movement() updates the amount in place
                # for the existing source-linked movement.
                try:
                    from current_account.services import (
                        get_or_create_cari_for_order, post_order_movement,
                    )
                    member = getattr(self.request.user, "member", None)
                    cari = get_or_create_cari_for_order(self.object, member=member)
                    if cari and self.object.cari_id != cari.pk:
                        self.object.cari = cari
                        self.object.save(update_fields=["cari"])
                    post_order_movement(self.object, member=member)
                except Exception as _e:
                    messages.warning(self.request, f"Order updated but cari sync had an issue: {_e}")

                messages.success(self.request, "Order updated successfully.")
                if self.request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse({
                        "success": True,
                        "redirect_url": reverse("operating:order_detail", args=[self.object.pk]),
                    })
                return redirect("operating:order_detail", pk=self.object.pk)

        except Exception as e:
            messages.error(self.request, f"Order update failed: {e}")
            print("form is invalid")
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse_lazy("operating:order_detail", kwargs={"pk": self.object.pk})


class OrderList(ListView):
    model = Order
    template_name = "operating/order_list.html"
    context_object_name = "orders"
    ordering = ["-created_at"]
    
    def get_queryset(self):
        # Optimize with select_related and prefetch_related to avoid N+1 queries.
        # `items__product` and `items__product_variant` are needed because
        # the gross_profit() helper reads cost from those — without
        # prefetching, each order row would fire two extra queries per
        # line item.
        return (
            Order.objects
            .select_related('contact', 'company', 'web_client')
            .prefetch_related('items__product', 'items__product_variant')
            .order_by("-created_at")
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Use the already-evaluated list to share prefetch cache across tabs
        all_orders = context.get('orders', [])
        
        # B2B orders: Has contact OR company, but NO web_client
        b2b_orders = [
            order for order in all_orders 
            if (order.contact or order.company) and not order.web_client
        ]
        
        # B2C orders: Has web_client (regardless of contact/company)
        b2c_orders = [
            order for order in all_orders 
            if order.web_client
        ]
        
        # Add to context
        context['b2b_orders'] = b2b_orders
        context['b2c_orders'] = b2c_orders
        context['total_count'] = len(all_orders)
        context['b2b_count'] = len(b2b_orders)
        context['b2c_count'] = len(b2c_orders)
        
        return context


class OrderProduction(View):
    template_name = "operating/order_production.html"

    def get(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        order_items = OrderItem.objects.filter(order=order)

        return render(
            request, self.template_name, {"order": order, "order_items": order_items}
        )


# This is how machines read and update the order status.
@csrf_exempt
def machine_update_status(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    try:
        data = json.loads(request.body)
        order_id = data.get("order_id")
        item_ids = data.get("item_ids", [])
        new_status = data.get("new_status")

        order = Order.objects.get(pk=order_id)
        updated_count = 0

        for item in order.items.filter(pk__in=item_ids):
            item.status = new_status
            item.save()
            updated_count += 1

        order.update_status_from_items()
        return JsonResponse({"success": True, "updated_items": updated_count})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# machines should not need csrf
@method_decorator(csrf_exempt, name="dispatch")
class MachineStatusUpdate(View):
    def post(self, request, item_id):
        api_key = request.headers.get("X-API-KEY")
        if not api_key:
            return JsonResponse({"error": "API key missing"}, status=400)

        machine = get_machine_from_api_key(api_key)
        if not machine:
            return JsonResponse({"error": "Unauthorized"}, status=401)

        try:
            data = json.loads(request.body)
            new_status = data.get("new_status")
            if not new_status:
                return JsonResponse({"error": "Missing 'new_status'"}, status=400)

            item = OrderItem.objects.get(pk=item_id)
            item.status = new_status
            item.save()

            item.order.update_status_from_items()

            return JsonResponse(
                {
                    "message": f"Status updated to '{new_status}' by machine '{machine.name}'"
                }
            )

        except OrderItem.DoesNotExist:
            return JsonResponse({"error": "OrderItem not found"}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)


class OrderItemUnitScan(View):
    choices = STATUS_CHOICES
    template_name = "operating/scan_order_item_unit.html"

    def get(self, request):
        # return render(request,"")
        return render(request, self.template_name, {"choices": STATUS_CHOICES})


class OrderItemUnitScanPack(View):

    template_name = "operating/scan_order_item_unit_pack.html"

    def get(self, request):
        context = {"max_pack_range": range(1, 51)}  # loops from 1 to 50
        return render(request, self.template_name, context)


class OrderPackingList(View):
    """Modern packing UI — pack at the OrderItem level.

    GET  → render order items + their pack assignments. Auto-creates
           Pack #1 the first time the page is opened.
    POST → JSON API for the inline UI:
           - action=add_pack            → create the next-numbered Pack
           - action=assign  + item_id + pack_id → move item to that pack
           - action=remove  + item_id   → unassign from any pack
           - action=delete_pack + pack_id → remove pack (items go back to unassigned)
    """
    template_name = "operating/order_packing_list.html"

    def get(self, request, pk):
        from .models import PackedOrderItem
        order = get_object_or_404(
            Order.objects.prefetch_related(
                "items__product", "items__product_variant",
            ),
            pk=pk,
        )
        # Auto-create Pack #1 if none exist — the UI assumes at least one.
        if not order.packs.exists():
            Pack.objects.create(order=order, pack_number=1)
        packs = list(order.packs.order_by("pack_number")
                                .prefetch_related("packed_order_items__order_item__product"))

        items = list(order.items.select_related("product", "product_variant"))
        assigned = {
            poi.order_item_id: poi.pack_id
            for poi in PackedOrderItem.objects.filter(pack__order=order)
        }
        for it in items:
            it.assigned_pack_id = assigned.get(it.pk)

        return render(request, self.template_name, {
            "order": order,
            "packs": packs,
            "items": items,
        })

    def post(self, request, pk):
        from .models import PackedOrderItem
        order = get_object_or_404(Order, pk=pk)
        action = (request.POST.get("action") or "").strip()

        if action == "add_pack":
            next_n = (order.packs.aggregate(m=models.Max("pack_number"))["m"] or 0) + 1
            pack = Pack.objects.create(order=order, pack_number=next_n)
            return JsonResponse({
                "ok": True,
                "pack": {
                    "id": pack.pk,
                    "number": pack.pack_number,
                    "code": pack.code,
                    "qr_code_url": pack.qr_code_url,
                }
            })

        if action == "delete_pack":
            try:
                pack_id = int(request.POST.get("pack_id") or 0)
            except (TypeError, ValueError):
                return JsonResponse({"ok": False, "error": "Bad pack_id"}, status=400)
            pack = get_object_or_404(Pack, pk=pack_id, order=order)
            # Refuse to delete the last remaining pack.
            if order.packs.count() <= 1:
                return JsonResponse({"ok": False, "error": "Cannot delete the only pack"}, status=400)
            pack.delete()  # CASCADE deletes the packed_order_items too
            return JsonResponse({"ok": True})

        if action == "assign":
            try:
                item_id = int(request.POST.get("item_id") or 0)
                pack_id = int(request.POST.get("pack_id") or 0)
            except (TypeError, ValueError):
                return JsonResponse({"ok": False, "error": "Bad params"}, status=400)
            item = get_object_or_404(OrderItem, pk=item_id, order=order)
            pack = get_object_or_404(Pack, pk=pack_id, order=order)
            # OneToOneField guarantees one-pack-per-item: delete any
            # existing assignment, then create the new one. This is
            # the "auto-move" the user asked for.
            PackedOrderItem.objects.filter(order_item=item).delete()
            PackedOrderItem.objects.create(pack=pack, order_item=item)
            return JsonResponse({
                "ok": True,
                "item_id": item.pk,
                "pack_id": pack.pk,
                "pack_number": pack.pack_number,
            })

        if action == "remove":
            try:
                item_id = int(request.POST.get("item_id") or 0)
            except (TypeError, ValueError):
                return JsonResponse({"ok": False, "error": "Bad item_id"}, status=400)
            PackedOrderItem.objects.filter(order_item_id=item_id,
                                           pack__order=order).delete()
            return JsonResponse({"ok": True, "item_id": item_id})

        return JsonResponse({"ok": False, "error": "Unknown action"}, status=400)


# below is for receiving goods
# In future you need to combine them
class RawMaterialGoodCreate(CreateView):
    model = RawMaterialGood
    fields = "__all__"
    # form_class = RawMaterialGoodReceiptForm
    template_name = "operating/create_raw_material_good.html"
    success_url = reverse_lazy("operating:index")


class RawMaterialGoodReceiptCreate(CreateView):
    model = RawMaterialGoodReceipt
    form_class = RawMaterialGoodReceiptForm
    template_name = "operating/create_raw_material_good_receipt.html"
    success_url = reverse_lazy("operating:create_raw_material_good_receipt")


class RawMaterialGoodItemCreate(CreateView):
    model = RawMaterialGoodItem
    form_class = RawMaterialGoodItemForm
    template_name = "operating/create_raw_material_good_item.html"
    success_url = reverse_lazy("operating:create_raw_material_good_item")


# -------------------------------- function based views  -------------------------------- #

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def export_packing_list_excel(request, pk):
    print("export packing list excel for order with pk: ", pk)
    order = Order.objects.get(pk=pk)
    packs = order.packs.prefetch_related("items__order_item_unit__order_item")

    # create excel workbook and worksheet
    wb = openpyxl.Workbook()
    # select the active worksheet
    ws = wb.active
    ws.title = f"Packing List - Order {pk}"

    # write the title row
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Packing List for Order #{pk} — {order.get_client()} "
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Pack Number",
        "SKU",
        "Description",
        "Quantity",
        "Unit of Measure",
        "Unit ID",
    ]
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

    # End of first rows.
    row = 3
    for pack in packs:
        pack_start_row = row

        for packed_item in pack.items.all():
            unit = packed_item.order_item_unit
            order_item = unit.order_item

            values = [
                pack.pack_number,
                order_item.display_sku(),
                order_item.description or "N/A",
                unit.quantity,
                order_item.product.unit_of_measurement or "units",
                unit.pk,
            ]

            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_index, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1

        # Add thick border around entire pack row range

        # Add thick border around entire pack row range
        for r in range(pack_start_row, row):
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = thick_border

        # Merge cells in "Pack Number" column (A) for the same pack
        if row - pack_start_row > 1:
            ws.merge_cells(
                start_row=pack_start_row, start_column=1, end_row=row - 1, end_column=1
            )
            merged_cell = ws.cell(row=pack_start_row, column=1)
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            # merged_cell.alignment = Alignment(vertical="center", horizontal="center")

    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Export
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"packing_list_order_{pk}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    # ws.cell().alignment = Alignment(horizontal="center", vertical="center")
    wb.save(response)
    return response

    # return HttpResponse("Packing list export is not implemented yet.")


def delete_order(request, pk):
    try:
        order = Order.objects.get(pk=pk)
        order.delete()
        messages.success(request, "Order deleted successfully.")
    except Order.DoesNotExist:
        messages.error(request, "Order not found.")
    return redirect("operating:order_list")


@require_POST
def bulk_delete_orders(request):
    """Delete a set of orders by id. Used by the order_list checkbox
    selection + floating "Delete selected" bar. Returns JSON with the
    count actually deleted so the JS can flash a confirmation.

    OrderItem.post_delete signals (stock restore + cari reversal) fire
    naturally as the cascade unfolds — no extra work needed here."""
    ids_raw = request.POST.getlist("order_ids[]") or request.POST.getlist("order_ids")
    if not ids_raw and request.content_type == "application/json":
        try:
            payload = json.loads(request.body.decode() or "{}")
            ids_raw = payload.get("order_ids") or []
        except (json.JSONDecodeError, UnicodeDecodeError):
            ids_raw = []
    ids = []
    for v in ids_raw:
        try:
            ids.append(int(v))
        except (TypeError, ValueError):
            pass
    if not ids:
        return JsonResponse({"ok": False, "error": "No order_ids supplied"}, status=400)
    qs = Order.objects.filter(pk__in=ids)
    count = qs.count()
    # Iterate so each Order.delete() triggers its cascade signals
    # (OrderItem post_delete restores catalog stock; cari movement is
    # reversed). bulk delete via qs.delete() would still cascade but
    # we keep it per-row for predictable signal ordering on Postgres.
    for o in qs:
        o.delete()
    return JsonResponse({"ok": True, "deleted": count})



def webclient_autocomplete(request):
    from authentication.models import WebClient
    query = request.GET.get("webclient_q", "").strip().lower()
    if not query:
        return HttpResponse("")
    clients = WebClient.objects.filter(
        Q(email__icontains=query) | Q(username__icontains=query) | Q(name__icontains=query)
    )[:10]
    if not clients:
        return HttpResponse("<div class='product_autocomplete_list'><ul><li style='color:#9ca3af;padding:10px;'>No web clients found</li></ul></div>")
    items = []
    for c in clients:
        name = escape(c.name or c.username)
        email = escape(c.email)
        items.append(
            f"<li onclick=\"selectWebClient({c.pk}, '{name.replace(chr(39), chr(92)+chr(39))}')\">"
            f"<strong>{name}</strong> <span style='color:#6b7280;'>— {email}</span></li>"
        )
    return HttpResponse(f"<div class='product_autocomplete_list'><ul>{''.join(items)}</ul></div>")


# escape(product.title) will turn < into &lt;, > into &gt;, & into &amp;, etc.
# This ensures that if a product title or variant SKU contains special characters (like <, >, &, or quotes), they won't break your HTML or allow malicious code to run.
def product_autocomplete(request):
    """Order-create/edit product search.

    Sources, in priority order:
      1. WAREHOUSE products — anything held in any warehouse (searched by
         warehouse name/SKU and by the linked catalog title/variant-sku,
         Turkish İ/ı-folded). Stock comes from the WAREHOUSE, shown per
         warehouse. Products existing in both places surface HERE.
      2. Catalog-only products — products/variants with no warehouse
         link, so the whole assortment stays searchable.
    Every row shows name + SKU."""
    from functools import reduce
    import operator as _op
    from .views_warehouse import _tr_ci_variants
    from .models import WarehouseProduct

    query = request.GET.get("product", "").strip().lower()
    if not query:
        return HttpResponse("")

    q_variants = _tr_ci_variants(query)

    def _fq(field):
        return reduce(_op.or_, (Q(**{f"{field}__icontains": v}) for v in q_variants))

    # ── 1) Warehouse hits, grouped by their catalog variant ──────────
    wh_rows = (
        WarehouseProduct.objects
        .filter(catalog_variant__isnull=False)
        .filter(_fq("name") | _fq("sku")
                | _fq("catalog_variant__variant_sku")
                | _fq("catalog_variant__product__title"))
        .select_related("warehouse", "catalog_variant__product__category")
        .order_by("name")[:60]
    )
    wh_groups = {}          # variant_id -> {"wp": first wp, "stocks": [(wh, qty)]}
    for wp in wh_rows:
        g = wh_groups.setdefault(wp.catalog_variant_id, {"wp": wp, "stocks": []})
        g["stocks"].append((wp.warehouse.name, wp.quantity or 0))
    wh_variant_ids = set(wh_groups.keys())

    # ── 2) Catalog-only fallback (nothing already shown as warehouse) ─
    products = (
        Product.objects.filter(
            _fq("title") | _fq("sku") | _fq("variants__variant_sku")
        )
        .select_related("category")
        .distinct()
        .prefetch_related("variants")[:12]
    )

    def js_str(s):
        """Escape string for safe use inside JS onclick single quotes"""
        return (s or "").replace("\\", "\\\\").replace("'", "\\'").replace('"', '&quot;').replace("\n", " ")

    def stock_badge(qty, allow_oversell):
        """Render a small "Stok: N" pill so the user sees availability
        right in the autocomplete row. None → unlimited (no badge).
        Zero → red badge. Positive → grey/green badge. allow_oversell
        flips the negative-stock case to neutral instead of red."""
        if qty is None:
            return ""
        try:
            q = float(qty)
        except (TypeError, ValueError):
            return ""
        if q <= 0 and not allow_oversell:
            color, bg = "#B91C1C", "#FEE2E2"
        elif q < 10:
            color, bg = "#92400E", "#FEF3C7"
        else:
            color, bg = "#065F46", "#D1FAE5"
        return (
            f"<span style='font-size:10.5px;font-weight:600;padding:2px 6px;"
            f"border-radius:6px;color:{color};background:{bg};margin-right:8px;white-space:nowrap;'>"
            f"Stok: {q:g}</span>"
        )

    def render_item(product, variant=None):
        title = escape(product.title or "")
        cat_name = escape(product.category.name if product.category else "")
        title_js = js_str(product.title or "")
        cat_js = js_str(product.category.name if product.category else "")
        allow_oversell = bool(getattr(product, "selling_while_out_of_stock", False))

        if variant:
            sku = escape(variant.variant_sku or "")
            price = variant.variant_price or product.price or 0
            stock = variant.variant_quantity
            stock_arg = "null" if stock is None else f"{float(stock):g}"
            oversell_arg = "true" if allow_oversell else "false"
            attr_info = variant.attribute_summary() if hasattr(variant, 'attribute_summary') else ''
            attr_display = f' <span style="color:#6b7280;font-size:11px;">({escape(attr_info)})</span>' if attr_info else ''
            return (
                f"<li onclick=\"selectProduct('{sku}',true,'{title_js}',{price},'{cat_js}',{stock_arg},{oversell_arg})\" style='display:flex;align-items:center;justify-content:space-between;gap:12px;'>"
                f"<span style='min-width:0;flex-grow:1;word-break:break-word;'>"
                f"<strong>{title}</strong> - <code>{sku}</code>{attr_display}"
                f"</span>"
                f"<span style='display:inline-flex;align-items:center;white-space:nowrap;flex-shrink:0;'>"
                f"{stock_badge(stock, allow_oversell)}"
                f"<span style='color:#059669;font-weight:600;'>${price}</span>"
                f"</span></li>"
            )
        else:
            sku = escape(product.sku or "")
            price = product.price or 0
            stock = product.quantity
            stock_arg = "null" if stock is None else f"{float(stock):g}"
            oversell_arg = "true" if allow_oversell else "false"
            return (
                f"<li onclick=\"selectProduct('{sku}',false,'{title_js}',{price},'{cat_js}',{stock_arg},{oversell_arg})\" style='display:flex;align-items:center;justify-content:space-between;gap:12px;'>"
                f"<span style='min-width:0;flex-grow:1;word-break:break-word;'>"
                f"<strong>{title}</strong> - <code>{sku}</code>"
                f"</span>"
                f"<span style='display:inline-flex;align-items:center;white-space:nowrap;flex-shrink:0;'>"
                f"{stock_badge(stock, allow_oversell)}"
                f"<span style='color:#059669;font-weight:600;'>${price}</span>"
                f"</span></li>"
            )

    def render_wh_item(group):
        """One row per catalog variant held in warehouses. Stock badges
        are PER WAREHOUSE (name + metres); the JS stock arg is the total."""
        wp = group["wp"]
        variant = wp.catalog_variant
        parent = variant.product
        display_name = wp.name or parent.title or ""
        title = escape(display_name)
        sku = escape(variant.variant_sku or "")
        title_js = js_str(parent.title or display_name)
        cat_js = js_str(parent.category.name if parent.category else "")
        allow_oversell = bool(getattr(parent, "selling_while_out_of_stock", False))
        price = variant.variant_price or parent.price or 0
        total = sum((q or 0) for _, q in group["stocks"])
        stock_arg = f"{float(total):g}"
        oversell_arg = "true" if allow_oversell else "false"
        wh_badges = "".join(
            f"<span style='font-size:10px;font-weight:700;padding:2px 7px;border-radius:6px;"
            f"color:#005354;background:#E5F1F0;margin-right:5px;white-space:nowrap;'>"
            f"{escape((wh or '')[:14])}: {float(q or 0):g} m</span>"
            for wh, q in group["stocks"][:3]
        )
        hint = ""
        if parent.title and (parent.title or "").strip().lower() != display_name.strip().lower():
            hint = f" <span style='color:#6b7280;font-size:11px;'>→ {escape(parent.title)}</span>"
        return (
            f"<li onclick=\"selectProduct('{sku}',true,'{title_js}',{price},'{cat_js}',{stock_arg},{oversell_arg})\" style='display:flex;align-items:center;justify-content:space-between;gap:12px;'>"
            f"<span style='min-width:0;flex-grow:1;word-break:break-word;'>"
            f"<span style='font-size:9.5px;font-weight:800;letter-spacing:.05em;color:#0F766E;background:#CCFBF1;border-radius:5px;padding:1px 6px;margin-right:6px;vertical-align:middle;'>DEPO</span>"
            f"<strong>{title}</strong> - <code>{sku}</code>{hint}"
            f"</span>"
            f"<span style='display:inline-flex;align-items:center;white-space:nowrap;flex-shrink:0;'>"
            f"{wh_badges}"
            f"<span style='color:#059669;font-weight:600;'>${price}</span>"
            f"</span></li>"
        )

    items = []

    # Warehouse rows first — shared products get picked FROM the warehouse.
    for group in list(wh_groups.values())[:8]:
        items.append(render_wh_item(group))

    # Catalog-only: skip every variant that lives in a warehouse (those
    # are the rows above); keep variant-less products and unlinked variants.
    cand_variant_ids = [v.id for p in products for v in p.variants.all()]
    linked_ids = set(
        WarehouseProduct.objects
        .filter(catalog_variant_id__in=cand_variant_ids)
        .values_list("catalog_variant_id", flat=True)
    ) | wh_variant_ids

    shown_catalog = 0
    for product in products:
        if shown_catalog >= 6:
            break
        variants = [v for v in product.variants.all() if v.id not in linked_ids]

        matches_parent = (
            query in (product.title or "").lower()
            or query in (product.sku or "").lower()
        )

        if not product.variants.all():
            if matches_parent:
                items.append(render_item(product))
                shown_catalog += 1
        else:
            for variant in variants:
                if matches_parent or query in (variant.variant_sku or "").lower():
                    items.append(render_item(product, variant))
                    shown_catalog += 1

    if not items:
        items.append("<li style='color:#9ca3af;padding:12px;'>Sonuç bulunamadı</li>")

    html = f"<div class='product_autocomplete_list'><ul>{''.join(items)}</ul></div>"
    return HttpResponse(html)


@require_POST
def start_production(request):
    try:
        with transaction.atomic():
            print("your row index")
            print("got your post request bro")
            # for key, value in request.POST.items():
            #     print(f"{key}: {value}")
            print("pack count is:", request.POST.get("pack_count"))
            pack_count = request.POST.get("pack_count")
            order_item_id = request.POST.get("order_item_id")
            target_quantity_per_pack = request.POST.get("target_quantity_per_pack")
            print("target_quantity_per_pack:", target_quantity_per_pack)

            if not pack_count or not target_quantity_per_pack or not order_item_id:
                return HttpResponseBadRequest("Missing required fields")

            try:
                pack_count = int(pack_count)
                target_quantity_per_pack = int(target_quantity_per_pack)
            except (ValueError, TypeError):
                return HttpResponseBadRequest("Invalid number format")

            order_item = OrderItem.objects.get(pk=order_item_id)

            order_item.target_quantity_per_pack = target_quantity_per_pack
            order_item.save(update_fields=["target_quantity_per_pack"])

            print("your order item id is:", order_item_id)
            for _ in range(pack_count):
                order_item_unit = OrderItemUnit(
                    order_item=order_item,
                    quantity=target_quantity_per_pack,
                    status="scheduled",
                )
                # order_item_unit.save()
                # order_item_unit.qr_code_url = generate_qr_for_order_item_unit(
                #     order_item_unit, status=order_item_unit.status
                # )
                # order_item_unit.save(update_fields="qr_code_url")
                try:
                    order_item_unit.full_clean()  # Trigger model field validation
                    order_item_unit.save()
                    order_item_unit.qr_code_url = generate_qr_for_order_item_unit(
                        order_item_unit, status=order_item_unit.status
                    )
                    order_item_unit.save(update_fields=["qr_code_url"])
                except ValidationError as ve:
                    print("💥 Validation Error on OrderItemUnit:")
                    print(ve.message_dict)
                    traceback.print_exc()
                    return HttpResponse(
                        f"<span style='color:red;'>❌ Validation Error: {ve.message_dict}</span>",
                        status=400,
                    )
            qr_url = reverse(
                "operating:generate_pdf_qr_for_order_item_units", args=[order_item.pk]
            )
            return HttpResponse(
                # '<a href="{% url \'operating:generate_pdf_qr_for_order_item_units\' item.pk %}" class="print_barcode_button" target="_blank">Print QR Labels</a>'
                f'<a href="{qr_url}" class="print_barcode_button" target="_blank">Print QR Labels</a>'
                # "<span>Hello</span>"
            )
    except Exception as e:
        return HttpResponse(
            f"<span style='color:red;'>❌ Error: {str(e)}</span>", status=500
        )


from io import BytesIO
from PIL import Image
import requests
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
import os


def generate_pdf_qr_for_order_item_units(request, pk):
    order_item = OrderItem.objects.get(pk=pk)
    order_item_units = order_item.units.all()

    # This object is an in-memory buffer that will hold the file content. It acts as a stream, allowing us to read from or write to it as if it were a file on disk.
    buffer = BytesIO()
    # page_width, page_height = letter
    # page_width = 2 * inch
    # page_height = 3 * inch
    # 2 by 3 inches
    width, height = 2 * inch, 3 * inch
    pdf = canvas.Canvas(buffer, pagesize=(width, height))

    for unit in order_item_units:
        # load qr image from url
        qr_url = unit.qr_code_url
        try:
            qr_response = requests.get(qr_url)
            qr_response.raise_for_status()
            qr_image_pil = Image.open(BytesIO(qr_response.content)).convert("RGB")
        except Exception as e:
            # continue  # Skip this unit if QR can't be loaded
            return HttpResponse('<p class="error">something went wrong</p>')

        # # save temporarily to save into PDF
        # temp_path = f"/tmp/temp_qr_pdf_{unit.pk}.png"
        # qr_img.save(temp_path)
        qr_width, qr_height = qr_image_pil.size
        scale = (1.6 * inch) / qr_width  # scale to fit
        qr_width_scaled = qr_width * scale
        qr_height_scaled = qr_height * scale

        # Draw QR image
        pdf.drawInlineImage(
            qr_image_pil,
            x=(width - qr_width_scaled) / 2,
            y=height - qr_height_scaled - 0.7 * inch,
            width=qr_width_scaled,
            height=qr_height_scaled,
        )

        # Add OrderItem display name
        pdf.setFont("Helvetica", 9)
        pdf.drawCentredString(width / 2, 0.7 * inch, order_item.display_name())

        # Add Unit ID
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawCentredString(width / 2, 0.5 * inch, f"Unit #{unit.pk}")

        pdf.showPage()

    pdf.save()
    buffer.seek(0)

    return HttpResponse(
        buffer,
        content_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="order_item_{pk}_qr.pdf"'
        },
    )


def get_pack_qr_image(pack):
    import requests
    from PIL import Image
    from io import BytesIO
    import json
    import segno
    from marketing.utils.bunny_storage import upload_to_bunny

    if pack.qr_code_url:
        try:
            response = requests.get(pack.qr_code_url, timeout=5)
            response.raise_for_status()
            return Image.open(BytesIO(response.content)).convert("RGB")
        except Exception:
            pass
    # Generate on the fly if fetch fails or URL is empty
    payload = {
        "pack_id": pack.pk,
        "code": pack.code,
        "order_id": pack.order.pk,
        "pack_number": pack.pack_number
    }
    qr = segno.make(json.dumps(payload))
    buf = BytesIO()
    qr.save(buf, kind="png", scale=5)
    buf.seek(0)
    # Also upload it to update/fix the url
    try:
        path = f"media/orders/{pack.order.pk}/packs/{pack.pk}/qr_{pack.pk}.png"
        buf.seek(0)
        url = upload_to_bunny(buf, path)
        pack.qr_code_url = url
        pack.save(update_fields=["qr_code_url"])
    except Exception:
        pass
    buf.seek(0)
    return Image.open(buf).convert("RGB")


def order_packing_list_pdf(request, pk):
    from django.utils.translation import gettext as _
    from django.utils.translation import get_language
    from .models import Order, Pack
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
    from reportlab.platypus import Image as RLImage
    from .order_notifications import _ensure_pdf_fonts
    
    order = get_object_or_404(Order, pk=pk)
    packs = order.packs.all().order_by("pack_number")
    
    # Determine the active language
    lang = (get_language() or 'tr').lower()
    is_tr = lang.startswith('tr')
    
    # Text labels mapping dynamically based on active language
    labels = {
        'title': "Sipariş Paket Listesi" if is_tr else "Order Packing List",
        'order_no': "Sipariş No" if is_tr else "Order No",
        'date': "Tarih" if is_tr else "Date",
        'customer': "Müşteri" if is_tr else "Customer",
        'package': "Paket" if is_tr else "Package",
        'product': "Ürün" if is_tr else "Product",
        'sku': "SKU",
        'qty': "Miktar" if is_tr else "Qty",
    }
    
    font = _ensure_pdf_fonts() or "Helvetica"
    font_bold = f"{font}-Bold" if font != "Helvetica" else "Helvetica-Bold"

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"{labels['title']} #{order.pk}"
    )
    
    story = []
    
    title_style = ParagraphStyle(
        name="TitleStyle",
        fontName=font_bold,
        fontSize=18,
        textColor=colors.HexColor("#111111"),
        spaceAfter=15
    )
    
    normal_style = ParagraphStyle(
        name="NormalStyle",
        fontName=font,
        fontSize=10,
        textColor=colors.HexColor("#333333"),
        leading=14
    )
    
    heading_style = ParagraphStyle(
        name="HeadingStyle",
        fontName=font_bold,
        fontSize=12,
        textColor=colors.HexColor("#00696A"),
        spaceBefore=15,
        spaceAfter=5
    )

    story.append(Paragraph(labels['title'], title_style))
    story.append(Paragraph(f"<b>{labels['order_no']}:</b> {order.order_number or order.pk}", normal_style))
    story.append(Paragraph(f"<b>{labels['date']}:</b> {order.created_at.strftime('%d.%m.%Y %H:%M')}", normal_style))
    
    cust_name = ""
    if order.contact:
        cust_name = order.contact.name
    elif order.company:
        cust_name = order.company.name
    elif order.web_client:
        cust_name = order.web_client.name or order.web_client.username
    story.append(Paragraph(f"<b>{labels['customer']}:</b> {cust_name or '-'}", normal_style))
    
    story.append(Spacer(1, 10))
    
    for pack in packs:
        pack_story = []
        pack_story.append(Paragraph(f"{labels['package']} #{pack.pack_number} ({pack.code})", heading_style))
        
        qr_pil = get_pack_qr_image(pack)
        
        items_data = [[
            Paragraph(f"<b>{labels['product']}</b>", normal_style),
            Paragraph(f"<b>{labels['sku']}</b>", normal_style),
            Paragraph(f"<b>{labels['qty']}</b>", normal_style)
        ]]
        
        pois = pack.packed_order_items.all()
        for poi in pois:
            it = poi.order_item
            title = getattr(it.product, "title", None) or it.description or "Product"
            sku = (it.product_variant.variant_sku if (it.product_variant_id and it.product_variant) 
                   else (it.product.sku if it.product else "-"))
            qty = str(it.quantity)
            items_data.append([
                Paragraph(title, normal_style),
                Paragraph(sku, normal_style),
                Paragraph(qty, normal_style)
            ])
            
        col_widths = [85 * mm, 38 * mm, 20 * mm]
        tbl = Table(items_data, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8EE")),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F9FAFB")),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        qr_io = BytesIO()
        qr_pil.save(qr_io, format="PNG")
        qr_io.seek(0)
        rl_qr_img = RLImage(qr_io, width=32 * mm, height=32 * mm)
        rl_qr_img.hAlign = 'CENTER'
        
        layout_table = Table([[tbl, rl_qr_img]], colWidths=[143 * mm, 37 * mm])
        layout_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        pack_story.append(layout_table)
        pack_story.append(Spacer(1, 10))
        
        story.append(KeepTogether(pack_story))
        
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="packing_list_{order.order_number or order.pk}.pdf"'
    return response


def pack_pdf(request, pack_pk):
    from django.utils.translation import gettext as _
    from django.utils.translation import get_language
    from .models import Pack
    from reportlab.lib.pagesizes import A6
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from .order_notifications import _ensure_pdf_fonts
    
    pack = get_object_or_404(Pack, pk=pack_pk)
    order = pack.order
    
    # Determine the active language
    lang = (get_language() or 'tr').lower()
    is_tr = lang.startswith('tr')
    
    # Text labels mapping dynamically based on active language
    labels = {
        'package': "PAKET" if is_tr else "PACKAGE",
        'order': "Sipariş" if is_tr else "Order",
        'contents': "İÇERİK" if is_tr else "CONTENTS",
        'product': "Ürün" if is_tr else "Product",
        'qty': "Miktar" if is_tr else "Qty",
    }
    
    font = _ensure_pdf_fonts() or "Helvetica"
    font_bold = f"{font}-Bold" if font != "Helvetica" else "Helvetica-Bold"

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A6,
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
        title=f"Pack {pack.code}"
    )
    
    story = []
    
    title_style = ParagraphStyle(
        name="LabelTitle",
        fontName=font_bold,
        fontSize=12,
        textColor=colors.HexColor("#111111"),
        alignment=1,
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        name="LabelSub",
        fontName=font,
        fontSize=8,
        textColor=colors.HexColor("#555555"),
        alignment=1,
        spaceAfter=10
    )
    
    normal_style = ParagraphStyle(
        name="LabelNormal",
        fontName=font,
        fontSize=8,
        textColor=colors.HexColor("#111111"),
        leading=11
    )
    
    code_style = ParagraphStyle(
        name="LabelCode",
        fontName=font_bold,
        fontSize=9,
        textColor=colors.HexColor("#00696A"),
        alignment=1,
        spaceBefore=6
    )

    story.append(Paragraph(f"<b>{labels['package']} #{pack.pack_number}</b>", title_style))
    story.append(Paragraph(f"{labels['order']}: {order.order_number or order.pk}", subtitle_style))
    
    qr_pil = get_pack_qr_image(pack)
    qr_io = BytesIO()
    qr_pil.save(qr_io, format="PNG")
    qr_io.seek(0)
    rl_qr_img = RLImage(qr_io, width=40 * mm, height=40 * mm)
    rl_qr_img.hAlign = 'CENTER'
    
    story.append(rl_qr_img)
    story.append(Paragraph(f"{pack.code}", code_style))
    story.append(Spacer(1, 8))
    
    story.append(Paragraph(f"<b>{labels['contents']}:</b>", normal_style))
    
    items_data = [[
        Paragraph(f"<b>{labels['product']}</b>", normal_style),
        Paragraph(f"<b>{labels['qty']}</b>", normal_style)
    ]]
    
    pois = pack.packed_order_items.all()
    for poi in pois:
        it = poi.order_item
        title = getattr(it.product, "title", None) or it.description or "Product"
        qty = str(it.quantity)
        items_data.append([
            Paragraph(title, normal_style),
            Paragraph(qty, normal_style)
        ])
        
    tbl = Table(items_data, colWidths=[65 * mm, 24 * mm])
    tbl.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="pack_{pack.code}.pdf"'
    return response


def get_order_status(request, order_id):

    try:
        order = Order.objects.get(pk=order_id)
        return JsonResponse(
            {
                "id": order.pk,
                "status": order.get_status_display(),
                "created_at": order.created_at.isoformat(),
                "updated_at": order.updated_at.isoformat(),
                "items": [
                    {
                        "id": item.pk,
                        "product_category": (
                            item.product.category.name if item.product else ""
                        ),
                        "product_sku": item.product.sku,
                        "product_title": item.product.title,
                        "description": item.description,
                        "quantity": str(item.quantity),
                        "status": item.get_status_display(),
                        "unit_of_measurement": item.product.unit_of_measurement,
                    }
                    for item in order.items.all()
                ],
            }
        )
    except Order.DoesNotExist:
        return JsonResponse({"error": "Order not found"}, status=404)


@csrf_exempt
def update_order_ettn(request, order_id):
    """Update order with ETTN (e-Arşiv invoice number)"""
    if request.method != 'PATCH':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        order = Order.objects.get(pk=order_id)
        data = json.loads(request.body)
        
        ettn = data.get('ettn')
        invoice_date = data.get('invoice_date')
        
        if ettn:
            order.ettn = ettn
        if invoice_date:
            order.invoice_date = invoice_date
        
        order.save()
        
        return JsonResponse({
            'success': True,
            'message': 'ETTN updated successfully',
            'order_id': order.pk,
            'ettn': order.ettn
        }, status=200)
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Order not found'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def get_order_detail_api(request, user_id, order_id):
    """Get order detail including ETTN for web clients"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        from authentication.models import WebClient
        
        # Get web client
        try:
            web_client = WebClient.objects.get(pk=user_id)
        except WebClient.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)
        
        # Get order
        order = Order.objects.get(pk=order_id, web_client=web_client)
        
        # Build order data
        order_data = {
            'id': order.pk,
            'order_number': f'ORD-{order.pk}',
            'status': order.status,
            'created_at': order.created_at.isoformat(),
            'updated_at': order.updated_at.isoformat(),
            'ettn': order.ettn,
            'invoice_date': order.invoice_date.isoformat() if order.invoice_date else None,
            
            # Payment info
            'payment_id': order.payment_id,
            'payment_status': order.payment_status,
            'payment_method': order.payment_method,
            
            # Pricing
            'original_currency': order.original_currency,
            'original_price': str(order.original_price) if order.original_price else None,
            'paid_currency': order.paid_currency,
            'paid_amount': str(order.paid_amount) if order.paid_amount else None,
            'exchange_rate': str(order.exchange_rate) if order.exchange_rate else None,
            
            # Card info
            'card_type': order.card_type,
            'card_association': order.card_association,
            'card_last_four': order.card_last_four,
            
            # Addresses
            'delivery_address_title': order.delivery_address_title,
            'delivery_address': order.delivery_address,
            'delivery_city': order.delivery_city,
            'delivery_country': order.delivery_country,
            'delivery_phone': order.delivery_phone,
            
            'billing_address_title': order.billing_address_title,
            'billing_address': order.billing_address,
            'billing_city': order.billing_city,
            'billing_country': order.billing_country,
            'billing_phone': order.billing_phone,
            
            # Items
            'items': [
                {
                    'id': item.pk,
                    'product_sku': item.product.sku,
                    'product_title': item.product.title,
                    'product_image': item.product.primary_image.file_url if item.product.primary_image else None,
                    'product_variant_sku': item.product_variant.variant_sku if item.product_variant else None,
                    'quantity': str(item.quantity),
                    'price': str(item.price),
                    'subtotal': str(item.subtotal()) if item.quantity and item.price else None,
                    'description': item.description,
                    'status': item.status,
                    # Custom Curtain Fields
                    'is_custom_curtain': item.is_custom_curtain,
                    'custom_fabric_used_meters': str(item.custom_fabric_used_meters) if item.custom_fabric_used_meters else None,
                    'custom_attributes': {
                        'mounting_type': item.custom_mounting_type,
                        'pleat_type': item.custom_pleat_type,
                        'pleat_density': item.custom_pleat_density,
                        'width': str(item.custom_width) if item.custom_width else None,
                        'height': str(item.custom_height) if item.custom_height else None,
                        'wing_type': item.custom_wing_type,
                    } if item.is_custom_curtain else None,
                }
                for item in order.items.all()
            ]
        }
        
        return JsonResponse(order_data, status=200)
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Order not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def create_web_order(request):
    """API endpoint to create orders from web checkout"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Get web client (optional for guest checkout)
        from authentication.models import WebClient
        web_client_id = data.get('web_client_id')
        is_guest_order = data.get('is_guest_order', False)
        web_client = None
        
        # For non-guest orders, require web_client
        if not is_guest_order:
            if not web_client_id:
                return JsonResponse({'error': 'web_client_id required for non-guest orders'}, status=400)
            
            try:
                web_client = WebClient.objects.get(pk=web_client_id)
            except WebClient.DoesNotExist:
                return JsonResponse({'error': 'Web client not found'}, status=404)
        
        # Create order
        with transaction.atomic():
            order = Order.objects.create(
                web_client=web_client,  # Can be None for guest orders
                status=data.get('status', 'pending'),
                notes=data.get('notes', ''),
                
                # Guest order info
                is_guest_order=is_guest_order,
                guest_email=data.get('guest_email'),
                guest_phone=data.get('guest_phone'),
                guest_first_name=data.get('guest_first_name'),
                guest_last_name=data.get('guest_last_name'),
                
                # Payment fields
                payment_id=data.get('payment_id'),
                payment_method=data.get('payment_method'),
                payment_status=data.get('payment_status'),
                card_type=data.get('card_type'),
                card_association=data.get('card_association'),
                card_last_four=data.get('card_last_four'),
                
                # Pricing fields
                original_currency=data.get('original_currency'),
                original_price=data.get('original_price'),
                paid_currency=data.get('paid_currency'),
                paid_amount=data.get('paid_amount'),
                exchange_rate=data.get('exchange_rate'),
                
                # Delivery address
                delivery_address_title=data.get('delivery_address_title'),
                delivery_address=data.get('delivery_address'),
                delivery_city=data.get('delivery_city'),
                delivery_country=data.get('delivery_country'),
                delivery_phone=data.get('delivery_phone'),
                
                # Billing address
                billing_address_title=data.get('billing_address_title'),
                billing_address=data.get('billing_address'),
                billing_city=data.get('billing_city'),
                billing_country=data.get('billing_country'),
                billing_phone=data.get('billing_phone'),
            )
            
            # Create order items
            items = data.get('items', [])
            for item_data in items:
                product_sku = item_data.get('product_sku')
                variant_sku = item_data.get('product_variant_sku')
                
                product = None
                variant = None
                
                # Get product
                if product_sku:
                    try:
                        product = Product.objects.get(sku=product_sku)
                    except Product.DoesNotExist:
                        pass

                # Get variant if specified — scope by product first to avoid
                # MultipleObjectsReturned when the same variant_sku string
                # (e.g. "red-m") exists across multiple products.
                if variant_sku:
                    qs = ProductVariant.objects.filter(variant_sku=variant_sku)
                    if product:
                        qs = qs.filter(product=product)
                    variant = qs.first()
                    if not product and variant:
                        product = variant.product
                
                if product:
                    # Get custom curtain attributes if present
                    custom_attrs = item_data.get('custom_attributes', {}) or {}
                    is_custom = item_data.get('is_custom_curtain', False)
                    
                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        product_variant=variant,
                        quantity=item_data.get('quantity', 1),
                        price=item_data.get('price', 0),
                        description=item_data.get('description', ''),
                        # Custom Curtain Fields
                        is_custom_curtain=is_custom,
                        custom_mounting_type=custom_attrs.get('mountingType') if is_custom else None,
                        custom_pleat_type=custom_attrs.get('pleatType') if is_custom else None,
                        custom_pleat_density=custom_attrs.get('pleatDensity') if is_custom else None,
                        custom_width=custom_attrs.get('width') if is_custom else None,
                        custom_height=custom_attrs.get('height') if is_custom else None,
                        custom_wing_type=custom_attrs.get('wingType') if is_custom else None,
                        custom_fabric_used_meters=item_data.get('custom_fabric_used_meters') if is_custom else None,
                    )

            # ── NO warehouse stock-out at web-order create ────────────
            # Web orders also start "Açık" and reserve nothing. Staff
            # process them through packing (scan rolls → reserve) and
            # shipping (cargo info → real cut) in the ERP, same as
            # manual orders.

            return JsonResponse({
                'success': True,
                'order_id': order.pk,
                'order_number': order.order_number or f"DK{str(order.pk).zfill(7)}"
            }, status=201)
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({
            'error': 'Order creation failed',
            'details': str(e)
        }, status=500)


# -------------------------------- Order Tracking API -------------------------------- #

@csrf_exempt
def track_order(request):
    """
    API endpoint for tracking orders by order_number.
    
    GET /operating/orders/track/?order_number=DK0000001
    
    Returns order details including status, tracking info, and items.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'GET method required'}, status=405)
    
    order_number = request.GET.get('order_number', '').strip()
    
    if not order_number:
        return JsonResponse({
            'error': 'order_number parameter is required'
        }, status=400)
    
    try:
        # Try to find by order_number first
        order = Order.objects.filter(order_number__iexact=order_number).first()
        
        # If not found, try to find by ID (for backward compatibility)
        if not order and order_number.isdigit():
            order = Order.objects.filter(pk=int(order_number)).first()
        
        if not order:
            return JsonResponse({
                'success': False,
                'error': 'Sipariş bulunamadı',
                'message': 'Bu sipariş numarası ile eşleşen sipariş bulunamadı.'
            }, status=404)
        
        # Get order items
        items = []
        for item in order.items.all():
            item_data = {
                'product_sku': item.product.sku if item.product else None,
                'product_title': item.product.title if item.product else 'Unknown',
                'product_image': item.product.primary_image.file_url if item.product and item.product.primary_image else None,
                'variant_sku': item.product_variant.variant_sku if item.product_variant else None,
                'quantity': str(item.quantity),
                'price': str(item.price),
                'status': item.status,
                # Custom Curtain Fields
                'is_custom_curtain': item.is_custom_curtain,
                'custom_fabric_used_meters': str(item.custom_fabric_used_meters) if item.custom_fabric_used_meters else None,
                'custom_attributes': {
                    'mounting_type': item.custom_mounting_type,
                    'pleat_type': item.custom_pleat_type,
                    'pleat_density': item.custom_pleat_density,
                    'width': str(item.custom_width) if item.custom_width else None,
                    'height': str(item.custom_height) if item.custom_height else None,
                    'wing_type': item.custom_wing_type,
                } if item.is_custom_curtain else None,
            }
            items.append(item_data)
        
        # Build response
        response_data = {
            'success': True,
            'order': {
                'id': order.pk,
                'order_number': order.order_number or f"DK{str(order.pk).zfill(7)}",
                'status': order.status,
                'order_status': order.order_status,
                'order_status_display': dict(order._meta.get_field('order_status').choices).get(order.order_status, order.order_status) if order.order_status else None,
                'created_at': order.created_at.isoformat() if order.created_at else None,
                'updated_at': order.updated_at.isoformat() if order.updated_at else None,
                
                # Tracking Info
                'carrier': order.carrier,
                'carrier_display': dict(order._meta.get_field('carrier').choices).get(order.carrier, order.carrier) if order.carrier else None,
                'tracking_number': order.tracking_number,
                'shipped_at': order.shipped_at.isoformat() if order.shipped_at else None,
                'delivered_at': order.delivered_at.isoformat() if order.delivered_at else None,
                
                # Pricing
                'original_currency': order.original_currency,
                'original_price': str(order.original_price) if order.original_price else None,
                'paid_currency': order.paid_currency,
                'paid_amount': str(order.paid_amount) if order.paid_amount else None,
                
                # Delivery Address
                'delivery_address': {
                    'title': order.delivery_address_title,
                    'address': order.delivery_address,
                    'city': order.delivery_city,
                    'country': order.delivery_country,
                    'phone': order.delivery_phone,
                },
                
                # Items
                'items': items,
                'total_items': len(items),
            }
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Sipariş sorgulama hatası',
            'details': str(e)
        }, status=500)


@csrf_exempt
def update_order_status(request, order_id):
    """
    API endpoint for updating order status and tracking information.
    
    POST /operating/orders/<order_id>/update-status/
    
    Body:
    {
        "order_status": "shipped",  # pending, confirmed, preparing, shipped, delivered, etc.
        "carrier": "yurtici",       # yurtici, mng, aras, ptt, ups
        "tracking_number": "123456789"
    }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        order = Order.objects.get(pk=order_id)

        data = json.loads(request.body)

        # Funnel through the shared atomic helper so this API can't ship
        # an order without cutting its reserved rolls or without cargo.
        from .views_warehouse import apply_order_status_change
        ok, code = apply_order_status_change(
            order,
            (data.get('order_status') or order.order_status),
            carrier=data.get('carrier'),
            tracking=data.get('tracking_number'),
            user=request.user if getattr(request, 'user', None) and request.user.is_authenticated else None,
        )
        if not ok:
            if code == "cargo_required":
                return JsonResponse({'error': 'Kargo şirketi ve takip numarası gerekli'}, status=400)
            if code == "packing_requires_scan":
                return JsonResponse({'error': 'Paketlemeye geçmek için en az bir top okutulmalı'}, status=400)
            return JsonResponse({'error': 'Durum güncelleme hatası', 'details': (code or '').replace('error:', '')}, status=500)

        return JsonResponse({
            'success': True,
            'message': 'Sipariş durumu güncellendi',
            'order_number': order.order_number,
            'order_status': order.order_status,
            'carrier': order.carrier,
            'tracking_number': order.tracking_number
        })

    except Order.DoesNotExist:
        return JsonResponse({
            'error': 'Sipariş bulunamadı'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({
            'error': 'Durum güncelleme hatası',
            'details': str(e)
        }, status=500)


# ============================================================
# ORDER ANALYTICS DASHBOARD
# ============================================================

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum, Count, Avg, F
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth, TruncYear
from datetime import datetime, timedelta
from decimal import Decimal


class OrderAnalytics(LoginRequiredMixin, View):
    """
    Optimized order analytics dashboard with caching for 10K+ users.
    - Uses Django cache for expensive queries
    - Minimized database queries with values_list
    - Single aggregated queries where possible
    """
    template_name = "operating/order_analytics.html"
    
    # Time-window presets exposed in the UI.
    RANGE_DAYS = {"7d": 7, "30d": 30, "90d": 90, "180d": 180, "365d": 365}

    @staticmethod
    def _granularity(days):
        """Pick chart bucket size so the trend line is readable for any
        window — daily for short ranges, weekly mid, monthly for a year."""
        if days <= 31:
            return "daily"
        if days <= 180:
            return "weekly"
        return "monthly"

    def get(self, request):
        from django.utils import timezone
        from django.db.models import Sum, Count, F, Value, DecimalField
        from django.db.models.functions import Coalesce, TruncDate, TruncWeek, TruncMonth
        from current_account.models import CariAccount
        from datetime import datetime, time

        DEC = DecimalField(max_digits=16, decimal_places=2)

        # ── Filters ──────────────────────────────────────────────
        channel = request.GET.get("channel", "all")        # all | web | manual
        if channel not in ("all", "web", "manual"):
            channel = "all"

        product_type = request.GET.get("product_type", "all")  # all | custom | fabric
        if product_type not in ("all", "custom", "fabric"):
            product_type = "all"

        # Time window — a custom from/to date range (YYYY-MM-DD) takes
        # priority over the preset buttons. The end date is inclusive
        # (covers the whole day). If only one bound or an invalid date
        # is given we fall back to the preset range.
        custom_start_raw = (request.GET.get("start") or "").strip()
        custom_end_raw = (request.GET.get("end") or "").strip()
        custom_range = False
        custom_start = custom_end = ""
        rng = request.GET.get("range", "30d")
        if rng not in self.RANGE_DAYS:
            rng = "30d"

        if custom_start_raw and custom_end_raw:
            try:
                sd = datetime.strptime(custom_start_raw, "%Y-%m-%d").date()
                ed = datetime.strptime(custom_end_raw, "%Y-%m-%d").date()
                if sd > ed:
                    sd, ed = ed, sd  # tolerate reversed inputs
                start_date = timezone.make_aware(datetime.combine(sd, time.min))
                end_date = timezone.make_aware(datetime.combine(ed, time.max))
                days = max((ed - sd).days, 1)
                custom_range = True
                custom_start, custom_end = sd.isoformat(), ed.isoformat()
            except (ValueError, TypeError):
                custom_range = False

        if not custom_range:
            days = self.RANGE_DAYS[rng]
            end_date = timezone.now()
            start_date = end_date - timedelta(days=days)

        # Sales = orders in window, excluding cancelled / returned (not revenue).
        EXCLUDED = ["cancelled", "returned"]
        base_orders = Order.objects.filter(
            created_at__gte=start_date, created_at__lte=end_date
        ).exclude(order_status__in=EXCLUDED)

        # Channel predicates — partition EVERY order so web + manual
        # always reconciles with the total. Web = placed through the
        # online store (has a web_client). Manual = anything entered by
        # staff (no web_client), whether or not a customer is attached
        # yet. This matches the business meaning: "internet satışı" vs
        # "manuel satış".
        web_q = Q(web_client_id__isnull=False)
        manual_q = Q(web_client_id__isnull=True)

        if channel == "web":
            filtered_orders = base_orders.filter(web_q)
        elif channel == "manual":
            filtered_orders = base_orders.filter(manual_q)
        else:
            filtered_orders = base_orders
        filtered_ids = list(filtered_orders.values_list("id", flat=True))

        # ── Money — computed from OrderItem lines so it's correct for
        #    BOTH web orders (carry original_price) and manual/B2B orders
        #    (which don't). unit_cost mirrors OrderItem.unit_cost():
        #    variant_cost preferred, else product.cost, else 0. ──
        unit_cost_expr = Coalesce(
            F("product_variant__variant_cost"), F("product__cost"),
            Value(Decimal("0")), output_field=DEC,
        )
        line_rev = F("quantity") * F("price")
        line_cost = F("quantity") * unit_cost_expr

        items = OrderItem.objects.filter(order_id__in=filtered_ids)
        if product_type == "custom":
            items = items.filter(is_custom_curtain=True)
        elif product_type == "fabric":
            items = items.filter(is_custom_curtain=False)

        money = items.aggregate(
            revenue=Coalesce(Sum(line_rev, output_field=DEC), Value(Decimal("0"), output_field=DEC)),
            cost=Coalesce(Sum(line_cost, output_field=DEC), Value(Decimal("0"), output_field=DEC)),
        )
        total_revenue = money["revenue"] or Decimal("0")
        total_cost = money["cost"] or Decimal("0")
        total_profit = total_revenue - total_cost
        total_orders = filtered_orders.count()
        avg_order = (total_revenue / total_orders) if total_orders else Decimal("0")
        margin_pct = (total_profit / total_revenue * 100) if total_revenue else Decimal("0")

        # ── Channel split — always from base_orders so both web & manual
        #    show regardless of the channel filter selection. ──
        def _channel_money(qfilter):
            ids = list(base_orders.filter(qfilter).values_list("id", flat=True))
            rev = OrderItem.objects.filter(order_id__in=ids).aggregate(
                r=Coalesce(Sum(line_rev, output_field=DEC), Value(Decimal("0"), output_field=DEC))
            )["r"] or Decimal("0")
            return {"orders": len(ids), "revenue": float(rev)}
        web_split = _channel_money(web_q)
        manual_split = _channel_money(manual_q)

        # ── Trend line ──
        gran = self._granularity(days)
        trunc = {"daily": TruncDate, "weekly": TruncWeek, "monthly": TruncMonth}[gran]("order__created_at")
        trend = list(
            OrderItem.objects.filter(order_id__in=filtered_ids)
            .annotate(bucket=trunc)
            .values("bucket")
            .annotate(
                revenue=Sum(line_rev, output_field=DEC),
                orders=Count("order", distinct=True),
            )
            .order_by("bucket")
        )
        fmt = {"daily": "%d %b", "weekly": "%d %b", "monthly": "%b %Y"}[gran]
        trend_labels, trend_rev, trend_cnt = [], [], []
        for t in trend:
            if t["bucket"]:
                trend_labels.append(t["bucket"].strftime(fmt))
                trend_rev.append(float(t["revenue"] or 0))
                trend_cnt.append(t["orders"])

        # ── Top products (by revenue) ──
        top_products = list(
            items.values("product__title", "product__sku")
            .annotate(
                qty=Coalesce(Sum("quantity"), Value(Decimal("0"), output_field=DEC)),
                revenue=Sum(line_rev, output_field=DEC),
                orders=Count("order", distinct=True),
            )
            .order_by("-revenue")[:12]
        )
        for p in top_products:
            p["qty"] = float(p["qty"] or 0)
            p["revenue"] = float(p["revenue"] or 0)
        prod_labels = [(p["product__title"] or "—")[:24] for p in top_products[:8]]
        prod_rev = [p["revenue"] for p in top_products[:8]]

        # ── Top customers — merge company / contact / web-client rankings ──
        cust = []
        for r in (OrderItem.objects.filter(order_id__in=filtered_ids, order__company__isnull=False)
                  .values("order__company__id", "order__company__name")
                  .annotate(spent=Sum(line_rev, output_field=DEC), orders=Count("order", distinct=True))):
            cust.append({"kind": "company", "id": r["order__company__id"],
                         "name": r["order__company__name"] or "—",
                         "spent": float(r["spent"] or 0), "orders": r["orders"]})
        for r in (OrderItem.objects.filter(order_id__in=filtered_ids,
                                            order__company__isnull=True,
                                            order__contact__isnull=False)
                  .values("order__contact__id", "order__contact__name")
                  .annotate(spent=Sum(line_rev, output_field=DEC), orders=Count("order", distinct=True))):
            cust.append({"kind": "contact", "id": r["order__contact__id"],
                         "name": r["order__contact__name"] or "—",
                         "spent": float(r["spent"] or 0), "orders": r["orders"]})
        for r in (OrderItem.objects.filter(order_id__in=filtered_ids, order__web_client__isnull=False)
                  .values("order__web_client__id", "order__web_client__name", "order__web_client__username")
                  .annotate(spent=Sum(line_rev, output_field=DEC), orders=Count("order", distinct=True))):
            nm = r["order__web_client__name"] or r["order__web_client__username"] or "Web customer"
            cust.append({"kind": "web", "id": r["order__web_client__id"],
                         "name": nm, "spent": float(r["spent"] or 0), "orders": r["orders"]})
        cust.sort(key=lambda c: c["spent"], reverse=True)
        top_customers = cust[:10]

        # Attach cari snapshot for company / contact customers.
        comp_ids = [c["id"] for c in top_customers if c["kind"] == "company"]
        cont_ids = [c["id"] for c in top_customers if c["kind"] == "contact"]
        cari_by_company, cari_by_contact = {}, {}
        if comp_ids:
            for ca in CariAccount.objects.filter(company_id__in=comp_ids).select_related("default_currency"):
                cari_by_company.setdefault(ca.company_id, ca)
        if cont_ids:
            for ca in CariAccount.objects.filter(contact_id__in=cont_ids).select_related("default_currency"):
                cari_by_contact.setdefault(ca.contact_id, ca)
        for c in top_customers:
            ca = cari_by_company.get(c["id"]) if c["kind"] == "company" else \
                (cari_by_contact.get(c["id"]) if c["kind"] == "contact" else None)
            if ca:
                c["cari"] = {
                    "pk": ca.pk, "code": ca.code, "name": ca.name,
                    "balance": float(ca.cached_balance or 0),
                    "abs_balance": float(ca.absolute_balance or 0),
                    "label": str(ca.balance_label),
                    "symbol": ca.display_currency_symbol or "",
                }
            else:
                c["cari"] = None

        # ── Status distribution (for the mix bar) ──
        status_rows = list(base_orders.values("order_status").annotate(n=Count("id")).order_by("-n"))

        context = {
            "summary": {
                "total_orders": total_orders,
                "total_revenue": float(total_revenue),
                "total_profit": float(total_profit),
                "total_cost": float(total_cost),
                "avg_order": float(avg_order),
                "margin_pct": round(float(margin_pct), 1),
            },
            "web_split": web_split,
            "manual_split": manual_split,
            "trend_labels": json.dumps(trend_labels),
            "trend_rev": json.dumps(trend_rev),
            "trend_cnt": json.dumps(trend_cnt),
            "channel_chart": json.dumps([web_split["revenue"], manual_split["revenue"]]),
            "prod_labels": json.dumps(prod_labels),
            "prod_rev": json.dumps(prod_rev),
            "top_products": top_products,
            "top_customers": top_customers,
            "status_rows": status_rows,
            "range": rng,
            "channel": channel,
            "product_type": product_type,
            "granularity": gran,
            "custom_range": custom_range,
            "custom_start": custom_start,
            "custom_end": custom_end,
        }
        return render(request, self.template_name, context)


class WebOrderStatusEdit(View):
    """
    Dedicated view for updating shipping status of web orders.
    Only accessible for  orders with web_client (web orders).
    """
    template_name = "operating/web_order_status.html"
    
    def get(self, request, pk):
        from .models import ORDER_STATUS_CHOICES, CARRIER_CHOICES
        
        order = get_object_or_404(Order, pk=pk)
        
        # Only allow web orders
        if not order.web_client and not order.is_guest_order:
            messages.error(request, "This is not a web order.")
            return redirect('operating:order_detail', pk=pk)
        
        # English status choices for admin UI
        status_choices_en = [
            ("pending", "Pending"),
            ("confirmed", "Confirmed"),
            ("preparing", "Preparing"),
            ("shipped", "Shipped"),
            ("in_transit", "In Transit"),
            ("out_for_delivery", "Out for Delivery"),
            ("delivered", "Delivered"),
            ("cancelled", "Cancelled"),
            ("returned", "Returned"),
        ]
        
        # English carrier choices
        carrier_choices_en = [
            ("yurtici", "Yurtiçi Kargo"),
            ("mng", "MNG Kargo"),
            ("aras", "Aras Kargo"),
            ("ptt", "PTT Kargo"),
            ("ups", "UPS"),
            ("other", "Other"),
        ]
        
        context = {
            'order': order,
            'status_choices': status_choices_en,
            'carrier_choices': carrier_choices_en,
        }
        return render(request, self.template_name, context)
    
    def post(self, request, pk):
        from django.utils import timezone
        
        order = get_object_or_404(Order, pk=pk)
        
        # Only allow web orders
        if not order.web_client and not order.is_guest_order:
            messages.error(request, "This is not a web order.")
            return redirect('operating:order_detail', pk=pk)
        
        # English status choices for admin UI
        status_choices_en = [
            ("pending", "Pending"),
            ("confirmed", "Confirmed"),
            ("preparing", "Preparing"),
            ("shipped", "Shipped"),
            ("in_transit", "In Transit"),
            ("out_for_delivery", "Out for Delivery"),
            ("delivered", "Delivered"),
            ("cancelled", "Cancelled"),
            ("returned", "Returned"),
        ]
        
        # English carrier choices
        carrier_choices_en = [
            ("yurtici", "Yurtiçi Kargo"),
            ("mng", "MNG Kargo"),
            ("aras", "Aras Kargo"),
            ("ptt", "PTT Kargo"),
            ("ups", "UPS"),
            ("other", "Other"),
        ]
        
        # Get form data
        new_status = request.POST.get('order_status')
        carrier = request.POST.get('carrier')
        tracking_number = request.POST.get('tracking_number')

        # Same atomic funnel as the detail page + JSON API: cargo gate +
        # reservation → stock-out conversion can't be bypassed here.
        from .views_warehouse import apply_order_status_change
        ok, code = apply_order_status_change(
            order, (new_status or order.order_status),
            carrier=carrier, tracking=tracking_number, user=request.user,
        )
        if not ok:
            if code == "cargo_required":
                messages.error(request, "Siparişi tamamlamak için kargo şirketi ve takip numarası gerekli.")
            elif code == "packing_requires_scan":
                messages.error(request, "Paketlemeye geçmek için önce en az bir top okutun.")
            else:
                messages.error(request, f"Sipariş güncellenemedi: {(code or '').replace('error:', '')}")
            next_url = request.POST.get('next') or request.GET.get('next')
            return redirect(next_url or reverse('operating:order_detail', kwargs={'pk': order.pk}))

        status_label = dict(status_choices_en).get(new_status, new_status)
        messages.success(request, f"Order status updated successfully: {status_label}")

        # Honour ?next= so callers can return the user to the detail page
        # (where the inline status changer lives) instead of the dedicated
        # legacy edit screen.
        next_url = request.POST.get('next') or request.GET.get('next')
        if next_url:
            return redirect(next_url)
        # Default: bounce back to the order detail page.
        return redirect('operating:order_detail', pk=order.pk)

# -----------------------------------------------------------------------------
# HTMX / API Views
# -----------------------------------------------------------------------------

def raw_material_search(request):
    """
    Returns JSON list of raw materials matching the query.
    Used for BOM autocomplete.
    """
    query = request.GET.get("q", "")
    if len(query) < 2:
        return JsonResponse([], safe=False)

    materials = RawMaterialGood.objects.filter(
        Q(name__icontains=query) | Q(sku__icontains=query)
    ).prefetch_related('items', 'items__receipt').values("id", "name", "sku", "unit_of_measurement", "quantity")[:10]

    results = []
    for mat in materials:
        # Calculate latest cost
        # We can't easily do this in the header query without complex subqueries or annotations
        # Since we limit to 10, a separate query per item is acceptable or we use the prefetch if we iterate objects
        # But we used .values(). Let's fetch objects instead to use methods or easy related lookups
        pass 

    # Re-doing the query to return objects to easily property/related access
    materials_objs = RawMaterialGood.objects.filter(
        Q(name__icontains=query) | Q(sku__icontains=query)
    ).prefetch_related('items', 'items__receipt')[:10]

    for mat in materials_objs:
        # Find latest item by receipt date
        latest_item = mat.items.select_related('receipt').order_by('-receipt__date', '-id').first()
        cost = latest_item.unit_cost if latest_item else 0

        results.append({
            "id": mat.id,
            "name": mat.name,
            "sku": mat.sku,
            "unit": mat.unit_of_measurement,
            "stock": mat.quantity or 0,
            "cost": cost
        })
    
    return JsonResponse(results, safe=False)


class RawMaterialGoodList(ListView):
    model = RawMaterialGood
    template_name = "operating/raw_material_good_list.html"
    context_object_name = "raw_materials"
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '').strip()
        
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(sku__icontains=search_query) |
                Q(supplier_sku__icontains=search_query)
            )
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context


@require_POST
def create_raw_material_good_json(request):
    from django.forms import modelform_factory
    
    try:
        RawMaterialGoodForm = modelform_factory(RawMaterialGood, fields="__all__")
        form = RawMaterialGoodForm(request.POST, request.FILES)
        
        if form.is_valid():
            instance = form.save()
            return JsonResponse({
                "success": True,
                "message": f"Raw Material '{instance.name}' created successfully.",
                "data": {
                    "id": instance.pk,
                    "name": instance.name,
                    "sku": instance.sku
                }
            })
        else:
            # Convert form errors to a simple string
            errors = dict(form.errors.items())
            error_msg = "Validation failed"
            if errors:
                first_field = list(errors.keys())[0]
                first_error = errors[first_field][0]
                error_msg = f"{first_field}: {first_error}"
                
            return JsonResponse({
                "success": False,
                "errors": error_msg,
                "field_errors": errors
            })
            
    except Exception as e:
        return JsonResponse({"success": False, "errors": str(e)})


@require_POST
def create_raw_material_good_receipt_json(request):
    from django.forms import modelform_factory
    
    try:
        RawMaterialGoodReceiptForm = modelform_factory(RawMaterialGoodReceipt, fields="__all__")
        form = RawMaterialGoodReceiptForm(request.POST, request.FILES)
        
        if form.is_valid():
            instance = form.save()
            return JsonResponse({
                "success": True,
                "message": f"Receipt #{instance.receipt_number} from {instance.supplier} created successfully.",
                "data": {
                    "id": instance.pk,
                    "receipt_number": instance.receipt_number,
                    "supplier": str(instance.supplier)
                }
            })
        else:
            errors = dict(form.errors.items())
            error_msg = "Validation failed"
            if errors:
                first_field = list(errors.keys())[0]
                first_error = errors[first_field][0]
                error_msg = f"{first_field}: {first_error}"
                
            return JsonResponse({
                "success": False,
                "errors": error_msg,
                "field_errors": errors
            })
            
    except Exception as e:
        return JsonResponse({"success": False, "errors": str(e)})


@require_POST
def create_raw_material_good_item_json(request):
    from django.forms import modelform_factory
    
    try:
        RawMaterialGoodItemForm = modelform_factory(RawMaterialGoodItem, fields="__all__")
        form = RawMaterialGoodItemForm(request.POST, request.FILES)
        
        if form.is_valid():
            instance = form.save()
            return JsonResponse({
                "success": True,
                "message": f"Added {instance.quantity} of {instance.raw_material_good.name} successfully.",
                "data": {
                    "id": instance.pk,
                    "quantity": instance.quantity,
                    "raw_material": instance.raw_material_good.name
                }
            })
        else:
            errors = dict(form.errors.items())
            error_msg = "Validation failed"
            if errors:
                first_field = list(errors.keys())[0]
                first_error = errors[first_field][0]
                error_msg = f"{first_field}: {first_error}"
                
            return JsonResponse({
                "success": False,
                "errors": error_msg,
                "field_errors": errors
            })
            
    except Exception as e:
        return JsonResponse({"success": False, "errors": str(e)})


def get_raw_material_good_json(request, pk):
    raw_material = get_object_or_404(RawMaterialGood, pk=pk)
    return JsonResponse({
        "success": True,
        "data": {
            "id": raw_material.pk,
            "name": raw_material.name,
            "sku": raw_material.sku,
            "supplier_sku": raw_material.supplier_sku or "",
            "raw_type": raw_material.raw_type,
            "unit_of_measurement": raw_material.unit_of_measurement,
            "quantity": raw_material.quantity
        }
    })


@require_POST
def update_raw_material_good_json(request, pk):
    raw_material = get_object_or_404(RawMaterialGood, pk=pk)
    from django.forms import modelform_factory
    
    try:
        RawMaterialGoodForm = modelform_factory(RawMaterialGood, fields="__all__")
        form = RawMaterialGoodForm(request.POST, request.FILES, instance=raw_material)
        
        if form.is_valid():
            instance = form.save()
            return JsonResponse({
                "success": True,
                "message": f"Raw Material '{instance.name}' updated successfully."
            })
        else:
            errors = dict(form.errors.items())
            error_msg = "Validation failed"
            if errors:
                first_field = list(errors.keys())[0]
                first_error = errors[first_field][0]
                error_msg = f"{first_field}: {first_error}"
                
            return JsonResponse({
                "success": False,
                "errors": error_msg,
                "field_errors": errors
            })
            
    except Exception as e:
        return JsonResponse({"success": False, "errors": str(e)})


@require_POST
def delete_raw_material_good_json(request, pk):
    try:
        raw_material = get_object_or_404(RawMaterialGood, pk=pk)
        raw_material.delete()
        return JsonResponse({
            "success": True, 
            "message": "Raw material deleted successfully."
        })
    except Exception as e:
        return JsonResponse({"success": False, "errors": str(e)})


def create_raw_material_receipt_partial(request):
    from django.forms import modelform_factory
    from django import forms
    from datetime import date
    RawMaterialGoodReceiptForm = modelform_factory(RawMaterialGoodReceipt, fields="__all__", widgets={
        "date": forms.DateInput(attrs={"type": "date"}),
    })
    form = RawMaterialGoodReceiptForm(initial={"date": date.today().strftime("%Y-%m-%d")})
    return render(request, "operating/partials/raw_material_receipt_form.html", {"form": form})


def create_raw_material_item_partial(request):
    from django.forms import modelform_factory
    RawMaterialGoodItemForm = modelform_factory(RawMaterialGoodItem, fields="__all__")
    
    # Pre-fill receipt if provided in GET
    initial = {}
    if 'receipt' in request.GET:
        initial['receipt'] = request.GET.get('receipt')
        
    form = RawMaterialGoodItemForm(initial=initial)
    return render(request, "operating/partials/raw_material_item_form.html", {"form": form})
