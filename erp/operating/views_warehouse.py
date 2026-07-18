"""Warehouse views — list, detail, create, import products from Excel."""
import json
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, F, DecimalField
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views import View

from .models import Warehouse, WarehouseProduct, WarehouseProductRoll, StockMovement


def _env(key, default=None):
    """Read an env var that may live in either os.environ OR the
    project's .env file (loaded by python-decouple in settings.py)."""
    import os
    val = os.environ.get(key)
    if val:
        return val
    try:
        from decouple import config as _cfg
        return _cfg(key, default=default)
    except Exception:
        return default


def _tr_upper(s):
    """Turkish-aware uppercase: i→İ and ı→I (Python's str.upper() turns
    i→I, dropping the dot). Used for SKUs like 'K48083İ.G93' so the dotted
    İ survives instead of becoming a dotless I."""
    if not s:
        return s
    return s.replace("ı", "I").replace("i", "İ").upper()


def _is_admin(user):
    """Only admins may DELETE warehouses / products — editing stays open to
    everyone. Admin = Django superuser/staff, or a Member carrying the
    'admin' permission (authentication.Permission name='admin')."""
    if not getattr(user, "is_authenticated", False):
        return False
    if user.is_superuser or user.is_staff:
        return True
    try:
        return user.member.permissions.filter(name="admin").exists()
    except Exception:
        return False


def _safe_decimal(value, default=None):
    """Convert any cell value to Decimal, returning default on failure."""
    if value is None or value == '':
        return default
    try:
        if isinstance(value, str):
            cleaned = value.strip().replace(',', '.').replace(' ', '')
            return Decimal(cleaned)
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return default


def _get_usd_try_rate():
    """Best-effort USD->TRY rate. Falls back to 1 if unavailable."""
    try:
        from accounting.services import get_exchange_rate
        rate = get_exchange_rate("USD", "TRY")
        if rate:
            return Decimal(str(rate))
    except Exception:
        pass
    return Decimal('1')


def _consonant_prefix(name):
    """Barcode prefix = the first 3 CONSONANTS of a supplier name, upper-cased
    and ASCII-folded (Turkish-aware). "Kızılırmak" -> "KZL", "Acme" -> "CM".
    Falls back to "GEN" when a name has no usable consonants."""
    from .catalog_sync import _fold
    folded = _fold(name or "")               # Turkish-aware UPPER + ASCII fold
    vowels = set("AEIOU")
    cons = [c for c in folded if c.isalpha() and c not in vowels]
    return ("".join(cons[:3]) or "GEN")[:6]


def _barcode_minter(prefix):
    """Return a closure that mints fresh, GLOBALLY-UNIQUE barcodes of the form
    PREFIX + 6 digits (e.g. KZL000123). Uniqueness is checked against every
    existing roll AND product barcode, plus the ones minted in this batch, so
    each top across the whole system gets a one-of-a-kind code."""
    import re as _re
    prefix = (prefix or "GEN").upper()[:6] or "GEN"
    existing = set()
    for qs in (
        WarehouseProductRoll.objects.filter(barcode__startswith=prefix)
        .values_list("barcode", flat=True),
        WarehouseProduct.objects.filter(barcode__startswith=prefix)
        .values_list("barcode", flat=True),
    ):
        existing.update(b for b in qs if b)
    # Seed the counter only from OUR own clean "PREFIX######" sequence — never
    # from unrelated long codes that merely start with the same letters (e.g.
    # a supplier whose consonants match an existing SKU prefix). Uniqueness is
    # still enforced against EVERY existing barcode via the `existing` set.
    seq_pat = _re.compile(r"^" + _re.escape(prefix) + r"(\d{6,})$")
    maxn = 0
    for b in existing:
        m = seq_pat.match(b or "")
        if m and len(m.group(1)) == 6:
            try:
                maxn = max(maxn, int(m.group(1)))
            except ValueError:
                pass
    state = {"n": maxn}

    def mint():
        while True:
            state["n"] += 1
            code = f"{prefix}{state['n']:06d}"
            if code not in existing:
                existing.add(code)
                return code

    return mint


def _slug_token(text):
    """Compact ASCII-upper alnum token for an auto variant-SKU suffix:
    'Açık Mavi' -> 'ACIKMAVI'."""
    from .catalog_sync import _fold
    folded = _fold(text or "")
    return "".join(ch for ch in folded if ch.isalnum())[:10]


def _supplier_choices():
    """Suppliers for the manual-add dropdown, each with its derived barcode
    prefix so the UI can preview it instantly."""
    try:
        from crm.models import Supplier
        rows = (Supplier.objects.all()
                .order_by("company_name", "contact_name")[:500])
        return [{"id": s.id, "name": str(s), "prefix": _consonant_prefix(str(s))}
                for s in rows]
    except Exception:
        return []


# Manual-add unit → a valid marketing.Product.unit_of_measurement choice
# (the model only allows units/mt/kg; the real per-roll quantity is generic).
_PRODUCT_UNIT_MAP = {"mt": "mt", "kg": "kg", "adet": "units", "paket": "units", "units": "units"}


def _product_sku_minter(prefix):
    """Closure that mints GLOBALLY-UNIQUE marketing.Product SKUs of the form
    PREFIX + 3+ digits (e.g. KZL004) — same supplier prefix as the barcodes.
    Mirrors _barcode_minter but targets the DB-unique Product.sku column,
    case-insensitively (matching catalog_sync._safe_sku's sku__iexact). Seeds
    the counter from our own clean "PREFIX###" sequence so unrelated typed SKUs
    that merely share the prefix don't inflate it, while collision-checking
    against EVERY existing Product.sku so a minted code never equals one. The
    caller must STILL guard create() with an IntegrityError retry — the DB
    UNIQUE constraint is the final arbiter under concurrency."""
    import re as _re
    from marketing.models import Product as _Prod
    prefix = (prefix or "GEN").upper()[:6] or "GEN"
    existing = set()
    for s in _Prod.objects.filter(sku__istartswith=prefix).values_list("sku", flat=True):
        if s:
            existing.add(s.upper())
    # Seed ONLY from OUR own clean "PREFIX" + 3..5-digit sequence (e.g. KZL004),
    # never from long real product codes that merely share the prefix (the
    # warehouse already has 8-digit "KZL36900959"-style SKUs). Uniqueness is
    # still enforced against EVERY existing sku via the `existing` set.
    seq_pat = _re.compile(r"^" + _re.escape(prefix) + r"(\d{3,5})$")
    maxn = 0
    for s in existing:
        m = seq_pat.match(s)
        if m:
            try:
                maxn = max(maxn, int(m.group(1)))
            except ValueError:
                pass
    state = {"n": maxn}

    def mint():
        while True:
            state["n"] += 1
            code = f"{prefix}{state['n']:03d}"[:20]
            if code.upper() not in existing:
                existing.add(code.upper())
                return code

    return mint


def _variant_sku_exists(sku):
    """True if a marketing ProductVariant already owns this (globally-unique)
    variant_sku — used to keep AUTO-generated variant SKUs collision-free."""
    try:
        from marketing.models import ProductVariant
        return ProductVariant.objects.filter(variant_sku__iexact=sku).exists()
    except Exception:
        return False


def _merge_warehouse_dupes_by_sku(warehouse, sku, keep=None):
    """Consolidate every WarehouseProduct in `warehouse` that shares `sku`
    (case-insensitive) into ONE record — the same SKU is the same variant and
    must never show as two rows. Moves the duplicates' rolls (tops) + stock
    movements onto the survivor, recomputes its quantity from all its rolls,
    keeps a single catalog variant (deleting any now-orphaned hidden ones +
    their emptied parent), and deletes the emptied duplicates. Survivor = `keep`
    when given, else the lowest-id match. Returns (survivor, merged_count)."""
    from django.db import transaction as _tx
    sku = (sku or "").strip()
    if not sku:
        return keep, 0
    dupes = list(WarehouseProduct.objects
                 .filter(warehouse=warehouse, sku__iexact=sku).order_by("id"))
    if len(dupes) <= 1:
        return (keep or (dupes[0] if dupes else None)), 0
    survivor = keep if (keep and any(d.pk == keep.pk for d in dupes)) else dupes[0]
    orphan_variant_ids = set()
    merged = 0
    with _tx.atomic():
        for dup in dupes:
            if dup.pk == survivor.pk:
                continue
            WarehouseProductRoll.objects.filter(product=dup).update(product=survivor)
            StockMovement.objects.filter(product=dup).update(product=survivor)
            if not survivor.catalog_variant_id and dup.catalog_variant_id:
                survivor.catalog_variant_id = dup.catalog_variant_id
            elif dup.catalog_variant_id and dup.catalog_variant_id != survivor.catalog_variant_id:
                orphan_variant_ids.add(dup.catalog_variant_id)
            dup.delete()
            merged += 1
        # Survivor quantity = sum of remaining metres across ALL its rolls.
        total = Decimal("0")
        for r in survivor.rolls.all():
            rem = r.meters_remaining if r.meters_remaining is not None else (r.meters or Decimal("0"))
            total += rem or Decimal("0")
        survivor.quantity = total
        survivor.save(update_fields=["quantity", "catalog_variant", "updated_at"])
        # Catalog cleanup: drop variants no warehouse product points at anymore,
        # then mirror the survivor's stock onto its variant + parent product.
        try:
            from marketing.models import ProductVariant, Product as _Prod
            from django.db.models import Sum as _Sum
            for vid in orphan_variant_ids:
                if WarehouseProduct.objects.filter(catalog_variant_id=vid).exists():
                    continue
                v = ProductVariant.objects.filter(pk=vid).first()
                if not v:
                    continue
                pid = v.product_id
                v.delete()
                if (_Prod.objects.filter(pk=pid, featured=False).exists()
                        and not ProductVariant.objects.filter(product_id=pid).exists()):
                    _Prod.objects.filter(pk=pid).delete()
            if survivor.catalog_variant_id:
                ProductVariant.objects.filter(pk=survivor.catalog_variant_id).update(variant_quantity=total)
                pv = ProductVariant.objects.filter(pk=survivor.catalog_variant_id).first()
                if pv:
                    agg = ProductVariant.objects.filter(product_id=pv.product_id).aggregate(s=_Sum("variant_quantity"))
                    _Prod.objects.filter(pk=pv.product_id).update(quantity=agg["s"] or Decimal("0"))
        except Exception:
            import traceback
            traceback.print_exc()
    return survivor, merged


@method_decorator(login_required, name='dispatch')
class WarehouseList(View):
    template_name = "operating/warehouse_list.html"

    def get(self, request):
        warehouses = Warehouse.objects.annotate(
            n_products=Count('products'),
        ).order_by('name')
        # "Son Hareketler" preview — the latest activity across every
        # warehouse, linking to the full filterable feed.
        recent = _decorate_movements(list(
            StockMovement.objects
            .select_related("product", "product__warehouse", "created_by")
            .order_by("-created_at")[:6]
        ))
        return render(request, self.template_name, {
            'warehouses': warehouses,
            'recent_movements': recent,
        })


def _get_book_choices():
    """Return list of accounting Book objects for dropdown (lazy import)."""
    try:
        from accounting.models import Book
        return Book.objects.all().order_by('name')
    except Exception:
        return []


@method_decorator(login_required, name='dispatch')
class WarehouseCreate(View):
    template_name = "operating/warehouse_form.html"

    def get(self, request):
        return render(request, self.template_name, {
            'warehouse': None,
            'books': _get_book_choices(),
        })

    def post(self, request):
        name = (request.POST.get('name') or '').strip()
        location = (request.POST.get('location') or '').strip()
        description = (request.POST.get('description') or '').strip()
        book_id = (request.POST.get('accounting_book') or '').strip()

        if not name:
            messages.error(request, "Name is required")
            return render(request, self.template_name, {
                'warehouse': None, 'books': _get_book_choices()
            })
        if Warehouse.objects.filter(name__iexact=name).exists():
            messages.error(request, "A warehouse with this name already exists")
            return render(request, self.template_name, {
                'warehouse': None, 'books': _get_book_choices()
            })

        book = None
        if book_id:
            try:
                from accounting.models import Book
                book = Book.objects.filter(pk=int(book_id)).first()
            except (ValueError, TypeError):
                pass

        wh = Warehouse.objects.create(
            name=name,
            location=location or None,
            description=description or None,
            accounting_book=book,
        )
        messages.success(request, f"Warehouse '{wh.name}' created")
        return redirect(reverse('operating:warehouse_detail', args=[wh.pk]))


@method_decorator(login_required, name='dispatch')
class WarehouseCreatePartial(View):
    """Sidebar-mounted Add Warehouse form.

    GET  → returns just the inner form HTML (loaded via HTMX into the
           sidebar overlay in base.html).
    POST → expects X-Requested-With:XMLHttpRequest and returns JSON
           {success, warehouse_id, redirect_url} so the front-end can
           show a toast and either close the sidebar or redirect.
    """
    template_name = "operating/_warehouse_form_partial.html"

    def get(self, request):
        return render(request, self.template_name, {
            'books': _get_book_choices(),
        })

    def post(self, request):
        name = (request.POST.get('name') or '').strip()
        location = (request.POST.get('location') or '').strip()
        description = (request.POST.get('description') or '').strip()
        book_id = (request.POST.get('accounting_book') or '').strip()

        if not name:
            return JsonResponse({"success": False, "errors": {"name": "Name is required"}}, status=400)
        if Warehouse.objects.filter(name__iexact=name).exists():
            return JsonResponse({
                "success": False,
                "errors": {"name": "A warehouse with this name already exists"},
            }, status=400)

        book = None
        if book_id:
            try:
                from accounting.models import Book
                book = Book.objects.filter(pk=int(book_id)).first()
            except (ValueError, TypeError):
                pass

        wh = Warehouse.objects.create(
            name=name,
            location=location or None,
            description=description or None,
            accounting_book=book,
        )
        return JsonResponse({
            "success": True,
            "warehouse_id": wh.pk,
            "name": wh.name,
            "redirect_url": reverse('operating:warehouse_detail', args=[wh.pk]),
        })


@method_decorator(login_required, name='dispatch')
class WarehouseEdit(View):
    template_name = "operating/warehouse_form.html"

    def get(self, request, pk):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        return render(request, self.template_name, {
            'warehouse': warehouse,
            'books': _get_book_choices(),
        })

    def post(self, request, pk):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        name = (request.POST.get('name') or '').strip()
        location = (request.POST.get('location') or '').strip()
        description = (request.POST.get('description') or '').strip()
        book_id = (request.POST.get('accounting_book') or '').strip()

        if not name:
            messages.error(request, "Name is required")
            return render(request, self.template_name, {
                'warehouse': warehouse, 'books': _get_book_choices()
            })

        # Allow same name as own, but block conflict with others
        if Warehouse.objects.filter(name__iexact=name).exclude(pk=warehouse.pk).exists():
            messages.error(request, "Another warehouse with this name already exists")
            return render(request, self.template_name, {
                'warehouse': warehouse, 'books': _get_book_choices()
            })

        warehouse.name = name
        warehouse.location = location or None
        warehouse.description = description or None

        if book_id:
            try:
                from accounting.models import Book
                warehouse.accounting_book = Book.objects.filter(pk=int(book_id)).first()
            except (ValueError, TypeError):
                warehouse.accounting_book = None
        else:
            warehouse.accounting_book = None

        warehouse.save()
        messages.success(request, f"Warehouse '{warehouse.name}' updated")
        return redirect(reverse('operating:warehouse_detail', args=[warehouse.pk]))


from django.utils.translation import gettext_lazy as _lz
# The sort options offered on the warehouse product list. Kept simple:
# alphabetical by name, by top (roll) quantity, and most-recently-updated.
SORT_OPTIONS = {
    'name_asc':   ('name', _lz('A → Z')),
    'name_desc':  ('-name', _lz('Z → A')),
    'qty_desc':   ('-quantity', _lz('Top quantity')),
    'price_desc': ('-cost_usd', _lz('Most expensive')),
    'price_asc':  ('cost_usd', _lz('Cheapest')),
    'recent':     ('-updated_at', _lz('Recently updated')),
}


PAGE_SIZE = 50


def _warehouse_base_expr():
    """Group key for the warehouse list = the MAIN product.

    Linked rows use the real catalog Product title. Unlinked rows fall
    back to the SAME rule the catalog uses (see ``derive_catalog``):

      * SKU has a dot  → the part BEFORE the first dot is the main product
        (``K24892İ.G157`` → ``K24892İ``, ``K24620.G33`` → ``K24620``).
        This is why GREK TUL / GREK TÜL / GREK TUL FIXE must NOT merge —
        their SKU bases (K24620 / K24892İ / K24892) differ.
      * SKU has no dot → the name's first word (``MT-3016 GÜMÜŞ`` → ``MT-3016``).

    The two branches agree with the catalog titles, so a linked roll and
    an unlinked sibling of the same product still land in one group.
    """
    from django.db.models import Func, Value, CharField, Case, When, Q, F
    from django.db.models.functions import Coalesce

    class _SplitPart(Func):
        function = "split_part"
        output_field = CharField()

    return Coalesce(
        F("catalog_variant__product__title"),
        Case(
            When(Q(sku__contains="."), then=_SplitPart("sku", Value("."), Value(1))),
            default=_SplitPart("name", Value(" "), Value(1)),
            output_field=CharField(),
        ),
        output_field=CharField(),
    )


def _tr_ci_variants(term):
    """Cased variants of a search term for case-insensitive matching that
    also covers Turkish dotted/dotless I (İ/ı), which ILIKE does NOT fold
    reliably across DB locales. Returns a de-duplicated list to OR
    together with __icontains. Standard letters (ü/Ü, ş/Ş, ç/Ç, ö/Ö, ğ/Ğ)
    already fold under Unicode ILIKE — only the I-family needs help."""
    term = (term or "").strip()
    if not term:
        return []

    def tr_upper(s):
        # Turkish upper: i→İ, ı→I (do these first, then upper the rest).
        return s.replace("ı", "I").replace("i", "İ").upper()

    def tr_lower(s):
        # Turkish lower: I→ı, İ→i (do these first, then lower the rest).
        return s.replace("I", "ı").replace("İ", "i").lower()

    seen, out = set(), []
    for v in (term, term.lower(), term.upper(), tr_lower(term), tr_upper(term)):
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


@method_decorator(login_required, name='dispatch')
class WarehouseDetail(View):
    template_name = "operating/warehouse_detail.html"

    def get(self, request, pk):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        search = (request.GET.get('search') or '').strip()
        sort = (request.GET.get('sort') or 'name_asc').strip()
        page_num = request.GET.get('page', '1')

        from django.db.models import Q

        base_expr = _warehouse_base_expr()

        base_qs = WarehouseProduct.objects.filter(warehouse=warehouse)
        if search:
            # Case-insensitive across name / SKU / product-barcode AND each
            # roll's (top's) own barcode. We OR a few Turkish-aware cased
            # variants of the term so dotted/dotless İ/ı match no matter how
            # the data was typed. Rolls are matched via a SUBQUERY (not a
            # join) so the grouped-view Sum() aggregates aren't multiplied.
            from functools import reduce
            import operator
            variants = _tr_ci_variants(search)

            def _field_q(field):
                return reduce(operator.or_,
                              (Q(**{f"{field}__icontains": v}) for v in variants))

            roll_match = (WarehouseProductRoll.objects
                          .filter(product__warehouse=warehouse)
                          .filter(_field_q("barcode"))
                          .values('product_id'))
            base_qs = base_qs.filter(
                _field_q("name") | _field_q("sku") | _field_q("barcode")
                | Q(id__in=roll_match)
            )

        # Two list modes: 'variants' (FLAT — every variant/roll-carrier on
        # its own row, the DEFAULT) and 'grouped' (main products, expandable).
        view_mode = (request.GET.get('view') or 'variants').strip()
        if view_mode not in ('grouped', 'variants'):
            view_mode = 'variants'

        groups_render = []
        variants_render = []

        if view_mode == 'variants':
            # ── FLAT: list EVERY variant (WarehouseProduct) individually. ──
            flat = (base_qs
                    .select_related('catalog_variant__product')
                    .annotate(base=base_expr, roll_count=Count('rolls'),
                              line_usd=F('quantity') * F('cost_usd'),
                              line_try=F('quantity') * F('cost_try'),
                              reserved=reserved_meters_subquery()))
            # A → Z by product NAME, top quantity, unit price, or recent.
            _flat_sort = {
                "name_asc": "name", "name_desc": "-name",
                "qty_desc": "-quantity", "qty_asc": "quantity",
                "recent": "-updated_at",
            }
            # Price sorts use the normalized USD unit cost with NULLs last
            # so uncosted products never top the "most expensive" list.
            if sort == "price_desc":
                flat = flat.order_by(F("cost_usd").desc(nulls_last=True), "name", "id")
            elif sort == "price_asc":
                flat = flat.order_by(F("cost_usd").asc(nulls_last=True), "name", "id")
            else:
                flat = flat.order_by(_flat_sort.get(sort, "name"), "sku", "id")
            paginator = Paginator(flat, PAGE_SIZE)
            try:
                page = paginator.page(int(page_num))
            except (PageNotAnInteger, ValueError):
                page = paginator.page(1)
            except EmptyPage:
                page = paginator.page(paginator.num_pages or 1)
            variants_render = list(page.object_list)
        else:
            # ── Group in SQL by main product (base name). One row per group →
            #    cheap to paginate even with thousands of variants. Each group
            #    expands to its variants (stock / SKU / rolls=tops/coupons). ──
            from django.db.models import Max, Avg
            grouped = (base_qs.annotate(base=base_expr).values("base").annotate(
                variant_count=Count("id"),
                linked=Count("catalog_variant"),
                total_qty=Coalesce(Sum("quantity"), Decimal("0"),
                                   output_field=DecimalField(max_digits=18, decimal_places=2)),
                total_usd=Coalesce(Sum(F("quantity") * F("cost_usd")), Decimal("0"),
                                   output_field=DecimalField(max_digits=20, decimal_places=4)),
                total_try=Coalesce(Sum(F("quantity") * F("cost_try")), Decimal("0"),
                                   output_field=DecimalField(max_digits=20, decimal_places=4)),
                updated=Max("updated_at"),
                avg_cost=Avg("cost_usd"),
                avg_cost_try=Avg("cost_try"),
            ))
            _sort_map = {
                "name_asc": "base", "name_desc": "-base",
                "qty_desc": "-total_qty", "qty_asc": "total_qty",
                "recent": "-updated",
            }
            # Price sorts by the group's average USD unit cost, NULLs last.
            if sort == "price_desc":
                grouped = grouped.order_by(F("avg_cost").desc(nulls_last=True), "base")
            elif sort == "price_asc":
                grouped = grouped.order_by(F("avg_cost").asc(nulls_last=True), "base")
            else:
                grouped = grouped.order_by(_sort_map.get(sort, "base"), "base")

            # Paginate the GROUPS (main products), not the variant rows.
            paginator = Paginator(grouped, PAGE_SIZE)
            try:
                page = paginator.page(int(page_num))
            except (PageNotAnInteger, ValueError):
                page = paginator.page(1)
            except EmptyPage:
                page = paginator.page(paginator.num_pages or 1)

            # Total tops (rolls) per group on THIS page — computed in a
            # SEPARATE query (Count only) so it doesn't multiply the Sum()s
            # above via the rolls join.
            _bases = [g["base"] for g in page.object_list]
            _roll_counts = {}
            _reserved_by_base = {}
            if _bases:
                _rc = (base_qs.annotate(base=base_expr)
                       .filter(base__in=_bases).values("base")
                       .annotate(rc=Count("rolls")))
                _roll_counts = {r["base"]: r["rc"] for r in _rc}

                # Active reserved metres per group — computed separately
                # (product-id keyed) so it doesn't multiply the Sum()s.
                _prod_rows = (base_qs.annotate(base=base_expr)
                              .filter(base__in=_bases).values("id", "base"))
                _pid_to_base = {row["id"]: row["base"] for row in _prod_rows}
                _res_by_pid = reserved_meters_for_products(list(_pid_to_base.keys()))
                for _pid, _m in _res_by_pid.items():
                    _b = _pid_to_base.get(_pid)
                    _reserved_by_base[_b] = _reserved_by_base.get(_b, Decimal("0")) + _m

            # Only the group HEADERS render up-front — variants load lazily on
            # expand (a single base can hold thousands of variants, so rendering
            # them all would freeze the page).
            groups_render = [{
                "base": g["base"] or "",
                "title": g["base"] or "—",
                "variant_count": g["variant_count"],
                "roll_total": _roll_counts.get(g["base"], 0),
                "total_qty": g["total_qty"],
                "reserved_total": _reserved_by_base.get(g["base"], Decimal("0")),
                "total_usd": g["total_usd"],
                "total_try": g["total_try"],
                "avg_cost_usd": g["avg_cost"],
                "avg_cost_try": g["avg_cost_try"],
                "is_catalog": (g.get("linked") or 0) > 0,
            } for g in page.object_list]

        is_htmx = bool(request.headers.get('HX-Request'))

        ctx = {
            'warehouse': warehouse,
            'groups': groups_render,      # main-product groups on this page
            'variants': variants_render,  # flat variant rows on this page (view=variants)
            'view_mode': view_mode,
            'page': page,
            'paginator': paginator,
            'search': search,
            'sort': sort,
            'sort_options': [(k, v[1]) for k, v in SORT_OPTIONS.items()],
            'filtered_count': paginator.count,
        }

        # HTMX partial refresh for search/sort/page — re-renders ONLY the
        # rows + pagination, so skip every header stat (net-worth rollups,
        # group/roll counts, suppliers, dup scan, recent movements): each
        # is its own round-trip to the DB and none of it appears in the
        # partial. This is what makes search-as-you-type feel instant.
        if is_htmx:
            if search:
                # The pagination footer shows "filtered from N total".
                ctx['product_count'] = WarehouseProduct.objects.filter(warehouse=warehouse).count()
            return render(request, "operating/partials/warehouse_product_table_body.html", ctx)

        # Aggregate totals across the whole warehouse (independent of
        # filter) — full page load only. Net-worth mirrors the per-product
        # purchase_price fallback, computed as ONE SQL aggregate.
        all_products = WarehouseProduct.objects.filter(warehouse=warehouse)
        counts = all_products.aggregate(
            n=Count('id'),
            qty=Coalesce(Sum('quantity'), Decimal('0'), output_field=DecimalField(max_digits=18, decimal_places=2)),
        )
        # Header counts: main products (packages), variants, and rolls (tops).
        group_count = (all_products.annotate(base=base_expr)
                       .values('base').distinct().count())
        roll_count = WarehouseProductRoll.objects.filter(product__warehouse=warehouse).count()
        total_value_usd, total_value_try = warehouse.total_values()

        # "Son Hareketler" preview — the latest activity in THIS warehouse
        # (all products), linking to the full filterable ledger.
        recent_movements = _decorate_movements(list(
            StockMovement.objects
            .filter(product__warehouse=warehouse)
            .select_related("product", "created_by")
            .order_by("-created_at")[:6]
        ))

        ctx.update({
            'recent_movements': recent_movements,
            'total_value_usd': total_value_usd,
            'total_value_try': total_value_try,
            'product_count': counts['n'],     # variants
            'variant_count': counts['n'],     # alias — variants
            'group_count': group_count,       # main products (packages)
            'roll_count': roll_count,         # rolls (tops)
            'is_admin': _is_admin(request.user),
            'suppliers': _supplier_choices(),   # manual-add dropdown
            'dup_count': _warehouse_dup_sku_count(warehouse),  # duplicate-SKU variant groups
            'total_quantity': counts['qty'],
        })
        return render(request, self.template_name, ctx)


@login_required
def warehouse_group_variants(request, pk):
    """Lazy-load the variant rows for ONE main-product group (base name).

    Returned as raw <tr> rows injected into the group's <tbody> when the
    user expands it — keeps the initial page light even when a base holds
    thousands of variants.
    """
    from django.db.models import Count, F
    warehouse = get_object_or_404(Warehouse, pk=pk)
    base = request.GET.get("base", "")

    base_expr = _warehouse_base_expr()
    CAP = 400
    qs = (WarehouseProduct.objects.filter(warehouse=warehouse)
          .select_related("catalog_variant__product")
          .annotate(base=base_expr, roll_count=Count("rolls"),
                    line_usd=F("quantity") * F("cost_usd"),
                    line_try=F("quantity") * F("cost_try"),
                    reserved=reserved_meters_subquery())
          .filter(base=base).order_by("name", "id"))
    total = qs.count()
    variants = list(qs[:CAP])
    return render(request, "operating/partials/warehouse_group_variant_rows.html", {
        "warehouse": warehouse,
        "variants": variants,
        "truncated": total > CAP,
        "shown": len(variants),
        "total": total,
        "base": base,
    })


@login_required
def warehouse_barcode_lookup(request, pk):
    """Barcode-only LOOKUP (separate from scanning to ADD). Given a scanned
    barcode, says whether that product is already in this warehouse — match a
    roll's (top's) barcode first, else the product barcode / SKU — and returns
    its info so staff can confirm "did I already add this?" without searching."""
    from django.db.models import Q
    warehouse = get_object_or_404(Warehouse, pk=pk)
    code = (request.GET.get("barcode") or "").strip()
    if not code:
        return JsonResponse({"found": False, "error": "No barcode"}, status=400)

    matched = None
    roll = (WarehouseProductRoll.objects
            .select_related("product", "product__catalog_variant__product")
            .filter(product__warehouse=warehouse, barcode__iexact=code)
            .first())
    if roll:
        product = roll.product
        matched = "roll"
    else:
        product = (WarehouseProduct.objects
                   .select_related("catalog_variant__product")
                   .filter(warehouse=warehouse)
                   .filter(Q(barcode__iexact=code) | Q(sku__iexact=code))
                   .first())
        matched = ("sku" if (product and (product.sku or "").lower() == code.lower())
                   else "product") if product else None

    if not product:
        return JsonResponse({"found": False, "barcode": code})

    main = product.catalog_variant.product if product.catalog_variant_id else None
    return JsonResponse({
        "found": True,
        "matched": matched,
        "barcode": code,
        "product": {
            "id": product.pk,
            "sku": product.sku,
            "name": product.name,
            "quantity": float(product.quantity or 0),
            "rolls_count": product.rolls.count(),
            "detail_url": reverse("operating:warehouse_product_detail",
                                  args=[warehouse.pk, product.pk]),
            "main_product": (main.title if main else None),
            "main_product_url": (reverse("marketing:product_detail", args=[main.pk]) if main else None),
        },
        "roll": ({
            "id": roll.pk,
            "barcode": roll.barcode,
            "meters": float(roll.meters or 0),
            "meters_remaining": (float(roll.meters_remaining) if roll.meters_remaining is not None else None),
            "lot_number": roll.lot_number,
        } if roll else None),
    })


@login_required
def catalog_base_search(request, pk):
    """Autocomplete for the scan screen's "Catalog — main product" field.
    Returns hidden catalog products (featured=False) matching the query so
    staff can attach a scanned variant to an EXISTING main product.

    The catalog Product/ProductVariant are an auto-generated MIRROR of
    warehouse data (see catalog_sync.py) — staff recognize stock by its
    WarehouseProduct name, not the catalog's title (which can drift, e.g.
    after a manual rename on one side only). So search matches either
    side, but the name shown is always sourced from the linked
    WarehouseProduct when one exists."""
    from django.db.models import Q
    from marketing.models import Product
    from .catalog_sync import parse_label_name
    q = (request.GET.get("q") or "").strip()
    results = []
    if q:
        # NOTE: variant COUNT is fetched per-product below rather than via
        # .annotate(Count("variants")) here — annotating a Count alongside
        # a filter that itself joins through "variants" (the warehouse-name
        # match) makes Django fold the WHERE into the same join, so the
        # count would reflect only the MATCHING variant(s), not the
        # product's true total. Capped at 12 results, so N+1 is cheap.
        qs = (Product.objects.filter(featured=False)
              .filter(Q(title__icontains=q) | Q(sku__icontains=q) |
                      Q(variants__warehouse_products__name__icontains=q) |
                      Q(variants__warehouse_products__sku__icontains=q))
              .distinct().order_by("title")[:12])
        for p in qs:
            wp = (WarehouseProduct.objects.filter(catalog_variant__product=p)
                  .order_by("id").first())
            title = p.title
            if wp:
                base = parse_label_name(wp.name)["base_name"]
                title = base or wp.name
            results.append({"id": p.id, "title": title, "sku": p.sku or "",
                            "variants": p.variants.count()})
    return JsonResponse({"results": results})


def _warehouse_variant_label(wp, base_name):
    """The bit of a WarehouseProduct's name that actually distinguishes
    this variant — its full name (e.g. "1026 / V-106 ALTIN GÜMÜŞ") with
    the shared base ("1026") stripped off the front, so staff see
    "V-106 ALTIN GÜMÜŞ" exactly as written in the warehouse, not an
    English-translated catalog attribute value."""
    name = wp.name or ""
    if base_name and name.upper().startswith(base_name.upper()):
        name = name[len(base_name):]
    return name.strip(" /-—–") or (wp.name or "")


@login_required
def catalog_product_variants(request, pk, product_id):
    """List the EXISTING variants of one catalog product, for the "Yeni ürün"
    panel's existing-main-product picker — lets staff see what's already on
    the product BEFORE typing a variant name, instead of guessing.

    Label is sourced from the linked WarehouseProduct's own name rather
    than the catalog's translated attribute value ("Silver") — staff
    recognize stock by what's written in the warehouse, not the English
    catalog mirror.

    Only variants with a REAL WarehouseProduct link are shown. Some
    catalog variants have none — leftover duplicates or test rows from
    earlier syncs that never got (or lost) their warehouse row — and
    showing those here is actively misleading: staff see "2 variants"
    for a product that only has 1 in the actual warehouse."""
    from marketing.models import Product
    # Scoped to hidden/warehouse products only — same as catalog_base_search —
    # so this intake endpoint can't be used to read out a real storefront
    # product's SKU/barcode/quantity by guessing its id.
    product = Product.objects.filter(pk=product_id, featured=False).first()
    if product is None:
        return JsonResponse({"results": []})
    variants = (product.variants
                .filter(warehouse_products__isnull=False)
                .prefetch_related("product_variant_attribute_values__product_variant_attribute",
                                  "warehouse_products")
                .distinct().order_by("variant_sku"))
    results = []
    for v in variants:
        av = (v.product_variant_attribute_values.all()[:1] or [None])[0]
        wp = next(iter(v.warehouse_products.all()), None)
        label = _warehouse_variant_label(wp, product.title) if wp else None
        if not label:
            label = (av.product_variant_attribute_value.replace("_", " ").title() if av else v.variant_sku)
        results.append({
            "id": v.id,
            "label": label,
            "attribute_name": (av.product_variant_attribute.name if av else None),
            "variant_sku": v.variant_sku,
            "variant_barcode": v.variant_barcode,
            "variant_quantity": float(v.variant_quantity or 0),
        })
    return JsonResponse({"results": results})


@login_required
def catalog_variant_match(request, pk, product_id):
    """Classify a variant NAME being typed into the "Yeni ürün" panel as
    EXISTING (adds stock to an already-catalogued variant) or NEW (will
    create one), using the SAME parse/translate logic sync_roll_to_catalog
    uses at save time — so the preview badge and the actual save never
    disagree on what counts as "the same variant".

    Only matches variants that have a real WarehouseProduct link — see
    catalog_product_variants for why. Matching an orphaned catalog
    variant would silently add real intake stock to a row with no
    warehouse row behind it, instead of properly creating/linking one."""
    from .catalog_sync import translate_color, _norm_attr, _norm_value
    from marketing.models import Product, ProductVariant

    name = (request.GET.get("name") or "").strip()
    if not name:
        return JsonResponse({"exists": False, "attribute_name": None, "attribute_value": None})

    eng = translate_color(name)
    attribute_name = "color" if eng else "model"
    attribute_value = eng or name

    # Scoped to hidden/warehouse products only — see catalog_product_variants.
    product = Product.objects.filter(pk=product_id, featured=False).first()
    match = None
    if product is not None:
        match = (ProductVariant.objects
                 .filter(product=product,
                         warehouse_products__isnull=False,
                         product_variant_attribute_values__product_variant_attribute__name=_norm_attr(attribute_name),
                         product_variant_attribute_values__product_variant_attribute_value=_norm_value(attribute_value))
                 .first())

    if match:
        return JsonResponse({
            "exists": True,
            "attribute_name": attribute_name,
            "attribute_value": attribute_value,
            "variant_sku": match.variant_sku,
            "variant_quantity": float(match.variant_quantity or 0),
        })
    return JsonResponse({
        "exists": False,
        "attribute_name": attribute_name,
        "attribute_value": attribute_value,
    })


@login_required
def warehouse_next_sku(request, pk):
    """Preview the next auto product SKU for a supplier prefix so the New-product
    panel can show it (and root the variant SKUs on it) BEFORE saving. Indicative
    only — the final SKU is minted authoritatively on save."""
    prefix = (request.GET.get("prefix") or "").strip().upper()[:6] or "GEN"
    try:
        sku = _product_sku_minter(prefix)()
    except Exception:
        sku = f"{prefix}001"
    return JsonResponse({"prefix": prefix, "sku": sku})


def _warehouse_dup_sku_count(warehouse):
    """How many SKUs in the warehouse are carried by >1 WarehouseProduct
    (i.e. duplicate variant rows that should be merged). Case-insensitive."""
    from django.db.models import Count
    from django.db.models.functions import Lower
    return (WarehouseProduct.objects.filter(warehouse=warehouse)
            .exclude(sku__isnull=True).exclude(sku="")
            .annotate(lsku=Lower("sku")).values("lsku")
            .annotate(n=Count("id")).filter(n__gt=1).count())


@method_decorator(login_required, name='dispatch')
class WarehouseMergeDuplicates(View):
    """Merge every same-SKU duplicate variant in the warehouse into one record.
    GET = dry-run PREVIEW (which products/tops would merge, which survives) so
    the user can approve; POST = actually merge (see _merge_warehouse_dupes_by_sku)."""

    def get(self, request, pk):
        from django.db.models import Count
        from django.db.models.functions import Lower
        warehouse = get_object_or_404(Warehouse, pk=pk)
        dup_skus = (WarehouseProduct.objects.filter(warehouse=warehouse)
                    .exclude(sku__isnull=True).exclude(sku="")
                    .annotate(lsku=Lower("sku")).values("lsku")
                    .annotate(n=Count("id")).filter(n__gt=1)
                    .values_list("lsku", flat=True))
        groups = []
        for lsku in list(dup_skus):
            prods = list(WarehouseProduct.objects
                         .filter(warehouse=warehouse, sku__iexact=lsku)
                         .select_related("catalog_variant__product").order_by("id"))
            if len(prods) <= 1:
                continue
            items, total_tops, total_qty, main = [], 0, Decimal("0"), ""
            for i, p in enumerate(prods):
                tc = p.rolls.count()
                total_tops += tc
                total_qty += (p.quantity or Decimal("0"))
                if not main and p.catalog_variant_id and p.catalog_variant and p.catalog_variant.product_id:
                    main = p.catalog_variant.product.title or ""
                items.append({
                    "id": p.id, "name": p.name, "sku": p.sku,
                    "quantity": float(p.quantity or 0), "tops": tc,
                    "survivor": i == 0,
                })
            groups.append({
                "sku": prods[0].sku, "main_product": main,
                "variant_name": prods[0].name, "count": len(prods),
                "total_tops": total_tops, "total_quantity": float(total_qty),
                "products": items,
            })
        return JsonResponse({"success": True, "groups": groups})

    def post(self, request, pk):
        from django.db.models import Count
        from django.db.models.functions import Lower
        warehouse = get_object_or_404(Warehouse, pk=pk)
        dup_rows = (WarehouseProduct.objects.filter(warehouse=warehouse)
                    .exclude(sku__isnull=True).exclude(sku="")
                    .annotate(lsku=Lower("sku")).values("lsku")
                    .annotate(n=Count("id")).filter(n__gt=1))
        groups = 0
        merged_total = 0
        for row in list(dup_rows):
            _surv, m = _merge_warehouse_dupes_by_sku(warehouse, row["lsku"])
            if m:
                groups += 1
                merged_total += m
        return JsonResponse({"success": True, "groups": groups, "merged": merged_total})


@method_decorator(login_required, name='dispatch')
class WarehouseManualAdd(View):
    """Create one or more brand-new (un-labelled) products straight into the
    warehouse in a SINGLE atomic submit. A delivery often contains several
    unrelated products, not just variants of one — the whole batch is built
    in one panel and posted here as ONE request; if any product in the batch
    is invalid, nothing in the batch is saved.

    Each product = ONE main product (group) → its VARIANTS → each variant's
    TOPS (rolls), every top getting an auto-generated, globally-unique
    barcode whose prefix is the supplier's consonants ("Kızılırmak" →
    KZL000001). Unit-agnostic so it fits fabric (m), bedding (pcs), etc.

    JSON body:
      {
        "supplier_id": 5,                 # optional → barcode prefix
        "barcode_prefix": "KZL",          # optional explicit override
        "unit": "mt",                     # mt | adet | kg | paket | ...
        "products": [
          {
            "main_product": {"mode": "new"|"existing", "id": 12,
                              "name": "GREK", "sku": "GREK"},
            "has_variants": true,
            "variants": [
              {"name": "Beyaz", "sku": "GREK-BEYAZ",
               "tops": [{"qty": 160}, {"qty": 150}]}
            ]
          }
        ]
      }
    """

    def post(self, request, pk):
        import re as _re_sku
        from django.db import transaction, IntegrityError
        from .catalog_sync import (
            translate_color, sync_roll_to_catalog, CatalogSyncConflict,
        )
        from marketing.models import Product as _Prod

        warehouse = get_object_or_404(Warehouse, pk=pk)
        try:
            data = json.loads((request.body or b"").decode("utf-8") or "{}")
        except (ValueError, UnicodeDecodeError):
            return JsonResponse({"success": False, "error": "Geçersiz veri."}, status=400)

        unit = (data.get("unit") or "mt").strip()[:20] or "mt"
        products_in = data.get("products")
        if not isinstance(products_in, list) or not products_in:
            return JsonResponse({"success": False, "error": "En az bir ürün ekleyin."}, status=400)

        # ── Supplier → barcode prefix + purchase (alım) posting below ──
        supplier_name = ""
        supplier_obj = None
        supplier_id = data.get("supplier_id")
        if supplier_id:
            try:
                from crm.models import Supplier
                supplier_obj = Supplier.objects.filter(pk=int(supplier_id)).first()
                if supplier_obj:
                    supplier_name = str(supplier_obj)
            except Exception:
                pass
        prefix = (data.get("barcode_prefix") or "").strip().upper()
        if not prefix:
            prefix = _consonant_prefix(supplier_name) if supplier_name else "GEN"
        prefix = (prefix[:6] or "GEN")

        # ── Pass 1: resolve + validate EVERY product before writing anything,
        # so a mistake on product #3 never leaves #1/#2 half-saved. ──
        resolved = []
        for i, p_in in enumerate(products_in, start=1):
            variants_in = p_in.get("variants") or []
            if not isinstance(variants_in, list) or not variants_in:
                return JsonResponse({"success": False, "error": f"Ürün {i}: en az bir varyant ekleyin."}, status=400)
            mp = p_in.get("main_product") or {}
            mode = (mp.get("mode") or "new").strip()
            main_product = None
            if mode == "existing" and str(mp.get("id") or "").isdigit():
                main_product = _Prod.objects.filter(pk=int(mp["id"]), featured=False).first()
            base_name = (mp.get("name") or "").strip()
            if main_product is not None:
                base_name = main_product.title
            elif not base_name:
                return JsonResponse({"success": False, "error": f"Ürün {i}: ana ürün adı gerekli."}, status=400)

            has_variants = p_in.get("has_variants")
            if has_variants is None:
                has_variants = True
            # Previewed auto SKU the panel showed (so what the user sees is what
            # gets saved). Honoured only if it's a valid prefix+digits auto-code
            # AND still free; otherwise we mint the next one.
            desired_sku = (mp.get("sku") or "").strip().upper()
            if not _re_sku.match(r"^" + _re_sku.escape(prefix) + r"\d{3,}$", desired_sku or ""):
                desired_sku = ""

            resolved.append({
                "main_product": main_product, "base_name": base_name,
                "desired_sku": desired_sku, "has_variants": has_variants,
                "variants_in": variants_in,
            })

        created_list = []
        warnings = []
        purchase_lines = []   # aggregated across the WHOLE batch → one alış faturası
        mint = _barcode_minter(prefix)
        prod_unit = _PRODUCT_UNIT_MAP.get(unit, "units")
        usd_try = _get_usd_try_rate() or Decimal("1")
        user = request.user if request.user.is_authenticated else None

        try:
            with transaction.atomic():
                for item in resolved:
                    main_product = item["main_product"]
                    base_name = item["base_name"]
                    desired_sku = item["desired_sku"]
                    has_variants = item["has_variants"]
                    variants_in = item["variants_in"]

                    created = {"main_product": None, "variants": 0, "tops": 0,
                               "barcodes": [], "variant_skus": []}

                    # NEW main product → AUTO, globally-unique SKU = supplier prefix
                    # + number (e.g. KZL004). Try the previewed code first, then mint;
                    # create with an IntegrityError retry so a concurrent request can't
                    # win the DB-unique Product.sku race. We always pre-create WITH the
                    # sku and pass existing_base_product below, so sync_roll_to_catalog
                    # never re-derives or nulls it.
                    if main_product is None:
                        sku_mint = _product_sku_minter(prefix)
                        for _attempt in range(8):
                            if _attempt == 0 and desired_sku and not _Prod.objects.filter(sku__iexact=desired_sku).exists():
                                candidate = desired_sku
                            else:
                                candidate = sku_mint()
                            try:
                                with transaction.atomic():   # savepoint
                                    main_product = _Prod.objects.create(
                                        title=base_name, sku=candidate, featured=False,
                                        unit_of_measurement=prod_unit, quantity=Decimal("0"),
                                    )
                                break
                            except IntegrityError:
                                main_product = None
                                continue
                        if main_product is None:
                            raise RuntimeError("Benzersiz ürün SKU üretilemedi, tekrar deneyin.")
                    created["main_product"] = {
                        "id": main_product.id, "title": main_product.title,
                        "sku": main_product.sku,
                    }

                    # No-variant ("simple product") mode → collapse to ONE implicit
                    # variant that IS the product (no colour/model, sku = main sku).
                    if not has_variants:
                        src = variants_in[0] if variants_in else {}
                        variants_in = [{
                            "name": "", "sku": main_product.sku or "",
                            "tops": src.get("tops") or [],
                            "price": src.get("price"), "currency": src.get("currency"),
                        }]

                    seen_skus = set()
                    for idx, v in enumerate(variants_in, start=1):
                        v_name = (v.get("name") or "").strip()
                        typed_sku = (v.get("sku") or "").strip()[:20]
                        v_sku = typed_sku
                        tops = v.get("tops") or []
                        if not v_name and not v_sku and not tops:
                            continue
                        if not v_sku:
                            # AUTO variant SKU rooted on the (minted) main product SKU.
                            root = (main_product.sku or base_name or "SKU").strip()
                            suffix = _slug_token(v_name) or str(idx)
                            v_sku = f"{root}-{suffix}"[:20]

                        def _bump(s, n):
                            tail = str(n)
                            return f"{s[:max(1, 20 - len(tail))]}{tail}"

                        base_v = v_sku
                        dup = 1
                        if typed_sku:
                            # Respect a typed SKU; only avoid clashing within THIS product
                            # (a global clash surfaces as a CatalogSyncConflict warning).
                            while v_sku in seen_skus:
                                dup += 1
                                v_sku = _bump(base_v, dup)
                        else:
                            # AUTO SKU: keep it unique product-wide AND globally.
                            while v_sku in seen_skus or _variant_sku_exists(v_sku):
                                dup += 1
                                v_sku = _bump(base_v, dup)
                        seen_skus.add(v_sku)

                        # Colour vs model attribute, derived from the variant name.
                        eng = translate_color(v_name) if v_name else None
                        attr_name = ("color" if eng else ("model" if v_name else None))
                        attr_value = (eng or v_name) or None

                        # Purchase price (alış fiyatı) → unit cost in USD/TRY so the
                        # warehouse value rollup reflects it.
                        price = _safe_decimal(v.get("price"))
                        currency = (v.get("currency") or "USD").strip().upper()
                        if currency not in ("USD", "TRY", "EUR"):
                            currency = "USD"
                        cost_usd = cost_try = None
                        if price is not None and price > 0:
                            if currency == "USD":
                                cost_usd = price
                                cost_try = (price * usd_try) if usd_try else None
                            elif currency == "TRY":
                                cost_try = price
                                if usd_try and usd_try > 0:
                                    cost_usd = (price / usd_try).quantize(Decimal("0.0001"))
                            elif currency == "EUR":
                                cost_usd = price   # coarse EUR≈USD for the rollup

                        wp_name = (f"{base_name} {v_name}".strip()) or v_sku
                        # Reuse an existing same-SKU product in this warehouse rather
                        # than creating a duplicate row (same SKU = same variant).
                        wp = (WarehouseProduct.objects
                              .filter(warehouse=warehouse, sku__iexact=v_sku).first())
                        if wp is None:
                            wp = WarehouseProduct.objects.create(
                                warehouse=warehouse, name=wp_name, sku=v_sku,
                                quantity=Decimal("0"),
                                purchase_price=(price if (price and price > 0) else None),
                                purchase_currency=currency,
                                cost_usd=cost_usd, cost_try=cost_try,
                            )
                        elif price and price > 0:
                            wp.purchase_price = price
                            wp.purchase_currency = currency
                            wp.cost_usd = cost_usd
                            wp.cost_try = cost_try
                            wp.save(update_fields=["purchase_price", "purchase_currency",
                                                   "cost_usd", "cost_try", "updated_at"])

                        first_barcode = None
                        added_qty = Decimal("0")
                        new_roll_ids = []
                        for t in tops:
                            qty = _safe_decimal(t.get("qty"))
                            if qty is None or qty <= 0:
                                continue
                            code = mint()
                            if first_barcode is None:
                                first_barcode = code
                            roll = WarehouseProductRoll.objects.create(
                                product=wp, meters=qty, meters_remaining=qty,
                                barcode=code,
                                notes=(f"Supplier: {supplier_name}" if supplier_name else None),
                                scanned_by=user,
                            )
                            StockMovement.objects.create(
                                product=wp, roll=roll, movement_type="in",
                                quantity=qty, reason="Manual add",
                                reference=code, created_by=user,
                            )
                            added_qty += qty
                            new_roll_ids.append(roll.pk)
                            created["tops"] += 1
                            created["barcodes"].append(code)

                        # Quantity = sum of remaining metres across ALL rolls (covers
                        # reusing an existing same-SKU product, not just the new tops).
                        total = Decimal("0")
                        for rr in wp.rolls.all():
                            rem = rr.meters_remaining if rr.meters_remaining is not None else (rr.meters or Decimal("0"))
                            total += rem or Decimal("0")
                        wp.quantity = total
                        if first_barcode and not wp.barcode:
                            wp.barcode = first_barcode[:64]
                        wp.save(update_fields=["quantity", "barcode", "updated_at"])

                        cat_variant_obj = None
                        try:
                            _p, cat_variant, _pc, _vc = sync_roll_to_catalog(
                                base_name=base_name,
                                attribute_name=attr_name,
                                attribute_value=attr_value,
                                variant_sku=v_sku, variant_barcode=first_barcode,
                                quantity=total, cost=cost_usd,
                                existing_base_product=main_product,
                            )
                            wp.catalog_variant = cat_variant
                            wp.save(update_fields=["catalog_variant"])
                            cat_variant_obj = cat_variant
                        except CatalogSyncConflict as exc:
                            warnings.append(f"{v_sku}: {exc}")
                        created["variants"] += 1
                        created["variant_skus"].append(v_sku)

                        # Purchase-invoice line — only for stock actually added
                        # in THIS request (added_qty), never the product's full
                        # roll total, so re-adding to an existing SKU doesn't
                        # re-bill previous intakes.
                        if supplier_obj and added_qty > 0:
                            purchase_lines.append({
                                "description": wp_name,
                                "quantity": added_qty,
                                "unit": unit,
                                "unit_price": price if (price and price > 0) else Decimal("0"),
                                "currency": currency,
                                "product": main_product,
                                "variant": cat_variant_obj,
                                "roll_ids": new_roll_ids,
                            })

                    created_list.append(created)
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return JsonResponse({"success": False, "error": f"Kayıt hatası: {exc}"}, status=500)

        # ── Alış faturası (purchase invoice) — the intake above IS a
        # purchase: we now owe the supplier for these goods, across every
        # product in the batch. Posted to the supplier's cari as ONE issued
        # purchase invoice so the alım shows up in the invoice list
        # (type=purchase) and the cari statement links straight back to it.
        # Best-effort: a bookkeeping hiccup must never roll back the physical
        # stock that was just added.
        purchase_info = None
        if supplier_obj and purchase_lines:
            try:
                from current_account.services import create_purchase_invoice_for_intake
                member = getattr(request.user, "member", None)
                inv = create_purchase_invoice_for_intake(
                    supplier_obj, purchase_lines, member=member, user=user,
                )
                purchase_info = {
                    "invoice_id": inv.pk,
                    "number": f"{inv.series}-{inv.number}",
                    "total": float(inv.total or 0),
                    "currency": inv.currency.code,
                }
                # Link each physical top back to the invoice line it was
                # received against — items are created in the same order
                # as purchase_lines, so zipping by line_no is exact.
                inv_items = list(inv.items.order_by("line_no"))
                for line, item in zip(purchase_lines, inv_items):
                    roll_ids = line.get("roll_ids") or []
                    if roll_ids:
                        WarehouseProductRoll.objects.filter(pk__in=roll_ids).update(
                            purchase_invoice_item=item
                        )
                if not any(l["unit_price"] > 0 for l in purchase_lines):
                    warnings.append(
                        "Alış faturası 0 tutarla oluşturuldu — fiyat girilmedi. "
                        "Faturayı cari sayfasından düzenleyebilirsiniz."
                    )
            except Exception as exc:
                warnings.append(f"Stok eklendi ama alış faturası oluşturulamadı: {exc}")

        return JsonResponse({
            "success": True, "created": created_list,
            "warnings": warnings, "prefix": prefix,
            "purchase": purchase_info,
        })


@method_decorator(login_required, name='dispatch')
class WarehouseProductImport(View):
    """Import products from Excel — streaming progress + bulk operations.

    Speed strategy:
      1. Read all rows in one pass (openpyxl read_only)
      2. Pre-fetch existing products for this warehouse, indexed by SKU
      3. Categorize each row: skip / update / create (compare quantity to skip unchanged)
      4. Use bulk_create + bulk_update (chunks of 1000)
      5. Stream progress as newline-delimited JSON

    Expected columns (case-insensitive TR/EN):
      Adı/Name, Kodu/SKU, Miktar/Quantity, Alış Fiyatı/Purchase Price,
      Para Birimi/Currency, Barkod (optional), Model (optional)
    """

    def post(self, request, pk):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        upload = request.FILES.get('file')
        if not upload:
            return JsonResponse({'success': False, 'error': 'No file uploaded'}, status=400)

        try:
            from openpyxl import load_workbook
        except ImportError:
            return JsonResponse({'success': False, 'error': 'openpyxl not installed'}, status=500)

        try:
            wb = load_workbook(upload, data_only=True, read_only=True)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Invalid xlsx: {e}'}, status=400)

        ws = wb.active

        # Find header row
        header_row = None
        rows_iter = ws.iter_rows(values_only=True)
        for row in rows_iter:
            if row and any(c is not None and str(c).strip() != '' for c in row):
                header_row = [str(c).strip().lower() if c is not None else '' for c in row]
                break

        if not header_row:
            return JsonResponse({'success': False, 'error': 'Empty sheet'}, status=400)

        def find_col(*candidates):
            for cand in candidates:
                cand = cand.lower()
                for i, h in enumerate(header_row):
                    if cand in h:
                        return i
            return None

        col_name = find_col('adı', 'adi', 'name', 'ürün adı', 'urun adi')
        col_sku = find_col('kodu', 'kod', 'sku', 'code')
        col_qty = find_col('miktar', 'quantity', 'qty', 'stok')
        col_price = find_col('alış fiyatı', 'alis fiyati', 'purchase price', 'a. fiyatı', 'a.fiyat')
        col_currency = find_col('para birimi', 'currency', 'a.f.b', 'döviz')
        col_barcode = find_col('barkod', 'barcode')
        col_model = find_col('model')

        if col_name is None or col_sku is None or col_qty is None or col_price is None:
            return JsonResponse({
                'success': False,
                'error': 'Required columns not found. Need: Name, Code/SKU, Quantity, Purchase Price.',
                'detected_headers': header_row,
            }, status=400)

        # Buffer all data rows into memory (fast — openpyxl read_only mode)
        data_rows = []
        for row in rows_iter:
            if row and any(c is not None and str(c).strip() != '' for c in row):
                data_rows.append(row)

        wb.close()
        total_rows = len(data_rows)

        if total_rows == 0:
            return JsonResponse({'success': True, 'created': 0, 'updated': 0, 'skipped': 0, 'unchanged': 0, 'total': 0})

        usd_to_try = _get_usd_try_rate()
        try_to_usd = (Decimal('1') / usd_to_try) if usd_to_try and usd_to_try != 0 else Decimal('0')

        # Pre-fetch existing rows once — by SKU
        existing_by_sku = {
            p.sku: p
            for p in WarehouseProduct.objects.filter(warehouse=warehouse).only(
                'id', 'sku', 'name', 'barcode', 'model', 'quantity',
                'purchase_price', 'purchase_currency', 'cost_usd', 'cost_try',
            )
            if p.sku
        }

        def stream():
            to_create = []
            to_update = []
            update_fields = ['name', 'barcode', 'model', 'quantity', 'purchase_price',
                             'purchase_currency', 'cost_usd', 'cost_try']

            counters = {'created': 0, 'updated': 0, 'unchanged': 0, 'skipped': 0}
            errors = []

            yield (json.dumps({'phase': 'start', 'total': total_rows}) + '\n').encode('utf-8')

            for idx, row in enumerate(data_rows, start=1):
                try:
                    name = row[col_name] if col_name < len(row) else None
                    sku = row[col_sku] if col_sku < len(row) else None

                    if not name or str(name).strip() == '':
                        counters['skipped'] += 1
                        continue

                    name = str(name).strip()
                    sku_str = str(sku).strip() if sku is not None else None

                    qty = _safe_decimal(row[col_qty] if col_qty < len(row) else None, Decimal('0'))
                    price = _safe_decimal(row[col_price] if col_price < len(row) else None)
                    currency_raw = row[col_currency] if (col_currency is not None and col_currency < len(row)) else None
                    barcode = row[col_barcode] if (col_barcode is not None and col_barcode < len(row)) else None
                    model = row[col_model] if (col_model is not None and col_model < len(row)) else None

                    barcode = str(barcode).strip() if barcode is not None else None
                    model = str(model).strip() if model is not None else None

                    # Currency
                    currency = 'USD'
                    if currency_raw is not None:
                        cu = str(currency_raw).strip().upper()
                        if 'TRY' in cu or 'TL' in cu or '₺' in cu:
                            currency = 'TRY'
                        elif 'EUR' in cu or '€' in cu:
                            currency = 'EUR'

                    # Costs
                    cost_usd = None
                    cost_try = None
                    if price is not None:
                        if currency == 'USD':
                            cost_usd = price
                            cost_try = price * usd_to_try
                        elif currency == 'TRY':
                            cost_try = price
                            cost_usd = price * try_to_usd
                        else:
                            cost_usd = price
                            cost_try = price * usd_to_try

                    existing = existing_by_sku.get(sku_str) if sku_str else None

                    if existing:
                        # Compare quantity (and price) — skip if unchanged
                        same_qty = (existing.quantity or 0) == (qty or 0)
                        same_price = (existing.purchase_price or 0) == (price or 0)
                        same_curr = (existing.purchase_currency or '') == currency
                        same_name = (existing.name or '') == name

                        if same_qty and same_price and same_curr and same_name:
                            counters['unchanged'] += 1
                        else:
                            existing.name = name
                            existing.barcode = barcode
                            existing.model = model
                            existing.quantity = qty
                            existing.purchase_price = price
                            existing.purchase_currency = currency
                            existing.cost_usd = cost_usd
                            existing.cost_try = cost_try
                            to_update.append(existing)
                            counters['updated'] += 1
                    else:
                        to_create.append(WarehouseProduct(
                            warehouse=warehouse,
                            name=name,
                            sku=sku_str,
                            barcode=barcode,
                            model=model,
                            quantity=qty,
                            purchase_price=price,
                            purchase_currency=currency,
                            cost_usd=cost_usd,
                            cost_try=cost_try,
                        ))
                        counters['created'] += 1

                except Exception as e:
                    errors.append(f"Row {idx}: {e}")
                    counters['skipped'] += 1

                # Stream progress every 200 rows
                if idx % 200 == 0 or idx == total_rows:
                    yield (json.dumps({
                        'phase': 'processing',
                        'current': idx,
                        'total': total_rows,
                        **counters,
                    }) + '\n').encode('utf-8')

            # Bulk write to DB
            yield (json.dumps({'phase': 'saving', 'message': 'Writing to database...'}) + '\n').encode('utf-8')

            if to_create:
                WarehouseProduct.objects.bulk_create(to_create, batch_size=1000, ignore_conflicts=False)

            if to_update:
                WarehouseProduct.objects.bulk_update(to_update, update_fields, batch_size=1000)

            # Link every imported row to the marketing catalog (match by
            # SKU, create hidden products/variants for the rest) so the
            # invariant "every warehouse product has a catalog link"
            # survives imports. Never let linking break the import itself.
            linked_info = {}
            touched_skus = [p.sku for p in to_create if p.sku] + \
                           [p.sku for p in to_update if p.sku]
            if touched_skus:
                yield (json.dumps({'phase': 'linking',
                                   'message': 'Linking to catalog...'}) + '\n').encode('utf-8')
                try:
                    from .catalog_reconcile import reconcile_all_warehouse_links
                    rs = reconcile_all_warehouse_links(apply=True, skus=touched_skus)
                    linked_info = {
                        'catalog_linked': rs['linked_wps'],
                        'catalog_created': rs['products_created'],
                        'catalog_conflicts': len(rs['conflicts']),
                    }
                except Exception as _exc:
                    linked_info = {'catalog_link_error': str(_exc)}

            yield (json.dumps({
                'phase': 'done',
                'success': True,
                'total': total_rows,
                **counters,
                **linked_info,
                'usd_to_try_rate': str(usd_to_try),
                'errors': errors[:10],
            }) + '\n').encode('utf-8')

        response = StreamingHttpResponse(stream(), content_type='application/x-ndjson')
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        return response


@method_decorator(login_required, name='dispatch')
class WarehouseDelete(View):
    def post(self, request, pk):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        if not _is_admin(request.user):
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Sadece admin depo silebilir."}, status=403)
            messages.error(request, "Sadece admin depo silebilir.")
            return redirect("operating:warehouse_detail", warehouse.pk)
        name = warehouse.name
        warehouse.delete()
        messages.success(request, f"Warehouse '{name}' deleted")
        return redirect(reverse('operating:warehouse_list'))


# ────────────────────────────────────────────────────────────────────
# Roll scanning — camera-driven OCR of fabric roll labels
# ────────────────────────────────────────────────────────────────────

_LABEL_PROMPT = (
    "This is a photo of a Turkish fabric roll label. Extract these "
    "fields and return ONLY a single JSON object — no commentary, "
    "no markdown fences:\n\n"
    '{"sku": "...", "name": "...", "meters": 48.5, "barcode": "...", "color": "...", "coupon": "..."}\n\n'
    "Field rules:\n"
    "- sku: the product CODE. Usually an uppercase prefix + zeros + digits "
    "(e.g. 'İPK0000174', 'K24614', 'RK48060RW9', 'K48083İ.G93'). It is "
    "printed on the label, NOT the long number under the barcode.\n"
    "- name: the product/fabric name as printed — typically a model "
    "name + colour (e.g. 'S-LINE 1106 EKRU', 'GREK TÜL'). Do NOT include "
    "the brand name (KARVEN, etc.).\n"
    "- meters: roll length as a decimal number. Turkish labels write "
    "'25,00 Metre' or '48,5 m' — interpret comma as decimal point, so "
    "'25,00' → 25.0, '48,5' → 48.5. Do not include the unit.\n"
    "- color: the colour/variant text if printed on its own line "
    "(e.g. 'AÇIK KREM', 'EKRU', 'GÜMÜŞ'); else null.\n"
    "- coupon: the value after 'Kupon No' / 'Kupon' if present; else null.\n\n"
    "CRITICAL — Turkish dotted 'İ' vs plain 'I': these are TURKISH labels and "
    "the SKU very often contains the dotted capital 'İ' (a capital I WITH a "
    "dot above it). Real examples: 'K24620İ.G52', 'K25318İ.G38', 'K24828İT', "
    "'K48083İ'. Inspect EVERY I-shaped letter in the SKU: if it has a dot "
    "above it you MUST output 'İ' (Unicode U+0130), NEVER the plain dotless "
    "'I'. Output plain 'I' only when there is clearly no dot. Zoom in mentally "
    "on the top of each vertical stroke. Likewise keep Ş Ğ Ü Ö Ç and the "
    "dotless lowercase 'ı' exactly as printed. Do NOT transliterate or drop "
    "any diacritic.\n"
    "If a field is illegible or absent in THIS image, set it to null. "
    "Do not invent values."
)


def _ocr_label_openai(image_file):
    """OpenAI Vision OCR — fast path. Requires OPENAI_API_KEY. Model via
    the OPENAI_MODEL env var (default 'gpt-4o-mini' — fast + accurate on
    fabric labels). Uses detail='high' so the tiny dot on a Turkish 'İ'
    is actually visible to the model. Returns (raw_text, parsed_dict)."""
    import base64
    import json as _json
    import re as _re

    try:
        from openai import OpenAI
    except ImportError:
        return "", {"error": "OpenAI SDK not installed (pip install openai)"}

    api_key = _env("OPENAI_API_KEY")
    if not api_key:
        return "", {"error": "OPENAI_API_KEY env var not set"}

    try:
        image_file.seek(0)
    except Exception:
        pass
    raw_bytes = image_file.read()
    if not raw_bytes:
        return "", {"error": "Empty image"}

    media_type = "image/jpeg"
    if raw_bytes[:8] == b"\x89PNG\r\n\x1a\n":
        media_type = "image/png"
    elif raw_bytes[:6] in (b"GIF87a", b"GIF89a"):
        media_type = "image/gif"
    elif raw_bytes[:4] == b"RIFF" and raw_bytes[8:12] == b"WEBP":
        media_type = "image/webp"
    b64 = base64.b64encode(raw_bytes).decode("ascii")
    data_uri = f"data:{media_type};base64,{b64}"

    # The fastest OpenAI vision model. Override with OPENAI_MODEL
    # (e.g. 'gpt-4.1-mini' for a touch more accuracy).
    model_name = _env("OPENAI_MODEL") or "gpt-4o-mini"

    try:
        client = OpenAI(api_key=api_key, timeout=30)
        resp = client.chat.completions.create(
            model=model_name,
            temperature=0,
            max_tokens=400,
            response_format={"type": "json_object"},
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": _LABEL_PROMPT},
                    {"type": "image_url",
                     "image_url": {"url": data_uri, "detail": "high"}},
                ],
            }],
        )
    except Exception as e:
        return "", {"error": f"OpenAI call failed: {e}"}

    try:
        text_out = (resp.choices[0].message.content or "").strip()
    except Exception:
        text_out = ""

    start = text_out.find("{")
    end = text_out.rfind("}")
    if start == -1 or end == -1:
        return text_out, {"error": "OpenAI did not return JSON",
                          "sku": None, "name": None, "meters": None}
    try:
        data = _json.loads(text_out[start:end + 1])
    except Exception:
        return text_out, {"error": "OpenAI returned malformed JSON",
                          "sku": None, "name": None, "meters": None}

    sku = (data.get("sku") or None)
    if sku:
        sku = _tr_upper(str(sku).strip()) or None
    name = (data.get("name") or None)
    if name:
        name = str(name).strip() or None
    meters = data.get("meters")
    try:
        meters = float(meters) if meters is not None else None
    except (TypeError, ValueError):
        meters = None
    barcode = (data.get("barcode") or None)
    if barcode:
        cleaned = _re.sub(r"\D", "", str(barcode))
        barcode = cleaned or None
    color = (data.get("color") or None)
    if color:
        color = str(color).strip() or None
    coupon = (data.get("coupon") or None)
    if coupon:
        coupon = str(coupon).strip() or None

    return text_out, {"sku": sku, "name": name, "meters": meters,
                      "barcode": barcode, "color": color, "coupon": coupon}


def _ocr_label_claude(image_file):
    """Send the label image to Claude Vision and ask it to extract
    {sku, name, meters} as structured JSON. Far more reliable than
    Tesseract on real-world fabric labels (angles, lighting,
    reflective surfaces, mixed fonts).

    Returns (raw_text, parsed_dict). On error the parsed dict has an
    "error" key.

    Requires ANTHROPIC_API_KEY env var.
    """
    import base64
    import json as _json

    try:
        import anthropic
    except ImportError:
        return "", {"error": "Anthropic SDK not installed (pip install anthropic)"}

    api_key = _env("ANTHROPIC_API_KEY")
    if not api_key:
        return "", {"error": "ANTHROPIC_API_KEY env var not set"}

    # Read + b64 the image.
    try:
        image_file.seek(0)
    except Exception:
        pass
    raw_bytes = image_file.read()
    if not raw_bytes:
        return "", {"error": "Empty image"}

    # Detect media type from the first few bytes.
    media_type = "image/jpeg"
    if raw_bytes[:3] == b"\x89PN" or raw_bytes[:8] == b"\x89PNG\r\n\x1a\n":
        media_type = "image/png"
    elif raw_bytes[:6] in (b"GIF87a", b"GIF89a"):
        media_type = "image/gif"
    elif raw_bytes[:4] == b"RIFF" and raw_bytes[8:12] == b"WEBP":
        media_type = "image/webp"

    b64 = base64.b64encode(raw_bytes).decode("ascii")

    prompt = (
        "This is a photo of a Turkish fabric roll label. Extract these "
        "fields and return ONLY a single JSON object — no commentary, "
        "no markdown fences:\n\n"
        '{"sku": "...", "name": "...", "meters": 48.5, "barcode": "...", "color": "...", "coupon": "..."}\n\n'
        "Field rules:\n"
        "- sku: the product CODE. Usually an uppercase prefix + zeros + digits "
        "(e.g. 'İPK0000174', 'K24614', 'RK48060RW9', 'K48083İ.G93'). It is "
        "printed on the label, NOT the long number under the barcode. If the "
        "SKU contains a Turkish dotted İ, preserve it exactly (İ, not I).\n"
        "- name: the product/fabric name as printed — typically a model "
        "name + colour (e.g. 'S-LINE 1106 EKRU', 'GREK TÜL'). Do NOT "
        "include the brand name (KARVEN, etc.) here.\n"
        "- meters: roll length as a decimal number. Turkish labels write "
        "'25,00 Metre' or '48,5 m' — interpret comma as decimal point, so "
        "'25,00' → 25.0, '48,5' → 48.5. Do not include the unit.\n"
        "- color: the colour/variant text if printed on its own line "
        "(e.g. 'AÇIK KREM', 'EKRU', 'GÜMÜŞ'); else null.\n"
        "- coupon: the value after 'Kupon No' / 'Kupon' if present; else null.\n\n"
        "If a field is illegible or absent, set it to null. "
        "Do not invent values."
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=400,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": b64,
                            },
                        },
                        {"type": "text", "text": prompt},
                    ],
                }
            ],
        )
    except anthropic.APIError as e:
        return "", {"error": f"Claude API error: {e}"}
    except Exception as e:
        return "", {"error": f"Claude call failed: {e}"}

    # Pull the text content out of the response.
    text_out = ""
    for block in resp.content:
        if getattr(block, "type", None) == "text":
            text_out += block.text

    # Parse the JSON — be tolerant of code fences or stray text.
    json_str = text_out.strip()
    if "```" in json_str:
        # extract the first ``` block
        parts = json_str.split("```")
        for p in parts:
            if "{" in p:
                # strip optional language tag
                if p.startswith("json"):
                    p = p[4:]
                json_str = p
                break
    start = json_str.find("{")
    end = json_str.rfind("}")
    if start == -1 or end == -1:
        return text_out, {"error": "Claude did not return JSON", "sku": None, "name": None, "meters": None}
    try:
        data = _json.loads(json_str[start:end + 1])
    except Exception:
        return text_out, {"error": "Claude returned malformed JSON", "sku": None, "name": None, "meters": None}

    sku = (data.get("sku") or None)
    if sku:
        sku = _tr_upper(str(sku).strip()) or None
    name = (data.get("name") or None)
    if name:
        name = str(name).strip() or None
    meters = data.get("meters")
    try:
        meters = float(meters) if meters is not None else None
    except (TypeError, ValueError):
        meters = None
    barcode = (data.get("barcode") or None)
    if barcode:
        # Digits only — strip whitespace / punctuation the LLM may add.
        import re as _re
        cleaned = _re.sub(r"\D", "", str(barcode))
        barcode = cleaned or None
    color = (data.get("color") or None)
    if color:
        color = str(color).strip() or None
    coupon = (data.get("coupon") or None)
    if coupon:
        coupon = str(coupon).strip() or None

    return text_out, {"sku": sku, "name": name, "meters": meters,
                      "barcode": barcode, "color": color, "coupon": coupon}


def _ocr_label_gemini(image_file):
    """Send the label image to Gemini Vision and ask it to extract
    {sku, name, meters} as structured JSON. Cheap + fast option.

    Requires GEMINI_API_KEY (or GOOGLE_API_KEY) env var.
    """
    import json as _json

    try:
        from google import genai
        from google.genai import types as genai_types
    except ImportError:
        return "", {"error": "google-genai SDK not installed (pip install google-genai)"}

    api_key = _env("GEMINI_API_KEY") or _env("GOOGLE_API_KEY")
    if not api_key:
        return "", {"error": "GEMINI_API_KEY env var not set"}

    try:
        image_file.seek(0)
    except Exception:
        pass
    raw_bytes = image_file.read()
    if not raw_bytes:
        return "", {"error": "Empty image"}

    media_type = "image/jpeg"
    if raw_bytes[:8] == b"\x89PNG\r\n\x1a\n":
        media_type = "image/png"
    elif raw_bytes[:6] in (b"GIF87a", b"GIF89a"):
        media_type = "image/gif"
    elif raw_bytes[:4] == b"RIFF" and raw_bytes[8:12] == b"WEBP":
        media_type = "image/webp"

    prompt = (
        "This is a photo of a Turkish fabric roll label. Read EXACTLY "
        "what is printed in THIS image — do not invent, do not "
        "remember anything from previous images. Return ONLY a single "
        "JSON object — no commentary, no markdown fences:\n\n"
        '{"sku": "...", "name": "...", "meters": 48.5, "barcode": "...", "color": "...", "coupon": "..."}\n\n'
        "LABEL STRUCTURE (top → bottom):\n"
        "  1. BRAND   (e.g. KARVEN, DEMFIRAT, BELINO)  ← IGNORE, never extract\n"
        "  2. SKU     (short alphanumeric code)        → goes in 'sku'\n"
        "  3. MODEL   (descriptive product/model name) → goes in 'name'\n"
        "  4. METERS  (length in metres)               → goes in 'meters'\n"
        "  5. BARCODE (the code under the barcode bars) → goes in 'barcode'\n\n"
        "Field rules:\n"
        "- sku: the short product CODE on the second line (e.g. "
        "'İPK0000174', 'K24614', 'RK48060RW9'). Preserve Turkish "
        "letters (İ, Ş, Ğ, Ç, Ö, Ü). NOT the long barcode, NOT the "
        "brand.\n"
        "- name: the MODEL NAME printed DIRECTLY UNDER the sku. "
        "Examples: 'S-LINE 1106 EKRU', 'KADİFE KREM', 'PANAMA 4500 "
        "GRİ', 'YAĞMUR DAMLASI ANTRASİT', 'POLO-203 / KIRIK BEYAZ'. "
        "Read VERBATIM, exactly as printed. INCLUDE special "
        "characters (- / + & . ' Turkish letters İŞĞÇÖÜ) and any "
        "numbers that are part of the name. DO NOT return the brand "
        "(KARVEN/DEMFIRAT/BELINO etc.) here — the brand is the FIRST "
        "line and must be skipped. Only return null if there really "
        "is no model line between the sku and the metres/barcode.\n"
        "- meters: roll length as a decimal number. Turkish labels "
        "write '25,00 Metre' or '48,5 m' — interpret comma as decimal "
        "point ('25,00' → 25.0). No unit.\n"
        "- barcode: the code printed directly under the barcode strip. "
        "ALPHANUMERIC — may have a letter prefix like 'KL055555' or be "
        "all digits like '017400000228'. Copy EXACTLY what is printed, "
        "including any letters at the start. Typically 8 to 16 chars.\n"
        "- color: a colour/variant printed on its own line if present "
        "(e.g. 'AÇIK KREM', 'EKRU', 'GÜMÜŞ'); else null.\n"
        "- coupon: the value after 'Kupon No' / 'Kupon' if present; else null.\n\n"
        "If a field is illegible or absent in THIS image, set it to "
        "null. Do not invent values."
    )

    # Allow overriding the model from env. Flash is the sweet spot.
    model_name = _env("GEMINI_MODEL") or "gemini-2.5-flash"

    # Build the config. Disable thinking mode for Gemini 2.5 — without
    # this the model burns most of max_output_tokens on internal
    # reasoning and the real JSON comes back truncated (we saw
    # responses like '{"sku' or 'Here is the JSON requested:\n```').
    cfg_kwargs = dict(
        response_mime_type="application/json",
        temperature=0,
        # Output JSON is short (~120 chars). Keep this small to cut
        # latency — Gemini 2.5 returns faster with a tight budget.
        max_output_tokens=400,
    )
    try:
        cfg_kwargs["thinking_config"] = genai_types.ThinkingConfig(thinking_budget=0)
    except Exception:
        # Older SDKs without ThinkingConfig — leave it off.
        pass

    try:
        client = genai.Client(api_key=api_key)
        resp = client.models.generate_content(
            model=model_name,
            contents=[
                genai_types.Part.from_bytes(data=raw_bytes, mime_type=media_type),
                prompt,
            ],
            config=genai_types.GenerateContentConfig(**cfg_kwargs),
        )
    except Exception as e:
        return "", {"error": f"Gemini call failed: {e}"}

    text_out = (getattr(resp, "text", None) or "").strip()
    if not text_out:
        # Fallback: dig through candidates
        try:
            text_out = resp.candidates[0].content.parts[0].text or ""
        except Exception:
            text_out = ""

    json_str = text_out.strip()
    if "```" in json_str:
        parts = json_str.split("```")
        for p in parts:
            if "{" in p:
                if p.startswith("json"):
                    p = p[4:]
                json_str = p
                break
    start = json_str.find("{")
    end = json_str.rfind("}")
    if start == -1 or end == -1:
        return text_out, {"error": "Gemini did not return JSON", "sku": None, "name": None, "meters": None}
    try:
        data = _json.loads(json_str[start:end + 1])
    except Exception:
        return text_out, {"error": "Gemini returned malformed JSON", "sku": None, "name": None, "meters": None}

    sku = (data.get("sku") or None)
    if sku:
        sku = _tr_upper(str(sku).strip()) or None
    name = (data.get("name") or None)
    if name:
        name = str(name).strip() or None
    meters = data.get("meters")
    try:
        meters = float(meters) if meters is not None else None
    except (TypeError, ValueError):
        meters = None
    barcode = (data.get("barcode") or None)
    if barcode:
        # Keep alphanumeric (letter prefixes like 'KL055555' are valid).
        import re as _re
        cleaned = _re.sub(r"[^A-Za-z0-9]", "", str(barcode)).upper()
        barcode = cleaned or None
    color = (data.get("color") or None)
    if color:
        color = str(color).strip() or None
    coupon = (data.get("coupon") or None)
    if coupon:
        coupon = str(coupon).strip() or None

    return text_out, {"sku": sku, "name": name, "meters": meters,
                      "barcode": barcode, "color": color, "coupon": coupon}


def _ocr_label_xai(image_file):
    """Send the label image to xAI Grok Vision and ask it to extract
    {sku, name, meters, barcode} as structured JSON.

    Uses xAI's OpenAI-compatible chat completions endpoint
    (https://api.x.ai/v1/chat/completions). No SDK needed.

    Requires XAI_API_KEY env var. Optional XAI_MODEL env var
    (defaults to 'grok-2-vision-latest').
    """
    import base64
    import json as _json

    try:
        import requests
    except ImportError:
        return "", {"error": "requests not installed"}

    api_key = _env("XAI_API_KEY")
    if not api_key:
        return "", {"error": "XAI_API_KEY env var not set"}

    try:
        image_file.seek(0)
    except Exception:
        pass
    raw_bytes = image_file.read()
    if not raw_bytes:
        return "", {"error": "Empty image"}

    media_type = "image/jpeg"
    if raw_bytes[:8] == b"\x89PNG\r\n\x1a\n":
        media_type = "image/png"
    elif raw_bytes[:6] in (b"GIF87a", b"GIF89a"):
        media_type = "image/gif"
    elif raw_bytes[:4] == b"RIFF" and raw_bytes[8:12] == b"WEBP":
        media_type = "image/webp"

    b64 = base64.b64encode(raw_bytes).decode("ascii")
    data_url = f"data:{media_type};base64,{b64}"

    prompt = (
        "This is a photo of a Turkish fabric roll label. Read EXACTLY "
        "what is printed in THIS image — do not invent, do not "
        "remember anything from previous images. Return ONLY a single "
        "JSON object — no commentary, no markdown fences:\n\n"
        '{"sku": "...", "name": "...", "meters": 48.5, "barcode": "...", "color": "...", "coupon": "..."}\n\n'
        "LABEL STRUCTURE (top → bottom):\n"
        "  1. BRAND   (e.g. KARVEN, DEMFIRAT, BELINO)  ← IGNORE, never extract\n"
        "  2. SKU     (short alphanumeric code)        → goes in 'sku'\n"
        "  3. MODEL   (descriptive product/model name) → goes in 'name'\n"
        "  4. METERS  (length in metres)               → goes in 'meters'\n"
        "  5. BARCODE (the code under the barcode bars) → goes in 'barcode'\n\n"
        "Field rules:\n"
        "- sku: the short product CODE printed on the second line "
        "(e.g. 'İPK0000174', 'K24614', 'RK48060RW9'). Alphanumeric, "
        "uppercase. Preserve Turkish letters (İ, Ş, Ğ, Ç, Ö, Ü) "
        "exactly. NOT the long barcode number, NOT the brand.\n"
        "- name: the MODEL NAME — the descriptive text printed "
        "DIRECTLY UNDER the sku. Examples: 'S-LINE 1106 EKRU', "
        "'KADİFE KREM', 'PANAMA 4500 GRİ', 'YAĞMUR DAMLASI ANTRASİT', "
        "'POLO-203 / KIRIK BEYAZ'. Read this descriptive line "
        "VERBATIM, exactly as printed. INCLUDE special characters "
        "(- / + & . ' Turkish letters İŞĞÇÖÜ) and any numbers that "
        "are part of the model name. DO NOT return the brand "
        "(KARVEN/DEMFIRAT/BELINO etc.) for this field — the brand "
        "is the FIRST line at the very top of the label and must be "
        "skipped. If there really is no model line between the sku "
        "and the meters/barcode area, return null.\n"
        "- meters: roll length as a decimal number. Turkish labels "
        "write '25,00 Metre' or '48,5 m' — interpret comma as decimal "
        "point ('25,00' → 25.0). No unit.\n"
        "- barcode: the code printed directly under the barcode strip. "
        "ALPHANUMERIC — may have a letter prefix like 'KL055555' or "
        "be all digits like '017400000228'. Copy EXACTLY what is "
        "printed, including any letters at the start. Typically 8 to "
        "16 chars. Different from the sku at the top.\n"
        "- color: a colour/variant on its own line if present "
        "(e.g. 'AÇIK KREM', 'EKRU', 'GÜMÜŞ'); else null.\n"
        "- coupon: the value after 'Kupon No' / 'Kupon' if present; else null.\n\n"
        "If a field is illegible or genuinely not in THIS image, set "
        "it to null. Do not invent values. Do not fall back to "
        "previous-scan values."
    )

    # xAI deprecated the dedicated grok-2-vision models — vision is now
    # built into the Grok 4 chat models. grok-4.3 is the fast non-reasoning
    # variant, ideal for OCR. Override with XAI_MODEL if you want
    # grok-4.20-0309-reasoning for tricky labels.
    model_name = _env("XAI_MODEL") or "grok-4.3"

    payload = {
        "model": model_name,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": data_url, "detail": "high"}},
                {"type": "text", "text": prompt},
            ],
        }],
        "temperature": 0,
        # The output JSON is short (~120 chars). Capping max_tokens
        # tight cuts response latency noticeably on Grok 4 — there's
        # no reason to budget for an essay.
        "max_tokens": 250,
        "response_format": {"type": "json_object"},
    }

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    try:
        resp = requests.post(
            "https://api.x.ai/v1/chat/completions",
            json=payload, headers=headers, timeout=30,
        )
    except requests.RequestException as e:
        return "", {"error": f"xAI call failed: {e}"}

    if resp.status_code != 200:
        snippet = (resp.text or "")[:400]
        return snippet, {"error": f"xAI HTTP {resp.status_code}: {snippet}"}

    try:
        body = resp.json()
        text_out = (body.get("choices", [{}])[0]
                        .get("message", {})
                        .get("content", "") or "").strip()
    except Exception as e:
        return resp.text, {"error": f"xAI bad response: {e}"}

    if not text_out:
        return "", {"error": "xAI returned empty content"}

    json_str = text_out.strip()
    if "```" in json_str:
        parts = json_str.split("```")
        for p in parts:
            if "{" in p:
                if p.startswith("json"):
                    p = p[4:]
                json_str = p
                break
    start = json_str.find("{")
    end = json_str.rfind("}")
    if start == -1 or end == -1:
        return text_out, {"error": "xAI did not return JSON", "sku": None, "name": None, "meters": None}
    try:
        data = _json.loads(json_str[start:end + 1])
    except Exception:
        return text_out, {"error": "xAI returned malformed JSON", "sku": None, "name": None, "meters": None}

    sku = (data.get("sku") or None)
    if sku:
        sku = _tr_upper(str(sku).strip()) or None
    name = (data.get("name") or None)
    if name:
        name = str(name).strip() or None
    meters = data.get("meters")
    try:
        meters = float(meters) if meters is not None else None
    except (TypeError, ValueError):
        meters = None
    barcode = (data.get("barcode") or None)
    if barcode:
        # Keep alphanumeric (letter prefixes like 'KL055555' are valid).
        # Strip whitespace/punctuation but preserve letters + digits.
        import re as _re
        cleaned = _re.sub(r"[^A-Za-z0-9]", "", str(barcode)).upper()
        barcode = cleaned or None
    color = (data.get("color") or None)
    if color:
        color = str(color).strip() or None
    coupon = (data.get("coupon") or None)
    if coupon:
        coupon = str(coupon).strip() or None

    return text_out, {"sku": sku, "name": name, "meters": meters,
                      "barcode": barcode, "color": color, "coupon": coupon}


def _ocr_label_tesseract(image_file):
    """Tesseract fallback — same return shape as _ocr_label_claude."""
    import os
    import re
    try:
        from PIL import Image, ImageOps
        import pytesseract
    except ImportError as e:
        return "", {"error": f"OCR libraries missing: {e}"}

    custom_cmd = os.environ.get("TESSERACT_CMD")
    if custom_cmd:
        pytesseract.pytesseract.tesseract_cmd = custom_cmd
    else:
        default_win = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(default_win):
            pytesseract.pytesseract.tesseract_cmd = default_win

    try:
        img = Image.open(image_file)
        img = ImageOps.exif_transpose(img)
        if max(img.size) < 1200:
            ratio = 1600 / max(img.size)
            img = img.resize(
                (int(img.size[0] * ratio), int(img.size[1] * ratio)),
                Image.LANCZOS,
            )
        img = ImageOps.grayscale(img)
        img = ImageOps.autocontrast(img, cutoff=2)
        try:
            installed_langs = pytesseract.get_languages(config="")
        except Exception:
            installed_langs = []
        lang = "eng+tur" if "tur" in installed_langs else "eng"
        raw = pytesseract.image_to_string(img, lang=lang, config="--psm 6")
    except FileNotFoundError:
        return "", {"error": "Tesseract binary not found"}
    except Exception as e:
        return "", {"error": f"OCR failed: {e}"}

    # ── Parsing ─────────────────────────────────────────────────
    # 1) SKU: alphanumeric, typically 5-15 chars, sometimes with hyphens or dots.
    # 2) Meters: a number followed by 'm' or 'mt' or 'metre' etc.
    # 3) Name: a line that's pure letters + spaces with no digits, longer than 3 chars.
    text = raw or ""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    sku = None
    meters = None
    name = None

    # Meters: 48.5 m / 48,5m / 48 metre / LEN 48.5
    meter_pat = re.compile(
        r"(\d{1,3}[.,]?\d{0,2})\s*(m|mt|metre|meters)\b",
        re.IGNORECASE,
    )

    # Pass 1 — explicit SKU/LEN labels
    for ln in lines:
        if not sku:
            m = re.search(r"\b(sku|kod|code|cod)\s*[:.\-]?\s*([A-Z0-9._-]{4,20})", ln, re.IGNORECASE)
            if m:
                sku = m.group(2).upper()
        if not meters:
            m = re.search(r"\b(len|length|metre|metraj|m)\s*[:.\-]?\s*(\d{1,3}[.,]?\d{0,2})", ln, re.IGNORECASE)
            if m:
                try:
                    meters = float(m.group(2).replace(",", "."))
                except ValueError:
                    pass

    # Pass 2 — bare-pattern fallbacks
    if not meters:
        for ln in lines:
            m = meter_pat.search(ln)
            if m:
                try:
                    meters = float(m.group(1).replace(",", "."))
                    break
                except ValueError:
                    pass
    if not sku:
        for ln in lines:
            # Heuristic: an upper-case alphanumeric token mixing letters AND digits
            for tok in re.findall(r"\b[A-Z0-9._-]{4,20}\b", ln.upper()):
                if any(c.isalpha() for c in tok) and any(c.isdigit() for c in tok):
                    sku = tok
                    break
            if sku:
                break
    if not name:
        for ln in lines:
            stripped = re.sub(r"[^A-Za-zÇĞİÖŞÜçğıöşü\s]", "", ln).strip()
            if len(stripped) >= 4 and any(c.isalpha() for c in stripped):
                # Avoid grabbing single-letter tokens by ensuring multi-word
                if len(stripped.split()) >= 1:
                    name = stripped
                    break

    # Tesseract fallback doesn't try to extract barcode reliably.
    parsed = {"sku": sku, "name": name, "meters": meters, "barcode": None}
    return raw, parsed


def _ocr_label(image_file):
    """Dispatcher — picks the best OCR backend available.

    Selection order:
      1. OCR_BACKEND env var ('gemini' | 'xai' | 'claude' | 'tesseract') wins if set.
      2. GEMINI_API_KEY or GOOGLE_API_KEY → Gemini (cheap + fast).
      3. XAI_API_KEY → xAI Grok Vision.
      4. ANTHROPIC_API_KEY → Claude (most accurate, slightly costlier).
      5. Tesseract fallback (works offline but mediocre on real labels).

    Vision-API errors auto-fall-back to the next backend so the user
    always gets *some* result back even on a network blip. The chain
    of attempts is recorded in `_attempts` so the debug panel can
    show exactly WHY each tried backend failed (e.g. "gemini: 503
    overload → falling back to xai").
    """
    backend = (_env("OCR_BACKEND") or "").lower().strip()
    has_openai = bool(_env("OPENAI_API_KEY"))
    has_gemini = bool(_env("GEMINI_API_KEY") or _env("GOOGLE_API_KEY"))
    has_xai = bool(_env("XAI_API_KEY"))
    has_claude = bool(_env("ANTHROPIC_API_KEY"))

    print(
        f"\n[OCR] ── new scan ── preferred={backend or 'auto'} "
        f"has_openai={has_openai} has_xai={has_xai} "
        f"has_gemini={has_gemini} has_claude={has_claude}",
        flush=True,
    )

    attempts = []  # list of {backend, error} for the debug panel

    def _safe_rewind():
        try:
            image_file.seek(0)
        except Exception:
            pass

    def _try(name, fn):
        _safe_rewind()
        print(f"[OCR] → trying {name}…", flush=True)
        raw, parsed = fn(image_file)
        if parsed.get("error"):
            err = parsed["error"]
            # Truncate huge HTML/error pages to keep the terminal readable.
            short_err = str(err)[:300]
            print(f"[OCR] ✗ {name} FAILED: {short_err}", flush=True)
            attempts.append({"backend": name, "error": err})
            return None
        print(
            f"[OCR] ✓ {name} OK — sku={parsed.get('sku')!r} "
            f"meters={parsed.get('meters')!r} barcode={parsed.get('barcode')!r} "
            f"name={parsed.get('name')!r}",
            flush=True,
        )
        # Show first 400 chars of raw response so the user can see what
        # the model literally returned (useful when JSON parse fails).
        if raw:
            raw_str = str(raw)
            print(f"[OCR]   raw: {raw_str[:400]}{'…' if len(raw_str) > 400 else ''}", flush=True)
        parsed["_backend"] = name
        parsed["_attempts"] = attempts + [{"backend": name, "error": None}]
        return raw, parsed

    # Build the try-order. The OCR_BACKEND env var (if set) just
    # promotes its choice to the front — it does NOT mean "only this".
    # If a preferred backend fails (no key, network error, quota), we
    # still walk the remaining vision APIs before dropping to
    # Tesseract. This is what the user actually wants: never silently
    # downgrade to bad OCR while a paid API is still available.
    #
    # Auto-order preference: openai → gemini → xai → claude → tesseract.
    # OpenAI (gpt-4o-mini, detail=high) leads by default — it's fast and
    # the user is out of Gemini credit. A set OCR_BACKEND still wins
    # (it's promoted to the front above). If a backend has no key / no
    # credit it errors and we fall through to the next.
    order = []
    if backend in ("openai", "gemini", "xai", "claude"):
        order.append(backend)
    if has_openai and "openai" not in order:
        order.append("openai")
    if has_gemini and "gemini" not in order:
        order.append("gemini")
    if has_xai and "xai" not in order:
        order.append("xai")
    if has_claude and "claude" not in order:
        order.append("claude")
    if backend == "tesseract":
        # User explicitly forces tesseract — honour it without trying APIs.
        order = ["tesseract"]
    else:
        order.append("tesseract")  # always the final fallback

    fns = {
        "openai": _ocr_label_openai,
        "gemini": _ocr_label_gemini,
        "xai": _ocr_label_xai,
        "claude": _ocr_label_claude,
        "tesseract": _ocr_label_tesseract,
    }

    for name in order:
        r = _try(name, fns[name])
        if r:
            return r

    return "", {"error": "all backends failed", "_attempts": attempts}


def reverse_consumption_for_order(order, user=None):
    """Restore stock that was previously consumed for `order`. Walks
    every StockMovement(out) whose reference matches the order's
    number/pk, returns the meters to the original roll (if still
    around), and writes a counter-balancing IN movement so the
    timeline shows both events. Used by OrderEdit + cancel flows."""
    from django.db import transaction as _tx
    if not order:
        return []

    order_ref = getattr(order, "order_number", None) or f"Order #{order.pk}"
    refs = [str(order_ref), str(order.pk), f"Order #{order.pk}"]
    if order.order_number:
        refs.append(order.order_number)

    moves = StockMovement.objects.filter(
        movement_type="out",
        reference__in=refs,
    ).select_related("product", "roll")

    restored = []
    with _tx.atomic():
        for mv in moves:
            wp = mv.product
            roll = mv.roll
            qty = mv.quantity or Decimal("0")
            if qty <= 0:
                continue
            # Return meters to the original roll, if it's still around.
            if roll and roll.product_id == wp.id:
                curr = roll.meters_remaining if roll.meters_remaining is not None else Decimal("0")
                roll.meters_remaining = curr + qty
                # Re-evaluate status.
                if roll.meters_remaining >= (roll.meters or Decimal("0")):
                    roll.status = "in_stock"
                else:
                    roll.status = "partial"
                roll.save(update_fields=["meters_remaining", "status"])
            wp.quantity = (wp.quantity or Decimal("0")) + qty
            wp.save(update_fields=["quantity", "updated_at"])

            StockMovement.objects.create(
                product=wp,
                roll=roll,
                movement_type="in",
                quantity=qty,
                reason=f"Order edit · reversed {order_ref}",
                reference=str(order_ref),
                created_by=user if (user and getattr(user, "is_authenticated", False)) else None,
            )
            restored.append({
                "product": wp.id, "roll": roll.id if roll else None,
                "qty": float(qty),
            })
    return restored


def consume_for_order_items(order, user=None, reason_prefix="Order"):
    """Decrement warehouse stock for every OrderItem in `order`.

    Algorithm:
      1. For each OrderItem find a WarehouseProduct whose sku/barcode
         matches the item's product/variant. Search ALL warehouses;
         prefer ones already holding stock (quantity > 0).
      2. Walk that product's rolls oldest-first, taking meters from
         each until the OrderItem.quantity is satisfied.
      3. Each consumption emits a StockMovement(out) referencing the
         order (its number/pk in the `reference` field) and the
         specific roll.
      4. If no warehouse has enough stock for an item, log the gap
         on the StockMovement.reason but still consume what we can.

    Returns a list of dicts describing each consumption (for logging /
    UI feedback).
    """
    from django.db import transaction as _tx
    results = []
    if not order:
        return results

    order_ref = (
        getattr(order, "order_number", None)
        or f"Order #{order.pk}"
    )

    items = list(order.items.all().select_related("product", "product_variant"))
    if not items:
        return results

    with _tx.atomic():
        for it in items:
            qty_needed = Decimal(str(it.quantity or 0))
            if qty_needed <= 0:
                continue
            # Resolve SKU — prefer variant SKU.
            sku = None
            if it.product_variant_id and it.product_variant:
                sku = getattr(it.product_variant, "variant_sku", None)
            if not sku and it.product_id and it.product:
                sku = getattr(it.product, "sku", None)
            if not sku:
                continue

            # Find candidate WarehouseProducts (any warehouse, with stock first).
            wp_qs = (
                WarehouseProduct.objects
                .filter(sku__iexact=sku)
                .order_by("-quantity", "id")
            )
            remaining_needed = qty_needed
            for wp in wp_qs:
                if remaining_needed <= 0:
                    break
                if (wp.quantity or Decimal("0")) <= 0:
                    continue
                # Walk this warehouse's active rolls oldest first.
                roll_qs = (
                    wp.rolls.exclude(status="consumed")
                    .order_by("scanned_at")
                )
                for roll in roll_qs:
                    if remaining_needed <= 0:
                        break
                    avail = (
                        roll.meters_remaining
                        if roll.meters_remaining is not None
                        else roll.meters
                    ) or Decimal("0")
                    if avail <= 0:
                        continue
                    take = min(avail, remaining_needed)
                    roll.meters_remaining = avail - take
                    if roll.meters_remaining <= 0:
                        roll.status = "consumed"
                    elif roll.meters_remaining < (roll.meters or Decimal("0")):
                        roll.status = "partial"
                    roll.save(update_fields=["meters_remaining", "status"])

                    wp.quantity = (wp.quantity or Decimal("0")) - take
                    wp.save(update_fields=["quantity", "updated_at"])

                    StockMovement.objects.create(
                        product=wp,
                        roll=roll,
                        movement_type="out",
                        quantity=take,
                        reason=f"{reason_prefix} {order_ref}",
                        reference=str(order_ref),
                        created_by=user if (user and getattr(user, "is_authenticated", False)) else None,
                    )
                    results.append({
                        "warehouse": wp.warehouse_id,
                        "product": wp.id,
                        "roll": roll.id,
                        "taken": float(take),
                        "sku": sku,
                    })
                    remaining_needed -= take
            if remaining_needed > 0:
                # Log unfulfilled — a "ghost" movement on the first matching
                # wp so the user can see it on the product timeline.
                wp = WarehouseProduct.objects.filter(sku__iexact=sku).first()
                if wp:
                    StockMovement.objects.create(
                        product=wp,
                        roll=None,
                        movement_type="adjustment",
                        quantity=remaining_needed,
                        reason=f"⚠️ Shortage: {reason_prefix} {order_ref} needed {remaining_needed}m more",
                        reference=str(order_ref),
                        created_by=user if (user and getattr(user, "is_authenticated", False)) else None,
                    )
                results.append({
                    "warehouse": None, "product": None, "roll": None,
                    "taken": 0, "sku": sku,
                    "shortage": float(remaining_needed),
                })

    return results


# ────────────────────────────────────────────────────────────────────
#  Packing reservations — the order lifecycle. Rolls are scanned during
#  the packing step and RESERVED (a soft hold: no physical cut). The
#  physical stock is only cut when the order ships (cargo info entered).
# ────────────────────────────────────────────────────────────────────
def reserved_meters_for_products(product_ids):
    """Return {warehouse_product_id: reserved_meters(Decimal)} for the
    given products, summing only ACTIVE (unconsumed) reservations."""
    from django.db.models import Sum
    from .models import OrderRollReservation
    ids = [p for p in (product_ids or []) if p]
    if not ids:
        return {}
    rows = (OrderRollReservation.objects
            .filter(consumed=False, warehouse_product_id__in=ids)
            .values("warehouse_product_id")
            .annotate(s=Sum("meters")))
    return {r["warehouse_product_id"]: (r["s"] or Decimal("0")) for r in rows}


def reserved_meters_subquery():
    """A Coalesced Subquery summing active reserved metres per
    WarehouseProduct — safe to .annotate() alongside Sum()/Count() over
    the rolls join without multiplying those aggregates."""
    from django.db.models import OuterRef, Subquery, Sum, DecimalField as _DF
    from .models import OrderRollReservation
    sq = (OrderRollReservation.objects
          .filter(warehouse_product_id=OuterRef("pk"), consumed=False)
          .values("warehouse_product_id")
          .annotate(s=Sum("meters"))
          .values("s")[:1])
    return Coalesce(
        Subquery(sq, output_field=_DF(max_digits=18, decimal_places=2)),
        Decimal("0"), output_field=_DF(max_digits=18, decimal_places=2),
    )


def consume_reservations_for_order(order, user=None, reason_prefix="Order ship"):
    """Convert every ACTIVE reservation on `order` into a real
    stock-out: cut the reserved metres from the exact scanned roll, drop
    WarehouseProduct.quantity, write StockMovement(out), and mark the
    reservation consumed. Idempotent — already-consumed reservations are
    skipped, so re-saving 'shipped' never double-cuts. Returns a list
    describing each cut."""
    from django.db import transaction as _tx
    from django.utils import timezone as _tz
    from .models import OrderRollReservation
    results = []
    if not order:
        return results
    order_ref = getattr(order, "order_number", None) or f"Order #{order.pk}"
    with _tx.atomic():
        resv = list(OrderRollReservation.objects
                    .select_for_update()
                    .filter(order=order, consumed=False))
        # Fetch each distinct roll / product ONCE (locked) and share the
        # instances across the loop. With select_related, every
        # reservation row carried its OWN pre-loop copy of the roll and
        # product, so two reservations on the same roll each subtracted
        # from the ORIGINAL remaining and the last save won — under-
        # cutting stock exactly when one order line is covered by
        # several rolls (which the coverage gate now requires).
        _roll_ids = {r.roll_id for r in resv if r.roll_id}
        _wp_ids = {r.warehouse_product_id for r in resv if r.warehouse_product_id}
        _rolls = {x.pk: x for x in WarehouseProductRoll.objects
                  .select_for_update().filter(pk__in=_roll_ids)}
        _wps = {x.pk: x for x in WarehouseProduct.objects
                .select_for_update().filter(pk__in=_wp_ids)}
        for r in resv:
            roll = _rolls.get(r.roll_id)
            wp = _wps.get(r.warehouse_product_id)
            take = Decimal(str(r.meters or 0))
            if take <= 0 or roll is None or wp is None:
                r.consumed = True
                r.consumed_at = _tz.now()
                r.save(update_fields=["consumed", "consumed_at"])
                continue
            avail = (roll.meters_remaining if roll.meters_remaining is not None else roll.meters) or Decimal("0")
            actual = take if take <= avail else avail  # clamp — never negative
            roll.meters_remaining = avail - actual
            if roll.meters_remaining <= 0:
                roll.status = "consumed"
            elif roll.meters_remaining < (roll.meters or Decimal("0")):
                roll.status = "partial"
            roll.save(update_fields=["meters_remaining", "status"])

            wp.quantity = (wp.quantity or Decimal("0")) - actual
            wp.save(update_fields=["quantity", "updated_at"])

            StockMovement.objects.create(
                product=wp, roll=roll, movement_type="out",
                quantity=actual, reason=f"{reason_prefix} {order_ref}",
                reference=str(order_ref),
                created_by=user if (user and getattr(user, "is_authenticated", False)) else None,
            )
            # If the roll physically had less than was reserved (it got cut
            # elsewhere after reservation), record the shortfall so it's
            # visible on the timeline instead of silently swallowed.
            if actual < take:
                StockMovement.objects.create(
                    product=wp, roll=None, movement_type="adjustment",
                    quantity=(take - actual),
                    reason=f"⚠️ Shortage: {reason_prefix} {order_ref} reserved {take}m but only {actual}m was available",
                    reference=str(order_ref),
                    created_by=user if (user and getattr(user, "is_authenticated", False)) else None,
                )
            # Pin the reservation to what was ACTUALLY cut so an un-ship
            # restores exactly that amount (never inflating the roll).
            r.meters = actual
            r.consumed = True
            r.consumed_at = _tz.now()
            r.save(update_fields=["meters", "consumed", "consumed_at"])
            results.append({"product": wp.id, "roll": roll.id, "taken": float(actual)})
    return results


def restore_reservations_for_order(order, user=None, reason_prefix="Order un-ship"):
    """Reverse a shipped order's stock-out: add the consumed metres back
    to each roll + WarehouseProduct, write StockMovement(in), and flip
    the reservation back to active (unconsumed). Used when an order
    moves OUT of 'shipped' (e.g. reverted to packing / cancelled)."""
    from django.db import transaction as _tx
    from .models import OrderRollReservation
    restored = []
    if not order:
        return restored
    order_ref = getattr(order, "order_number", None) or f"Order #{order.pk}"
    with _tx.atomic():
        resv = list(OrderRollReservation.objects
                    .select_for_update()
                    .filter(order=order, consumed=True))
        # Shared locked instances per roll/product — same stale-copy fix
        # as consume_reservations_for_order (see comment there).
        _roll_ids = {r.roll_id for r in resv if r.roll_id}
        _wp_ids = {r.warehouse_product_id for r in resv if r.warehouse_product_id}
        _rolls = {x.pk: x for x in WarehouseProductRoll.objects
                  .select_for_update().filter(pk__in=_roll_ids)}
        _wps = {x.pk: x for x in WarehouseProduct.objects
                .select_for_update().filter(pk__in=_wp_ids)}
        for r in resv:
            roll = _rolls.get(r.roll_id)
            wp = _wps.get(r.warehouse_product_id)
            qty = Decimal(str(r.meters or 0))
            if qty > 0 and roll is not None and wp is not None:
                curr = roll.meters_remaining if roll.meters_remaining is not None else Decimal("0")
                new_remaining = curr + qty
                # Never restore a roll above its original length.
                cap = roll.meters if roll.meters is not None else new_remaining
                if new_remaining > cap:
                    new_remaining = cap
                roll.meters_remaining = new_remaining
                if roll.meters_remaining >= (roll.meters or Decimal("0")):
                    roll.status = "in_stock"
                else:
                    roll.status = "partial"
                roll.save(update_fields=["meters_remaining", "status"])
                wp.quantity = (wp.quantity or Decimal("0")) + qty
                wp.save(update_fields=["quantity", "updated_at"])
                StockMovement.objects.create(
                    product=wp, roll=roll, movement_type="in",
                    quantity=qty, reason=f"{reason_prefix} {order_ref}",
                    reference=str(order_ref),
                    created_by=user if (user and getattr(user, "is_authenticated", False)) else None,
                )
            r.consumed = False
            r.consumed_at = None
            r.save(update_fields=["consumed", "consumed_at"])
            restored.append({"product": wp.id if wp else None, "roll": roll.id if roll else None, "qty": float(qty)})
    return restored


SHIPPED_CLASS = frozenset({"shipped", "in_transit", "out_for_delivery", "delivered"})


def order_reservation_shortfalls(order):
    """Items whose ACTIVE (unconsumed) roll reservations do NOT cover
    the ordered quantity. Returns [{item_id, name, required, reserved}]
    — an empty list means the order is fully covered and may complete
    packing / ship.

    Rules:
      * required metres = item.quantity, except custom-curtain lines,
        which use their fabric-used metres when set (their quantity
        counts curtains, not metres);
      * items whose product has NO warehouse presence at all (nothing
        that could ever be scanned) are skipped, so catalog-only lines
        can't dead-end an order. Presence is checked at the level the
        pack scan matches at: a line naming a specific variant needs
        THAT variant (or its SKU) in a warehouse — a sibling variant
        of the same parent doesn't count;
      * reservations count per order line (order_item FK)."""
    from functools import reduce
    import operator as _op
    from django.db.models import Sum, Q as _Q
    from .models import OrderRollReservation, WarehouseProduct

    items = list(order.items.all().select_related("product", "product_variant"))
    if not items:
        return []

    reserved = {
        r["order_item_id"]: (r["s"] or Decimal("0"))
        for r in (OrderRollReservation.objects
                  .filter(order=order, consumed=False, order_item__isnull=False)
                  .values("order_item_id").annotate(s=Sum("meters")))
    }

    variant_ids = {it.product_variant_id for it in items if it.product_variant_id}
    product_ids = {it.product_id for it in items if it.product_id}
    skus = set()
    for it in items:
        for s in ((getattr(it.product_variant, "variant_sku", None) if it.product_variant_id else None),
                  (getattr(it.product, "sku", None) if it.product_id else None)):
            if s and s.strip():
                skus.add(s.strip().lower())

    wq = _Q(pk__in=[])
    if variant_ids:
        wq |= _Q(catalog_variant_id__in=variant_ids)
    if product_ids:
        wq |= _Q(catalog_variant__product_id__in=product_ids)
    if skus:
        wq |= reduce(_op.or_, (_Q(sku__iexact=s) for s in skus))
    tracked_variants, tracked_products, tracked_skus = set(), set(), set()
    for cv, cvp, wsku in (WarehouseProduct.objects.filter(wq)
                          .values_list("catalog_variant_id",
                                       "catalog_variant__product_id", "sku")):
        if cv:
            tracked_variants.add(cv)
        if cvp:
            tracked_products.add(cvp)
        if wsku and wsku.strip():
            tracked_skus.add(wsku.strip().lower())

    shortfalls = []
    for it in items:
        vsku = ((getattr(it.product_variant, "variant_sku", "") or "").strip().lower()
                if it.product_variant_id else "")
        psku = ((getattr(it.product, "sku", "") or "").strip().lower()
                if it.product_id else "")
        if it.product_variant_id:
            tracked = (it.product_variant_id in tracked_variants
                       or (vsku and vsku in tracked_skus))
        else:
            tracked = ((it.product_id and it.product_id in tracked_products)
                       or (psku and psku in tracked_skus))
        if not tracked:
            continue
        required = it.quantity or Decimal("0")
        if getattr(it, "is_custom_curtain", False) and it.custom_fabric_used_meters:
            required = it.custom_fabric_used_meters
        if required <= 0:
            continue
        got = reserved.get(it.pk, Decimal("0"))
        if got < required:
            name = ((it.product.title if it.product_id else None)
                    or it.description or f"Kalem #{it.pk}")
            if it.product_variant_id and vsku:
                name = f"{name} ({vsku.upper()})"
            shortfalls.append({"item_id": it.pk, "name": name,
                               "required": required, "reserved": got})
    return shortfalls


def reservation_shortfall_message(order):
    """Turkish, user-facing summary of what's still missing — shared by
    every path that blocks on incomplete coverage."""
    rows = order_reservation_shortfalls(order)
    if not rows:
        return ""
    det = " · ".join(
        f"{r['name']}: {float(r['reserved']):g}/{float(r['required']):g} m"
        for r in rows[:4]
    )
    more = f" (+{len(rows) - 4} kalem)" if len(rows) > 4 else ""
    return (f"Okutulan toplar siparişi karşılamıyor — {det}{more}. "
            "Kalan metreleri paketleme sayfasından okutun.")


def apply_order_status_change(order, new_status, carrier=None, tracking=None,
                             user=None, require_cargo_for_ship=True):
    """THE single funnel for changing an order's order_status. Every
    write path (order detail page, JSON API, web-order edit) must go
    through here so the lifecycle rules can't diverge:

      * entering a shipped-class status requires carrier + tracking
        (unless require_cargo_for_ship=False) — else nothing changes;
      * entering 'packaging' AND entering a shipped-class status both
        require the scanned-roll reservations to fully COVER every
        order line's metres (see order_reservation_shortfalls) — a
        110 m line backed by a single 38 m roll cannot complete;
      * entering 'cancelled' releases every active roll reservation —
        cancelling a shipped order first restores its cut stock (the
        leaving_ship branch below), then this drops the (now-inactive)
        reservation rows entirely, so a cancelled order never keeps
        warehouse capacity falsely held for stock that'll never ship;
      * the status change, its catalog-stock signal, AND the warehouse
        reservation → stock-out conversion all happen inside ONE
        transaction, so a roll-cut failure rolls the whole thing back
        (no half-shipped, catalog-cut-but-warehouse-not state).

    Returns (ok: bool, code: str|None). code is a machine-readable
    reason on failure: 'cargo_required' | 'insufficient_reservation' |
    'error:<detail>'."""
    from django.db import transaction as _tx
    from django.utils import timezone as _tz
    from .models import ORDER_STATUS_CHOICES, CARRIER_CHOICES

    valid_statuses = {k for k, _ in ORDER_STATUS_CHOICES}
    valid_carriers = {k for k, _ in CARRIER_CHOICES}
    old_status = order.order_status

    # Stash any cargo info the caller carried.
    if carrier is not None:
        c = (carrier or "").strip()
        if c:
            order.carrier = c if c in valid_carriers else None
    if tracking is not None:
        tstr = (tracking or "").strip()
        if tstr:
            order.tracking_number = tstr

    changing = bool(new_status) and new_status in valid_statuses and new_status != old_status
    entering_ship = changing and new_status in SHIPPED_CLASS and old_status not in SHIPPED_CLASS
    leaving_ship = (old_status in SHIPPED_CLASS and new_status in valid_statuses
                    and new_status not in SHIPPED_CLASS and new_status != old_status)
    entering_cancelled = changing and new_status == "cancelled" and old_status != "cancelled"
    leaving_cancelled = changing and old_status == "cancelled"

    # Gate: shipping needs cargo info.
    if entering_ship and require_cargo_for_ship:
        if not (order.carrier or "").strip() or not (order.tracking_number or "").strip():
            order.save(update_fields=["carrier", "tracking_number", "updated_at"])
            return False, "cargo_required"

    # Gate: packing-complete AND ship both require the reservations to
    # fully cover every line's metres — not merely "a roll was scanned".
    # Coming BACK from a shipped status is exempt: the reservations are
    # still consumed at this point (restore runs inside the transaction
    # below), so the coverage check would always fail — and coverage was
    # already enforced on the way in.
    if entering_ship or (changing and new_status == "packaging"
                         and old_status != "packaging"
                         and old_status not in SHIPPED_CLASS):
        if order_reservation_shortfalls(order):
            # Persist any cargo edits but don't advance the status.
            order.save(update_fields=["carrier", "tracking_number", "updated_at"])
            return False, "insufficient_reservation"

    try:
        with _tx.atomic():
            if changing:
                order.order_status = new_status
                if new_status == "shipped" and not order.shipped_at:
                    order.shipped_at = _tz.now()
                if new_status == "delivered" and not order.delivered_at:
                    order.delivered_at = _tz.now()
            order.save(update_fields=["order_status", "carrier", "tracking_number",
                                      "shipped_at", "delivered_at", "updated_at"])
            if entering_ship:
                consume_reservations_for_order(order, user=user)
            elif leaving_ship:
                restore_reservations_for_order(order, user=user)
                # Un-shipped order is no longer a completed sale — its
                # invoice may not stay live (invoices only exist for
                # completed orders). Re-shipping issues a fresh one.
                for _inv in order.invoices.exclude(status="cancelled"):
                    _inv.cancel(user=user, reason=f"Order #{order.pk} un-shipped")
            if entering_cancelled:
                # Never any physical stock to restore here: leaving_ship
                # (if it also fired) already converted consumed rows back
                # to active ones above, and a non-shipped order's
                # reservations were never consumed in the first place —
                # either way they're plain soft holds now, safe to drop.
                order.roll_reservations.filter(consumed=False).delete()
                # A cancelled order must vanish from the books: the
                # order_sale movement comes off the cari (retail posts
                # at create too now), and any invoice cut from this
                # order is cancelled with it (its counter-movement is
                # 0-amount, the order movement carries the receivable —
                # see Invoice.issue()).
                from current_account.services import reverse_order_movement
                reverse_order_movement(order)
                for _inv in order.invoices.exclude(status="cancelled"):
                    _inv.cancel(user=user, reason=f"Order #{order.pk} cancelled")
            elif leaving_cancelled:
                # Re-opened order goes back on the books.
                if order.cari_id:
                    from current_account.services import post_order_movement
                    post_order_movement(order)
            # Retail money leg — completion posts the sale + auto
            # collection to the shared "Perakende Satışları" cari and
            # mirrors it into the Perakende defter; un-ship reverses
            # all of it. Inside the SAME transaction so status, stock
            # and money can never diverge.
            if getattr(order, "is_retail_order", False):
                from current_account.services import (
                    post_retail_order_financials,
                    reverse_retail_order_financials,
                )
                if entering_ship:
                    post_retail_order_financials(order, user=user)
                elif leaving_ship:
                    reverse_retail_order_financials(order, user=user)
            # Completed sale → its invoice cuts itself, retail included
            # — no manual step. Runs AFTER the retail leg so a legacy
            # retail order that only gets its cari linked there is still
            # invoiceable. Swallow-and-log: a numbering hiccup must
            # never un-ship the order (the order-detail button stays as
            # the manual fallback).
            if entering_ship:
                try:
                    from current_account.services import create_invoice_for_order
                    create_invoice_for_order(order, user=user)
                except Exception:
                    import traceback
                    traceback.print_exc()
    except Exception as _e:
        return False, f"error:{_e}"
    return True, None


@method_decorator(login_required, name='dispatch')
class WarehouseProductDetail(View):
    """Product detail under a warehouse — shows the product header
    info, every roll (top) with its barcode + remaining meters, and
    the full stock-movement timeline (in/out/adjustments)."""

    template_name = "operating/warehouse_product_detail.html"

    def get(self, request, warehouse_pk, product_pk):
        # One fetch resolves product + warehouse + the linked catalog
        # product (the template reads it) — each avoided query is a full
        # round-trip to the remote DB.
        product = get_object_or_404(
            WarehouseProduct.objects.select_related(
                "warehouse", "catalog_variant__product"),
            pk=product_pk, warehouse_id=warehouse_pk,
        )
        warehouse = product.warehouse
        rolls = list(product.rolls.all().order_by("-scanned_at"))
        movements = product.movements.all().select_related("roll", "created_by")[:200]

        # Aggregate quick stats — in/out totals in a single query.
        from django.db.models import Sum, Q as _Q
        from decimal import Decimal as _Dec
        from .models import OrderRollReservation
        _io = product.movements.aggregate(
            i=Sum("quantity", filter=_Q(movement_type="in")),
            o=Sum("quantity", filter=_Q(movement_type="out")),
        )
        in_total = _io["i"] or _Dec("0")
        out_total = _io["o"] or _Dec("0")

        # Active (unconsumed) reservations = packed-but-not-shipped metres
        # held against this product's rolls. Shown as a "Rezerv stok" card
        # and per-roll on the rolls list — each with the order it's held
        # for, so "Rezerv: 40 m" is traceable instead of a dead-end number.
        _resv_rows = (OrderRollReservation.objects
                      .filter(warehouse_product=product, consumed=False)
                      .select_related("order")
                      .order_by("-created_at"))
        reserved_by_roll = {}
        reservations_by_roll = {}
        for _resv in _resv_rows:
            reserved_by_roll[_resv.roll_id] = reserved_by_roll.get(_resv.roll_id, _Dec("0")) + (_resv.meters or _Dec("0"))
            reservations_by_roll.setdefault(_resv.roll_id, []).append(_resv)
        reserved_total = sum(reserved_by_roll.values(), _Dec("0"))
        for r in rolls:
            r.reserved_m = reserved_by_roll.get(r.id, _Dec("0"))
            r.reservation_list = reservations_by_roll.get(r.id, [])

        return render(request, self.template_name, {
            "warehouse": warehouse,
            "product": product,
            "rolls": rolls,
            "movements": movements,
            "rolls_count": len(rolls),
            "in_total": in_total,
            "out_total": out_total,
            "reserved_total": reserved_total,
            "active_rolls_count": sum(1 for r in rolls if r.status != "consumed"),
            "is_admin": _is_admin(request.user),
        })


# ────────────────────────────────────────────────────────────────────
#  Activity categorization — StockMovement rows carry their origin only
#  in free-text reason/reference, so the feed derives a "kind" from the
#  known prefixes each writer uses:
#    order      → ship/un-ship/edit-reversal cuts (reason "Order …" /
#                 "Web order …", or an order reference "DK…"/"Order #…")
#    in         → product intake (roll scan, manual add, import)
#    out        → manual stock-out
#    adjustment → roll delete / meters edit / manual override / edits
# ────────────────────────────────────────────────────────────────────
# Exact reason prefixes the order-lifecycle writers use. Deliberately
# NOT a bare "Order"/"DK" prefix match: manual stock-out reasons and
# roll barcodes are free-text (a "Deka" supplier mints DK-prefixed
# barcodes, a user may type "Ordered by X"), which would misclassify
# intake/adjustment rows as orders.
_ORDER_REASON_PREFIXES = ("Order ship", "Order un-ship", "Order edit")
# Order numbers are DK + zfill(7) digits (models.generate_order_number).
_ORDER_REF_RE = r"^DK\d{7,}$"


def _order_movement_q():
    from django.db.models import Q as _Q
    q = _Q(reference__regex=_ORDER_REF_RE) | _Q(reference__startswith="Order #")
    for p in _ORDER_REASON_PREFIXES:
        q |= _Q(reason__startswith=p)
    return q


def _movement_is_order(m):
    import re as _re
    r = (m.reason or "")
    ref = (m.reference or "")
    return (r.startswith(_ORDER_REASON_PREFIXES)
            or bool(_re.match(_ORDER_REF_RE, ref))
            or ref.startswith("Order #"))


# English reason prefixes (as stored in the DB) → Turkish display text.
# Order matters — longest/most-specific prefix first.
_REASON_TR = [
    ("Order edit · reversed", "Sipariş düzenleme iadesi"),
    ("Order un-ship", "Sipariş sevk iptali"),
    ("Order ship", "Sipariş sevkiyatı"),
    ("Web order", "Web siparişi"),
    ("Order", "Sipariş"),
    ("Roll scanned", "Top okutuldu"),
    ("Manual add", "Manuel ekleme"),
    ("Manual stock-out", "Manuel stok çıkışı"),
    ("Bulk roll delete", "Toplu top silme"),
    ("Roll deleted", "Top silindi"),
    ("Roll meters edited", "Top metresi düzenlendi"),
    ("Manual adjustment", "Manuel düzeltme"),
    ("Product edited", "Ürün düzenlendi"),
    ("⚠️ Shortage", "⚠️ Eksik"),
]


def _reason_display(reason):
    """Best-effort Turkish rendering of a stored English reason string.
    Unknown reasons pass through unchanged."""
    if not reason:
        return reason
    for en, tr in _REASON_TR:
        if reason.startswith(en):
            return tr + reason[len(en):]
    return reason


def _decorate_movements(movements):
    """Attach .kind / .reason_display / .order_pk to each page row so
    the templates can render category badges, Turkish reasons and a
    link to the source order."""
    from django.utils import translation
    import re as _re
    from .models import Order
    is_tr = (translation.get_language() or "").startswith("tr")
    refs = {m.reference for m in movements
            if m.reference and _re.match(_ORDER_REF_RE, m.reference)}
    order_map = dict(Order.objects.filter(order_number__in=refs)
                     .values_list("order_number", "pk")) if refs else {}
    for m in movements:
        m.kind = "order" if _movement_is_order(m) else m.movement_type
        m.reason_display = _reason_display(m.reason) if is_tr else m.reason
        ref = m.reference or ""
        m.order_pk = order_map.get(ref)
        if m.order_pk is None and ref.startswith("Order #") and ref[7:].strip().isdigit():
            m.order_pk = int(ref[7:].strip())
    return movements


@method_decorator(login_required, name='dispatch')
class WarehouseMovementsAll(View):
    """GLOBAL activity feed ("Son Hareketler") — every StockMovement
    across ALL warehouses, filterable by date range, category (order /
    intake / stock-out / adjustment), warehouse and free text.
    Query params:
      kind     : order | in | out | adjustment ('' = all)
      warehouse: Warehouse.pk ('' = all)
      from/to  : YYYY-MM-DD inclusive (timezone-aware via __date)
      q        : free text across reason/reference/product name/sku/roll
                 barcode (Turkish İ/ı-folded)
      page     : pagination
    """
    template_name = "operating/warehouse_movements_all.html"

    def get(self, request):
        from django.db.models import Q as _Q, Sum
        from datetime import datetime as _dt
        from functools import reduce
        import operator

        qs = (StockMovement.objects
              .select_related("product", "product__warehouse", "roll", "created_by"))

        wh_param = (request.GET.get("warehouse") or "").strip()
        if wh_param.isdigit():
            qs = qs.filter(product__warehouse_id=int(wh_param))

        date_from = (request.GET.get("from") or "").strip()
        date_to = (request.GET.get("to") or "").strip()
        try:
            if date_from:
                qs = qs.filter(created_at__date__gte=_dt.strptime(date_from, "%Y-%m-%d").date())
        except ValueError:
            pass
        try:
            if date_to:
                qs = qs.filter(created_at__date__lte=_dt.strptime(date_to, "%Y-%m-%d").date())
        except ValueError:
            pass

        q = (request.GET.get("q") or "").strip()
        if q:
            variants = _tr_ci_variants(q)

            def _fq(field):
                return reduce(operator.or_,
                              (_Q(**{f"{field}__icontains": v}) for v in variants))
            qs = qs.filter(_fq("reason") | _fq("reference") | _fq("product__name")
                           | _fq("product__sku") | _fq("roll__barcode"))

        # Stats reflect every filter EXCEPT the category, so picking
        # "Sipariş" doesn't zero the intake/out totals.
        stats_qs = qs
        in_total = stats_qs.filter(movement_type="in").aggregate(s=Sum("quantity"))["s"] or 0
        out_total = stats_qs.filter(movement_type="out").aggregate(s=Sum("quantity"))["s"] or 0
        order_count = stats_qs.filter(_order_movement_q()).count()

        kind = (request.GET.get("kind") or "").strip()
        if kind == "order":
            qs = qs.filter(_order_movement_q())
        elif kind in ("in", "out", "adjustment"):
            qs = qs.filter(movement_type=kind).exclude(_order_movement_q())

        qs = qs.order_by("-created_at")

        paginator = Paginator(qs, 50)
        try:
            page = paginator.page(request.GET.get("page", 1))
        except (EmptyPage, PageNotAnInteger, ValueError):
            page = paginator.page(1)

        movements = _decorate_movements(list(page.object_list))

        qd = request.GET.copy()
        qd.pop("page", None)
        filter_qs = qd.urlencode()

        return render(request, self.template_name, {
            "page_obj": page,
            "paginator": paginator,
            "movements": movements,
            "in_total": in_total,
            "out_total": out_total,
            "net": (in_total or 0) - (out_total or 0),
            "order_count": order_count,
            "warehouses": list(Warehouse.objects.order_by("name").values("id", "name")),
            "filter_qs": filter_qs,
            "filters": {
                "kind": kind,
                "warehouse": wh_param,
                "from": date_from,
                "to": date_to,
                "q": q,
            },
        })


@method_decorator(login_required, name='dispatch')
class WarehouseMovements(View):
    """Filterable, paginated stock-movement ledger for a warehouse.
    Query params:
      type   : in | out | adjustment (single or comma-separated)
      from   : YYYY-MM-DD inclusive
      to     : YYYY-MM-DD inclusive
      product: WarehouseProduct.pk
      q      : free-text search across reason/reference/product name/sku
      min_qty: minimum |quantity|
      max_qty: maximum |quantity|
      page   : pagination
    """
    template_name = "operating/warehouse_movements.html"

    def get(self, request, warehouse_pk):
        from django.db.models import Q as _Q
        from datetime import datetime as _dt
        warehouse = get_object_or_404(Warehouse, pk=warehouse_pk)

        qs = (
            StockMovement.objects
            .filter(product__warehouse=warehouse)
            .select_related("product", "roll", "created_by")
        )

        types_param = (request.GET.get("type") or "").strip()

        date_from = (request.GET.get("from") or "").strip()
        date_to = (request.GET.get("to") or "").strip()
        try:
            if date_from:
                d = _dt.strptime(date_from, "%Y-%m-%d").date()
                qs = qs.filter(created_at__date__gte=d)
        except ValueError:
            pass
        try:
            if date_to:
                d = _dt.strptime(date_to, "%Y-%m-%d").date()
                qs = qs.filter(created_at__date__lte=d)
        except ValueError:
            pass

        product_pk = (request.GET.get("product") or "").strip()
        if product_pk.isdigit():
            qs = qs.filter(product_id=int(product_pk))

        q = (request.GET.get("q") or "").strip()
        if q:
            # Turkish İ/ı-aware case folding, same as the product list search.
            from functools import reduce
            import operator
            variants = _tr_ci_variants(q)

            def _fq(field):
                return reduce(operator.or_,
                              (_Q(**{f"{field}__icontains": v}) for v in variants))
            qs = qs.filter(_fq("reason") | _fq("reference") | _fq("product__name")
                           | _fq("product__sku") | _fq("roll__barcode"))

        from decimal import Decimal as _D
        min_qty = (request.GET.get("min_qty") or "").strip().replace(",", ".")
        max_qty = (request.GET.get("max_qty") or "").strip().replace(",", ".")
        try:
            if min_qty:
                qs = qs.filter(quantity__gte=_D(min_qty))
        except Exception:
            pass
        try:
            if max_qty:
                qs = qs.filter(quantity__lte=_D(max_qty))
        except Exception:
            pass

        # Quick stats — computed BEFORE the category/type filter so
        # picking a category doesn't zero the other cards.
        from django.db.models import Sum
        in_total = qs.filter(movement_type="in").aggregate(s=Sum("quantity"))["s"] or 0
        out_total = qs.filter(movement_type="out").aggregate(s=Sum("quantity"))["s"] or 0
        net = (in_total or 0) - (out_total or 0)
        order_count = qs.filter(_order_movement_q()).count()

        # Category filter — same semantics as the global feed: order
        # movements are isolated by their reason/reference shapes and
        # excluded from the raw in/out/adjustment buckets.
        kind = (request.GET.get("kind") or "").strip()
        if kind == "order":
            qs = qs.filter(_order_movement_q())
        elif kind in ("in", "out", "adjustment"):
            qs = qs.filter(movement_type=kind).exclude(_order_movement_q())

        # Legacy raw type filter (kept for old bookmarked URLs).
        if types_param:
            wanted = [t.strip() for t in types_param.split(",") if t.strip()]
            qs = qs.filter(movement_type__in=wanted)

        qs = qs.order_by("-created_at")

        paginator = Paginator(qs, 50)
        try:
            page = paginator.page(request.GET.get("page", 1))
        except (EmptyPage, PageNotAnInteger, ValueError):
            page = paginator.page(1)

        # Product dropdown options
        products = WarehouseProduct.objects.filter(warehouse=warehouse).order_by("name").values("id", "name", "sku")

        # Preserve filter querystring for pagination links (without page key)
        qd = request.GET.copy()
        qd.pop("page", None)
        filter_qs = qd.urlencode()

        return render(request, self.template_name, {
            "warehouse": warehouse,
            "page_obj": page,
            "paginator": paginator,
            "movements": _decorate_movements(list(page.object_list)),
            "in_total": in_total,
            "out_total": out_total,
            "net": net,
            "order_count": order_count,
            "products": list(products),
            "filter_qs": filter_qs,
            "filters": {
                "type": types_param,
                "kind": kind,
                "from": date_from,
                "to": date_to,
                "product": product_pk,
                "q": q,
                "min_qty": min_qty,
                "max_qty": max_qty,
            },
        })


@method_decorator(login_required, name='dispatch')
class WarehouseProductEdit(View):
    """POST endpoint that updates editable WarehouseProduct fields.
    On price change, also recomputes cost_usd / cost_try so the
    warehouse total-value rollup is accurate. Emits an "adjustment"
    movement only when quantity is manually overridden so the
    timeline reflects the change."""

    def post(self, request, warehouse_pk, product_pk):
        warehouse = get_object_or_404(Warehouse, pk=warehouse_pk)
        product = get_object_or_404(WarehouseProduct, pk=product_pk, warehouse=warehouse)

        name = (request.POST.get("name") or "").strip()
        sku = (request.POST.get("sku") or "").strip()
        barcode = (request.POST.get("barcode") or "").strip()
        model = (request.POST.get("model") or "").strip()
        purchase_price_raw = (request.POST.get("purchase_price") or "").strip().replace(",", ".")
        purchase_currency = (request.POST.get("purchase_currency") or "USD").strip().upper()
        quantity_raw = (request.POST.get("quantity") or "").strip().replace(",", ".")
        # Optional: explicitly move this variant under a different catalog
        # main product (the base title typed/picked in the edit modal).
        catalog_base = (request.POST.get("catalog_base_name") or "").strip()

        changes = []
        update_fields = ["updated_at"]

        if name and name != product.name:
            product.name = name; update_fields.append("name"); changes.append(f"name → {name}")
        # SKU / barcode can be cleared by sending empty string.
        if (sku or "") != (product.sku or ""):
            product.sku = sku or None; update_fields.append("sku"); changes.append("sku updated")
        if (barcode or "") != (product.barcode or ""):
            product.barcode = barcode or None; update_fields.append("barcode"); changes.append("barcode updated")
        if (model or "") != (product.model or ""):
            product.model = model or None; update_fields.append("model"); changes.append("model updated")

        # Purchase price + currency — recompute cost_usd / cost_try so
        # warehouse value rollup stops showing $0 for products that
        # only had purchase_price set previously.
        if purchase_price_raw:
            try:
                pp = Decimal(purchase_price_raw)
                # The edit modal always re-posts the prefilled price, so
                # only record a change when the value/currency actually
                # differ — else every no-op save spams the activity feed.
                price_changed = (
                    product.purchase_price is None
                    or pp != product.purchase_price
                    or purchase_currency != (product.purchase_currency or "").upper()
                )
                product.purchase_price = pp
                product.purchase_currency = purchase_currency
                update_fields.extend(["purchase_price", "purchase_currency"])
                usd_try = _get_usd_try_rate() or Decimal("1")
                if purchase_currency == "USD":
                    product.cost_usd = pp
                    product.cost_try = (pp * usd_try) if usd_try else None
                elif purchase_currency == "TRY":
                    product.cost_try = pp
                    if usd_try and usd_try > 0:
                        product.cost_usd = (pp / usd_try).quantize(Decimal("0.0001"))
                elif purchase_currency == "EUR":
                    product.cost_usd = pp
                update_fields.extend(["cost_usd", "cost_try"])
                if price_changed:
                    changes.append(f"price → {pp} {purchase_currency}")
            except (InvalidOperation, TypeError):
                pass

        # Manual quantity override — record as adjustment movement so
        # the audit trail explains the delta.
        if quantity_raw:
            try:
                new_qty = Decimal(quantity_raw)
                old_qty = product.quantity or Decimal("0")
                if new_qty != old_qty:
                    delta = new_qty - old_qty
                    product.quantity = new_qty
                    update_fields.append("quantity")
                    StockMovement.objects.create(
                        product=product,
                        roll=None,
                        movement_type="adjustment",
                        quantity=abs(delta),
                        reason=f"Manual adjustment: {old_qty}m → {new_qty}m",
                        reference="Product edit",
                        created_by=request.user if request.user.is_authenticated else None,
                    )
                    changes.append(f"qty {old_qty} → {new_qty}")
            except (InvalidOperation, TypeError):
                pass

        product.save(update_fields=list(set(update_fields)))

        # Log non-quantity edits (name/SKU/barcode/model/price) as a
        # zero-metre adjustment so the activity feed shows every product
        # change. Quantity overrides already logged their own movement.
        non_qty_changes = [c for c in changes if not c.startswith("qty ")]
        if non_qty_changes:
            try:
                StockMovement.objects.create(
                    product=product,
                    roll=None,
                    movement_type="adjustment",
                    quantity=Decimal("0"),
                    reason=("Product edited: " + "; ".join(non_qty_changes))[:255],
                    reference="Product edit",
                    created_by=request.user if request.user.is_authenticated else None,
                )
            except Exception:
                pass  # audit logging must never block the edit itself

        # Keep the hidden catalog in sync: if the SKU/name changed (or the
        # user re-pointed the main product), re-link the catalog variant to
        # the correct main product and rename variant_sku to match.
        catalog_warning = None
        catalog_info = None
        sku_changed = "sku updated" in changes
        name_changed = any(c.startswith("name →") for c in changes)
        if (product.catalog_variant_id or catalog_base) and (sku_changed or name_changed or catalog_base):
            try:
                from .catalog_sync import resync_warehouse_product
                variant, catalog_warning = resync_warehouse_product(
                    product, base_override=catalog_base or None)
                if variant:
                    catalog_info = {
                        "variant_id": variant.id,
                        "variant_sku": variant.variant_sku,
                        "product_id": variant.product_id,
                        "product_title": variant.product.title,
                        "product_sku": variant.product.sku,
                    }
            except Exception as exc:
                import traceback
                traceback.print_exc()
                catalog_warning = f"Katalog senkron hatası: {exc}"

        # Same SKU = same variant: if editing made this product's SKU match
        # OTHER products in the warehouse, merge them into this one so they stop
        # appearing as duplicate rows (tops + movements moved here, qty re-rolled).
        merged_dupes = 0
        if product.sku:
            product, merged_dupes = _merge_warehouse_dupes_by_sku(
                warehouse, product.sku, keep=product)

        return JsonResponse({
            "success": True,
            "changes": changes,
            "merged_duplicates": merged_dupes,
            "catalog": catalog_info,
            "catalog_warning": catalog_warning,
            "product": {
                "id": product.pk,
                "name": product.name,
                "sku": product.sku,
                "barcode": product.barcode,
                "quantity": float(product.quantity or 0),
                "purchase_price": float(product.purchase_price or 0),
                "purchase_currency": product.purchase_currency,
                "cost_usd": float(product.cost_usd or 0),
            },
        })


@method_decorator(login_required, name='dispatch')
class WarehouseProductDelete(View):
    """Delete a WarehouseProduct entirely. Cascades to rolls + movements
    via the model FK delete behaviour. Before deleting we drop a final
    "adjustment" movement on the warehouse-wide ledger? — no, since
    deleting the product also deletes its movements. The user must
    confirm via POST; the warehouse total updates automatically."""

    def post(self, request, warehouse_pk, product_pk):
        warehouse = get_object_or_404(Warehouse, pk=warehouse_pk)
        product = get_object_or_404(WarehouseProduct, pk=product_pk, warehouse=warehouse)
        if not _is_admin(request.user):
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Sadece admin ürün silebilir."}, status=403)
            messages.error(request, "Sadece admin ürün silebilir.")
            return redirect("operating:warehouse_product_detail", warehouse.pk, product.pk)
        name = product.name
        product.delete()  # cascades to rolls + movements
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({
                "success": True,
                "redirect_url": reverse("operating:warehouse_detail", args=[warehouse.pk]),
            })
        messages.success(request, f"Product '{name}' deleted")
        return redirect("operating:warehouse_detail", warehouse.pk)


@method_decorator(login_required, name='dispatch')
class WarehouseRollDelete(View):
    """Delete a specific roll. This is a CORRECTION, not a sale: we log an
    ADJUSTMENT movement (so the timeline still shows the roll was removed)
    and decrement the parent quantity — but it must NOT count as a stock
    OUT. Real stock-out happens later, when an order is completed."""

    def post(self, request, warehouse_pk, product_pk, roll_pk):
        warehouse = get_object_or_404(Warehouse, pk=warehouse_pk)
        product = get_object_or_404(WarehouseProduct, pk=product_pk, warehouse=warehouse)
        roll = get_object_or_404(WarehouseProductRoll, pk=roll_pk, product=product)

        # Don't delete a roll that's actively reserved for an order.
        from .models import OrderRollReservation
        if OrderRollReservation.objects.filter(roll=roll, consumed=False).exists():
            return JsonResponse({"success": False,
                "error": "Bu top aktif bir siparişte rezerve edilmiş — önce ilgili siparişin paketlemesinden kaldırın."}, status=409)

        reason = (request.POST.get("reason") or "").strip() or "Roll deleted"
        remaining = (
            roll.meters_remaining
            if roll.meters_remaining is not None
            else (roll.meters or Decimal("0"))
        )

        # Decrement parent quantity for whatever was still on the roll.
        if remaining and remaining > 0:
            product.quantity = max(
                Decimal("0"), (product.quantity or Decimal("0")) - remaining
            )
            product.save(update_fields=["quantity", "updated_at"])

        # ADJUSTMENT (correction) — NOT an "out", so it never inflates the
        # Stock OUT total. Keep `roll=None` because the FK would dangle
        # after the delete; embed the roll's identity in the reason.
        StockMovement.objects.create(
            product=product,
            roll=None,
            movement_type="adjustment",
            quantity=remaining or Decimal("0"),
            reason=f"{reason} (roll #{roll.pk}"
                   + (f" · {roll.barcode}" if roll.barcode else "")
                   + ")",
            reference=roll.barcode or f"roll#{roll.pk}",
            created_by=request.user if request.user.is_authenticated else None,
        )

        roll.delete()

        return JsonResponse({
            "success": True,
            "product": {
                "id": product.pk,
                "quantity": float(product.quantity or 0),
                "rolls_count": product.rolls.count(),
                "active_rolls_count": product.rolls.exclude(status="consumed").count(),
            },
        })


@method_decorator(login_required, name='dispatch')
class WarehouseRollBulkDelete(View):
    """Delete MANY rolls (tops) at once — same CORRECTION semantics as the
    single delete: each removed roll logs an ADJUSTMENT movement and decrements
    the parent quantity once, all in one transaction. Accepts `roll_ids` as a
    repeated form field or a JSON list."""

    def post(self, request, warehouse_pk, product_pk):
        from django.db import transaction
        warehouse = get_object_or_404(Warehouse, pk=warehouse_pk)
        product = get_object_or_404(WarehouseProduct, pk=product_pk, warehouse=warehouse)

        raw_ids = request.POST.getlist("roll_ids")
        # Only touch request.body for a JSON payload — reading it after the form
        # POST stream was parsed raises RawPostDataException.
        if not raw_ids and "application/json" in (request.content_type or ""):
            try:
                payload = json.loads((request.body or b"").decode("utf-8") or "{}")
                raw_ids = payload.get("roll_ids") or []
            except (ValueError, UnicodeDecodeError):
                raw_ids = []
        ids = []
        for x in raw_ids:
            try:
                ids.append(int(x))
            except (TypeError, ValueError):
                pass
        if not ids:
            return JsonResponse({"success": False, "error": "Silinecek top seçilmedi."}, status=400)

        reason = (request.POST.get("reason") or "").strip() or "Bulk roll delete"
        rolls = list(WarehouseProductRoll.objects.filter(pk__in=ids, product=product))
        if not rolls:
            return JsonResponse({"success": False, "error": "Seçilen toplar bulunamadı."}, status=404)

        # Don't delete rolls actively reserved for an order — their hold
        # would be silently lost. Skip them and report the count.
        from .models import OrderRollReservation
        reserved_ids = set(OrderRollReservation.objects
                           .filter(roll__in=rolls, consumed=False)
                           .values_list("roll_id", flat=True))
        skipped = len(reserved_ids)
        rolls = [r for r in rolls if r.id not in reserved_ids]
        if not rolls:
            return JsonResponse({"success": False, "skipped": skipped,
                "error": "Seçilen toplar aktif siparişlerde rezerve — silinemedi. Önce paketlemeden kaldırın."}, status=409)

        deleted = 0
        with transaction.atomic():
            qty = product.quantity or Decimal("0")
            for roll in rolls:
                remaining = (roll.meters_remaining if roll.meters_remaining is not None
                             else (roll.meters or Decimal("0")))
                if remaining and remaining > 0:
                    qty = max(Decimal("0"), qty - remaining)
                StockMovement.objects.create(
                    product=product, roll=None, movement_type="adjustment",
                    quantity=remaining or Decimal("0"),
                    reason=f"{reason} (roll #{roll.pk}"
                           + (f" · {roll.barcode}" if roll.barcode else "") + ")",
                    reference=roll.barcode or f"roll#{roll.pk}",
                    created_by=request.user if request.user.is_authenticated else None,
                )
                roll.delete()
                deleted += 1
            product.quantity = qty
            product.save(update_fields=["quantity", "updated_at"])

        return JsonResponse({
            "success": True, "deleted": deleted, "skipped": skipped,
            "product": {
                "id": product.pk,
                "quantity": float(product.quantity or 0),
                "rolls_count": product.rolls.count(),
                "active_rolls_count": product.rolls.exclude(status="consumed").count(),
            },
        })


@method_decorator(login_required, name='dispatch')
class WarehouseRollEdit(View):
    """Edit ONE roll: barcode, meters, lot_number. Changing meters
    re-rolls the parent product.quantity (recomputed from all rolls) and
    logs an adjustment; the linked catalog variant quantity follows."""

    def post(self, request, warehouse_pk, product_pk, roll_pk):
        from django.db.models import Sum, F
        from django.db.models.functions import Coalesce
        warehouse = get_object_or_404(Warehouse, pk=warehouse_pk)
        product = get_object_or_404(WarehouseProduct, pk=product_pk, warehouse=warehouse)
        roll = get_object_or_404(WarehouseProductRoll, pk=roll_pk, product=product)

        changes = []
        roll_fields = []

        # ── Barcode (unique per roll within the warehouse) ──
        barcode = (request.POST.get("barcode") or "").strip()
        if (barcode or "") != (roll.barcode or ""):
            if barcode and (WarehouseProductRoll.objects
                            .filter(product__warehouse=warehouse, barcode=barcode)
                            .exclude(pk=roll.pk).exists()):
                return JsonResponse({"success": False,
                                     "error": "Bu barkod zaten başka bir topta kullanılıyor."},
                                    status=400)
            roll.barcode = barcode or None
            roll_fields.append("barcode"); changes.append("barcode")

        # ── Lot number ──
        lot = (request.POST.get("lot_number") or "").strip()
        if (lot or "") != (roll.lot_number or ""):
            roll.lot_number = lot or None
            roll_fields.append("lot_number"); changes.append("lot")

        # ── Meters (full length); keep any already-consumed amount ──
        meters_raw = (request.POST.get("meters") or "").strip().replace(",", ".")
        meters_changed = False
        if meters_raw:
            try:
                new_full = Decimal(meters_raw)
            except (InvalidOperation, TypeError):
                return JsonResponse({"success": False, "error": "Geçersiz metre değeri."}, status=400)
            if new_full <= 0:
                return JsonResponse({"success": False, "error": "Metre pozitif olmalı."}, status=400)
            old_full = roll.meters or Decimal("0")
            consumed = Decimal("0")
            if roll.meters_remaining is not None and old_full:
                consumed = max(Decimal("0"), old_full - roll.meters_remaining)
            if new_full != old_full:
                roll.meters = new_full
                roll.meters_remaining = max(Decimal("0"), new_full - consumed)
                roll_fields.extend(["meters", "meters_remaining"])
                meters_changed = True
                changes.append("meters")

        if roll_fields:
            roll.save(update_fields=list(set(roll_fields)))

        # Recompute the parent quantity authoritatively from its rolls
        # (current stock = remaining meters, falling back to full meters).
        if meters_changed:
            total = product.rolls.aggregate(
                s=Coalesce(Sum(Coalesce(F("meters_remaining"), F("meters"))),
                           Decimal("0"),
                           output_field=DecimalField(max_digits=18, decimal_places=2)))["s"]
            old_qty = product.quantity or Decimal("0")
            if total != old_qty:
                product.quantity = total
                product.save(update_fields=["quantity", "updated_at"])
                StockMovement.objects.create(
                    product=product, roll=roll, movement_type="adjustment",
                    quantity=abs(total - old_qty),
                    reason=f"Roll meters edited ({old_qty}m → {total}m)",
                    reference=roll.barcode or f"Roll #{roll.pk}",
                    created_by=request.user if request.user.is_authenticated else None,
                )
            # Mirror onto the linked catalog variant.
            if product.catalog_variant_id:
                v = product.catalog_variant
                v.variant_quantity = product.quantity
                v.save(update_fields=["variant_quantity"])

        return JsonResponse({
            "success": True,
            "changes": changes,
            "roll": {
                "id": roll.pk,
                "barcode": roll.barcode,
                "meters": float(roll.meters or 0),
                "meters_remaining": float(roll.meters_remaining or 0) if roll.meters_remaining is not None else None,
                "lot_number": roll.lot_number,
            },
            "product_quantity": float(product.quantity or 0),
        })


@method_decorator(login_required, name='dispatch')
class WarehouseStockOut(View):
    """POST endpoint to record manual stock-out from a product. Body:
        amount: decimal meters
        roll_id (optional): consume from a specific roll
        reason (optional)
        reference (optional)
    Returns JSON with the updated totals."""

    def post(self, request, warehouse_pk, product_pk):
        warehouse = get_object_or_404(Warehouse, pk=warehouse_pk)
        product = get_object_or_404(
            WarehouseProduct, pk=product_pk, warehouse=warehouse,
        )
        amount_raw = (request.POST.get("amount") or "").strip().replace(",", ".")
        try:
            amount = Decimal(amount_raw)
        except (InvalidOperation, TypeError):
            return JsonResponse({"success": False, "error": "Invalid amount"}, status=400)
        if amount <= 0:
            return JsonResponse({"success": False, "error": "Amount must be positive"}, status=400)
        if amount > (product.quantity or Decimal("0")):
            return JsonResponse({
                "success": False,
                "error": f"Only {product.quantity}m available",
            }, status=400)

        roll = None
        roll_id = request.POST.get("roll_id")
        if roll_id:
            roll = product.rolls.filter(pk=roll_id).first()
            if not roll:
                return JsonResponse({"success": False, "error": "Roll not found"}, status=404)
            roll_rem = roll.meters_remaining if roll.meters_remaining is not None else roll.meters
            if amount > (roll_rem or Decimal("0")):
                return JsonResponse({
                    "success": False,
                    "error": f"Only {roll_rem}m left on this roll",
                }, status=400)
            roll.meters_remaining = (roll_rem or Decimal("0")) - amount
            # Update status based on remaining.
            if roll.meters_remaining <= Decimal("0"):
                roll.status = "consumed"
            elif roll.meters_remaining < (roll.meters or Decimal("0")):
                roll.status = "partial"
            roll.save(update_fields=["meters_remaining", "status"])

        # Drop the parent quantity.
        product.quantity = (product.quantity or Decimal("0")) - amount
        product.save(update_fields=["quantity", "updated_at"])

        StockMovement.objects.create(
            product=product,
            roll=roll,
            movement_type="out",
            quantity=amount,
            reason=(request.POST.get("reason") or "Manual stock-out").strip() or "Manual stock-out",
            reference=(request.POST.get("reference") or "").strip() or None,
            created_by=request.user if request.user.is_authenticated else None,
        )

        return JsonResponse({
            "success": True,
            "product": {
                "id": product.pk,
                "quantity": float(product.quantity or 0),
                "rolls_count": product.rolls.count(),
                "active_rolls_count": product.rolls.exclude(status="consumed").count(),
            },
        })


@method_decorator(login_required, name='dispatch')
class WarehouseRollScan(View):
    """POST endpoint hit by the camera scanner.

    Expects multipart form-data with:
      - image: the captured frame (JPEG/PNG)
      - sku, name, meters (optional overrides if user has already confirmed in UI)
      - commit: 'true' to actually save; otherwise just OCR + return.

    Two-phase flow so the UI can confirm OCR'd values before writing.
    """

    def post(self, request, pk):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        image = request.FILES.get("image")
        commit = (request.POST.get("commit") or "").lower() == "true"

        # ── Phase 1: pure OCR (called with commit=false) ────────
        if not commit:
            if not image:
                return JsonResponse({"success": False, "error": "No image uploaded"}, status=400)
            raw, parsed = _ocr_label(image)
            if "error" in parsed:
                return JsonResponse({
                    "success": False,
                    "error": parsed["error"],
                    "raw": raw or "",
                    "attempts": parsed.get("_attempts", []),
                    "backend": "none",
                    "parsed": {},
                })

            # Trust the OCR'd SKU as printed — vision LLMs (xAI/Gemini/
            # Claude) read fabric-label SKUs very reliably, so the old
            # Levenshtein "auto-correct to nearest existing SKU" behaviour
            # ended up *changing* correct OCR output to a similar-but-
            # wrong DB entry (e.g. KZL000691 → KZL000131). Always show
            # the user exactly what was on the label.
            original_sku = parsed.get("sku")

            # Look up the existing product, if any. Exact match first;
            # only fall back to fuzzy when there is NO exact match — and
            # even then, we don't rewrite parsed["sku"]. Fuzzy just
            # surfaces a candidate so the UI can flag "looks similar to
            # SKU KZL000131 — same product?" without forcing it.
            match = None
            if parsed.get("sku"):
                match = WarehouseProduct.objects.filter(
                    warehouse=warehouse, sku__iexact=parsed["sku"],
                ).first()
            # Otherwise try by name as a last-resort heuristic.
            if not match and parsed.get("name"):
                match = WarehouseProduct.objects.filter(
                    warehouse=warehouse,
                    name__icontains=parsed["name"][:40],
                ).first()

            if match:
                print(
                    f"[OCR] matched existing product: #{match.pk} "
                    f"sku={match.sku!r} name={match.name!r} qty={match.quantity}",
                    flush=True,
                )
            else:
                print("[OCR] no existing product matched — UI will offer to create one", flush=True)

            backend = parsed.get("_backend") or "unknown"
            attempts = parsed.get("_attempts") or [{"backend": backend, "error": None}]
            print(f"[OCR] ── done ── final backend={backend}\n", flush=True)

            # Catalog preview: parse the scanned name into base product +
            # variant (color/model) and tell the UI whether that base
            # product already exists in the marketing catalog.
            from .catalog_sync import derive_catalog
            from marketing.models import Product as _CatProduct, ProductVariant as _CatVariant
            cp = derive_catalog(parsed.get("sku"), parsed.get("name"), parsed.get("color"))
            base_name = cp["base_name"]
            attribute_name = cp["attribute_name"]
            attribute_value = cp["attribute_value"]

            # If this EXACT variant_sku already exists in the catalog, this is
            # the SAME variant being re-scanned (a new roll/barcode of it) —
            # reuse the main product + colour/model from the FIRST time so the
            # user doesn't have to refill them (they can still edit). Only the
            # barcode differs roll-to-roll.
            variant_exists = False
            scanned_sku = (parsed.get("sku") or "").strip()
            try:
                if scanned_sku:
                    ev = (_CatVariant.objects.select_related("product")
                          .filter(variant_sku=scanned_sku).first())
                    if ev:
                        variant_exists = True
                        base_name = ev.product.title or base_name
                        av = (ev.product_variant_attribute_values
                              .select_related("product_variant_attribute")
                              .order_by("-id").first())
                        if av:
                            attribute_name = av.product_variant_attribute.name
                            # de-normalize for display ("dark_cream" → "DARK CREAM")
                            attribute_value = av.product_variant_attribute_value.replace("_", " ").upper()

                base_exists = bool(
                    base_name and _CatProduct.objects.filter(title__iexact=base_name).exists()
                )
            except Exception as exc:
                # A catalog hiccup must never break the OCR preview.
                print(f"[OCR] catalog_preview lookup failed: {exc}", flush=True)
                base_exists = False
            catalog_preview = {
                "base_name": base_name,
                "base_exists": base_exists,
                "attribute_name": attribute_name,
                "attribute_value": attribute_value,
                "original_token": cp["original_token"],
                "is_color": cp["is_color"],
                "variant_exists": variant_exists,
            }

            return JsonResponse({
                "success": True,
                "parsed": {k: v for k, v in parsed.items() if not k.startswith("_")},
                "catalog_preview": catalog_preview,
                "backend": backend,
                "attempts": attempts,  # full fallback chain for debug panel
                "raw": (raw or "")[:4000],  # OCR backend's raw response — for the debug panel
                "original_sku": original_sku,  # before fuzzy correction
                "sku_was_corrected": bool(
                    original_sku and parsed.get("sku")
                    and original_sku.upper() != (parsed.get("sku") or "").upper()
                ),
                "matched_product": (
                    {
                        "id": match.pk,
                        "name": match.name,
                        "sku": match.sku,
                        "quantity": float(match.quantity or 0),
                        "rolls_count": match.rolls.count(),
                    } if match else None
                ),
            })

        # ── Phase 2: commit — write the roll to DB ──────────────
        sku = (request.POST.get("sku") or "").strip()
        name = (request.POST.get("name") or "").strip()
        meters_raw = (request.POST.get("meters") or "").strip().replace(",", ".")
        try:
            meters = Decimal(meters_raw)
        except (InvalidOperation, TypeError):
            return JsonResponse({"success": False, "error": "Invalid meters value"}, status=400)
        if meters <= 0:
            return JsonResponse({"success": False, "error": "Meters must be positive"}, status=400)
        if not sku and not name:
            return JsonResponse({"success": False, "error": "SKU or name is required"}, status=400)

        # Reject a re-scan of the SAME roll: each physical roll's barcode is
        # unique, so if one already exists in this warehouse it's a duplicate
        # scan — don't add it again. (Same SKU with a DIFFERENT barcode is a
        # different roll of the same product and is allowed.)
        dup_barcode = (request.POST.get("barcode") or "").strip()
        if dup_barcode:
            existing_roll = (WarehouseProductRoll.objects
                             .filter(product__warehouse=warehouse, barcode=dup_barcode)
                             .select_related("product").first())
            if existing_roll:
                return JsonResponse({
                    "success": False,
                    "duplicate": True,
                    "error": "Bu barkod zaten okundu — tekrar eklenmedi (%s)." % (
                        existing_roll.product.name or existing_roll.product.sku or "?"),
                })

        # Find or create the WarehouseProduct in this warehouse.
        #
        # Match by SKU ONLY when a SKU is present. Physically different
        # products often share the same descriptive NAME — "GREK TÜL" is
        # printed on K24620İ.G52, K24620İ.G47, K24892İ.G157, K24620.G33 …
        # They are DIFFERENT products, distinguished by their SKU. The old
        # name fallback merged every "GREK TÜL" roll onto the first one
        # found (e.g. K24620.G33), so new rolls were saved under the wrong
        # product. Only use the name as the key when the label has NO SKU.
        product = None
        if sku:
            product = WarehouseProduct.objects.filter(
                warehouse=warehouse, sku__iexact=sku,
            ).first()
        elif name:
            product = WarehouseProduct.objects.filter(
                warehouse=warehouse, name__iexact=name,
            ).first()
        if not product:
            # No existing product for this SKU (or nameless label) → new one.
            product = WarehouseProduct.objects.create(
                warehouse=warehouse,
                name=name or sku,
                sku=sku or None,
                quantity=Decimal("0"),
            )

        # Optional manual fields: barcode (per-roll) + price (per-product).
        barcode = (request.POST.get("barcode") or "").strip() or None
        purchase_price_raw = (request.POST.get("purchase_price") or "").strip().replace(",", ".")
        purchase_currency = (request.POST.get("purchase_currency") or "").strip().upper() or product.purchase_currency or "USD"
        if purchase_price_raw:
            try:
                purchase_price = Decimal(purchase_price_raw)
                product.purchase_price = purchase_price
                product.purchase_currency = purchase_currency
                # ALSO populate cost_usd / cost_try so the warehouse
                # value rollup actually reflects the price (the
                # rollup multiplies quantity * cost_usd; if only
                # purchase_price is set the rollup stays $0).
                usd_try = _get_usd_try_rate() or Decimal("1")
                if purchase_currency == "USD":
                    product.cost_usd = purchase_price
                    product.cost_try = (purchase_price * usd_try) if usd_try else None
                elif purchase_currency == "TRY":
                    product.cost_try = purchase_price
                    if usd_try and usd_try > 0:
                        product.cost_usd = (purchase_price / usd_try).quantize(Decimal("0.0001"))
                elif purchase_currency == "EUR":
                    # Treat as a USD-equivalent for the value rollup —
                    # not perfect but better than $0. The accounting
                    # service can be extended later for true EUR.
                    product.cost_usd = purchase_price
            except (InvalidOperation, TypeError):
                pass
        # Update product-level barcode if it didn't have one (per-roll
        # barcode is the source of truth for stock-out scanning, but
        # the product-level field stays handy for product search).
        if barcode and not product.barcode:
            product.barcode = barcode[:64]

        # Save the roll.
        roll = WarehouseProductRoll.objects.create(
            product=product,
            meters=meters,
            meters_remaining=meters,
            barcode=barcode[:64] if barcode else None,
            lot_number=(request.POST.get("lot_number") or "").strip() or None,
            scanned_by=request.user if request.user.is_authenticated else None,
            source_image=image if image else None,
            ocr_raw=(request.POST.get("ocr_raw") or "")[:5000] or None,
        )

        # Bump the parent quantity (denormalised cache).
        product.quantity = (product.quantity or Decimal("0")) + meters
        product.save(update_fields=[
            "quantity", "updated_at", "barcode",
            "purchase_price", "purchase_currency",
            "cost_usd", "cost_try",
        ])

        # Stock-in ledger entry — drives the product detail timeline.
        StockMovement.objects.create(
            product=product,
            roll=roll,
            movement_type="in",
            quantity=meters,
            reason="Roll scanned",
            reference=barcode or roll.lot_number or "",
            created_by=request.user if request.user.is_authenticated else None,
        )

        # ── Mirror into the hidden marketing catalog as a variant of a
        #    main product — ONLY if the user left the "Also add to catalog"
        #    box ticked (default ON). Unchecked → warehouse-only.
        #    Never let a catalog hiccup roll back the stock write. ──
        catalog_info = None
        catalog_warning = None
        do_catalog = (request.POST.get("catalog_sync", "1") or "1").strip().lower() not in ("0", "false", "no", "off")
        variant_sku = (request.POST.get("catalog_variant_sku") or sku or "").strip()
        if not do_catalog:
            variant_sku = ""   # skip catalog sync entirely
        if variant_sku and len(variant_sku) > 20:
            catalog_warning = f"SKU '{variant_sku}' 20 karakteri aşıyor — katalog varyantı oluşturulmadı."
        elif variant_sku:
            try:
                from .catalog_sync import (
                    derive_catalog, sync_roll_to_catalog, CatalogSyncConflict,
                )
                from marketing.models import Product as _Prod
                cp = derive_catalog(variant_sku, name, request.POST.get("catalog_color"))
                base_name = (request.POST.get("catalog_base_name") or cp["base_name"] or name).strip()
                attribute_name = (request.POST.get("catalog_attribute") or cp["attribute_name"] or "").strip() or None
                attribute_value = (request.POST.get("catalog_value") or cp["attribute_value"] or "").strip() or None

                existing_base = None
                manual_pid = (request.POST.get("catalog_base_product_id") or "").strip()
                if product.catalog_variant_id and product.catalog_variant and product.catalog_variant.product_id:
                    existing_base = product.catalog_variant.product
                elif manual_pid.isdigit():
                    existing_base = _Prod.objects.filter(pk=int(manual_pid)).first()

                cat_product, cat_variant, p_created, v_created = sync_roll_to_catalog(
                    base_name=base_name,
                    attribute_name=attribute_name,
                    attribute_value=attribute_value,
                    variant_sku=variant_sku,
                    variant_barcode=barcode,
                    quantity=product.quantity,   # mirror this variant's warehouse stock
                    cost=product.cost_usd,
                    existing_base_product=existing_base,
                )
                if not product.catalog_variant_id:
                    product.catalog_variant = cat_variant
                    product.save(update_fields=["catalog_variant"])
                catalog_info = {
                    "product_id": cat_product.id,
                    "product_title": cat_product.title,
                    "product_created": p_created,
                    "variant_id": cat_variant.id,
                    "variant_sku": cat_variant.variant_sku,
                    "variant_created": v_created,
                    "attribute_name": attribute_name,
                    "attribute_value": attribute_value,
                }
            except CatalogSyncConflict as exc:
                catalog_warning = str(exc)
            except Exception as exc:
                import traceback
                traceback.print_exc()
                catalog_warning = f"Katalog senkron hatası: {exc}"

        return JsonResponse({
            "success": True,
            "roll": {
                "id": roll.pk,
                "meters": float(roll.meters),
                "barcode": roll.barcode,
                "lot_number": roll.lot_number,
                "scanned_at": roll.scanned_at.strftime("%H:%M:%S"),
            },
            "product": {
                "id": product.pk,
                "name": product.name,
                "sku": product.sku,
                "barcode": product.barcode,
                "quantity": float(product.quantity or 0),
                "rolls_count": product.rolls.count(),
                "main_product": (catalog_info["product_title"] if catalog_info else None),
            },
            "catalog": catalog_info,
            "catalog_warning": catalog_warning,
        })
